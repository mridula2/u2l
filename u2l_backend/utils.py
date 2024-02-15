from flask import Blueprint, jsonify, request, Response, Flask
import re
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from models import db, user_details, project_details, os_details, analysis_type, analysis_java, analysis_c, analysis_shell, analysis_status, source_code_inventory, migration_summary, artefacts_summary, java_data, contact_us, analysis_summary_java, celery_job_details, MiddlewareComponent
from datetime import datetime
import smtplib
import logging
from celery_setup import celery
from celery.utils.log import get_task_logger
from celery import Task
import subprocess
import time
import datetime
import os
import subprocess
import shutil   
import pandas as pd
from openpyxl import load_workbook
import re
import numpy as np
import zipfile
import openpyxl
import json

from flask_sse import sse
from celery import Celery
from threading import Lock

from bs4 import BeautifulSoup
import requests
import pandas as pd
import os
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# app = Flask(__name__)
logger = get_task_logger(__name__)

import redis

redis_client = redis.StrictRedis(host='localhost', port=6379, db=0)

def validate_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    logger.info("TASK HAS STARTED EXECUTION : Info")
    logger.warning("TASK HAS STARTED EXECUTION : Warning")
    logger.error("TASK HAS STARTED EXECUTION : Error")
    return re.match(pattern, email) is not None

def generate_workbook(df, save_location,sheet):
    book = load_workbook(save_location)
    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheet, index=True)
    writer.save()

def sheet_transfer(source_sheet, target_sheet, startrow, startcol, source_file, dest_file):
    df = pd.read_excel(source_file, sheet_name=source_sheet, engine='openpyxl')
    book = load_workbook(dest_file)
    writer = pd.ExcelWriter(dest_file, engine='openpyxl', mode='a')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=target_sheet, index=False, startrow=startrow, startcol=startcol, header=False)
    writer.save()

def insert_user_details_task(email, password, first_name, last_name, user_role):
    # user = user_details.query.filter_by(email=email).first()
    # if user:
    #     return
    # current_time = datetime.utcnow()
    # new_user = user_details(email, password, first_name, last_name, user_role, current_time, current_time)
    # db.session.add(new_user)
    # db.session.commit()

    # Your task implementation here...
    logger.info(f"Inserting user details for email: {email}")

    user = user_details.query.filter_by(email=email).first()
    if user:
        logger.warning("User already exists.")
        return

    current_time = datetime.datetime.now()
    new_user = user_details(email, password, first_name, last_name, user_role, current_time, current_time)
    db.session.add(new_user)
    db.session.commit()

    logger.info("User details inserted successfully.")

def fetch_store_java_data(form_project_name, form_application_name, filepath):

    sheet_names = ['OS Analysis Details', 'Compilation Error Report', 'Import Class Report', 'JDK Anaylsis Details', 'Import JSP Report', 'Compilation Warning Report']
    wb = load_workbook(filename=filepath, read_only=True, data_only=True)

    no_of_artefacts = []
    count_of_ids = 0
    total_count_of_ids = 0
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        count_of_ids = 0

        if sheet.max_row < 2:
            continue

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                count_of_ids += 1
                total_count_of_ids +=1

        no_of_artefacts.append({sheet_name: count_of_ids})
    print(no_of_artefacts)    

    sheetName = 'Source Code Inventory'
    selected_cols = ['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    df = df.apply(pd.to_numeric, errors='coerce')
    column_sums = df.sum()

    sheetName = 'Source Code Inventory'
    selected_cols = ['Type']
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    type_counts = df['Type'].value_counts().reset_index()
    counts = type_counts.rename(columns={'index': 'Type', 'Type': 'number'})
    counts_list = counts.to_dict('records')

    total_no_of_lines = column_sums['Total Nr LoC']
    percentage = (total_count_of_ids/total_no_of_lines) * 100
    rounded_percentage = round(percentage, 2)

    # Store each value into the database
    data_entry = java_data(
        inventory_sums_actual_nr_of_lines=column_sums['Actual Nr of Lines'],
        project_name = form_project_name,
        application_name = form_application_name,
        inventory_sums_nr_blank_lines=column_sums['Nr Blank Lines'],
        inventory_sums_nr_commented_lines=column_sums['Nr Commented Lines'],
        inventory_sums_total_nr_loc=column_sums['Total Nr LoC'],
        no_of_artefacts_os_analysis_details=no_of_artefacts[0]['OS Analysis Details'],
        no_of_artefacts_compilation_error_report=no_of_artefacts[1]['Compilation Error Report'],
        no_of_artefacts_import_class_report=no_of_artefacts[2]['Import Class Report'],
        no_of_artefacts_jdk_analysis_details=no_of_artefacts[3]['JDK Anaylsis Details'],
        no_of_artefacts_import_jsp_report=no_of_artefacts[4]['Import JSP Report'],
        no_of_artefacts_compilation_warning_report=no_of_artefacts[5]['Compilation Warning Report'],
        artefacts_count_type=counts_list[0]['Type'],
        artefacts_count_number=counts_list[0]['number'],
        percent=rounded_percentage
    )

    db.session.add(data_entry)
    db.session.commit()

def send_email(user_info):
    email_subject = 'New request for HPE Code Assessment Suite!'
    email_body = 'There has been a request for HPE Code Assessment Suite!\n\n'
    
    for key, value in user_info.items():
        email_body += f'{key.capitalize()}: {value}\n'
    
    # Connect to the email server
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()

    server.login('pratikjwani@gmail.com', 'zakxdvbnwrpjoxfb')

    email_message = f'Subject: {email_subject}\n\n{email_body}'

    # pratik-jayawant.wani@hpe.com
    # joy.pahari@hpe.com
    # mridula.krishnamurthy@hpe.com

    server.sendmail('pratikjwani@gmail.com', 'pratik-jayawant.wani@hpe.com', email_message)

    server.quit()

def summary_sheet_generation(source_excel_file, sheet_name_1, sheet_name_2, skiprows, column_index, target_excel_file, start_row, start_col):

    sheet_configs = {
        sheet_name_1: {'usecols': lambda x: x != 1, 'skiprows': skiprows, 'column_index': column_index}
    }

    for sheet_name_1, config in sheet_configs.items():
        df = pd.read_excel(source_excel_file, engine='openpyxl', sheet_name=sheet_name_1, usecols=config['usecols'], skiprows=config['skiprows'])
        df = df.drop(df.columns[0], axis=1)
        df = df.iloc[:, config['column_index']]

    row_counts = df.value_counts()
    unique_rows_df = pd.DataFrame({'Row': row_counts.index, 'Occurrences': row_counts.values})

    writer = pd.ExcelWriter(target_excel_file, engine='openpyxl', mode='a')
    book = load_workbook(target_excel_file)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    unique_rows_df.to_excel(writer, sheet_name=sheet_name_2, index=False, startrow=start_row, startcol=start_col, header=False)
    writer.save()

def get_task_queue_name(task_id):
    return f'logs:{task_id}'

@celery.task
def testing_task(yMD, hMS, tool_analysis_type, framework_type, file_name, script_path, file_path, form_project_name, form_application_name, form_project_client, form_project_manager, form_source_os, form_source_os_version, form_target_os, form_target_os_version, form_source_jdk, form_target_jdk, form_source_jsp, form_target_jsp, form_source_servlet, form_target_servlet, form_source_compiler, form_target_compiler, form_source_compiler_version, form_target_compiler_version, form_source_pre_compiler, form_target_pre_compiler, form_source_pre_compiler_version, form_target_pre_compiler_version, form_source_shell, form_target_shell, form_source_shell_version, form_target_shell_version):

    task_id = testing_task.request.id
    redis_queue_name = get_task_queue_name(task_id)
    
    logger.info("TASK HAS STARTED EXECUTION")
    logger.warning("TASK HAS STARTED EXECUTION")
    logger.error("TASK HAS STARTED EXECUTION")
    redis_client.rpush(redis_queue_name, '1/40::CODE ASSESSMENT STARTED')
    
    # Analysis endpoint logic starts here
    # Execute U2LTool_Install.sh Script
    try:
        command = ['sh', script_path, file_path, tool_analysis_type]
        process = subprocess.Popen(command, stdout=subprocess.PIPE)
        output = process.communicate()[0].decode().strip()
        lines = output.split('\n')
        logger.info("TOOL INSTALLATION COMPLETED")
        redis_client.rpush(redis_queue_name, '2/40::Tool Installation Completed.')
    except Exception as e:
        return jsonify({'error': str(e)})

    global PJHOME
    PJHOME = lines[5].split(':')[1]
    global APNAME
    APNAME = lines[6].split(':')[1]
    
    # Export 'PJHOME' & 'APNAME' variables
    os.environ['PJHOME'] = PJHOME
    os.environ['APNAME'] = APNAME

    source_path = '/usr/u2l/u2l_backend/projects/' + file_name +'/'+ file_name + '.zip'
    try:
        if(tool_analysis_type == 'javaanalysis'):
            destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
        elif(tool_analysis_type == 'canalysis'):
            destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/'
        elif(tool_analysis_type == 'shellanalysis'):
            destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_'+ yMD + '_' + hMS +'/work/shellanalysis/'
        shutil.unpack_archive(source_path, destination_path, 'zip')
        
    except Exception as e:
        response = jsonify({'error': str(e)})
        response.status_code = 500
        return response
    
    try:
        if(tool_analysis_type == 'canalysis' or tool_analysis_type == 'shellanalysis'):
            script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/U2LTool_Analysis.sh'

            redis_client.rpush(redis_queue_name, '3/40::Executing U2LTool_Analysis.sh script')
            command = ['sh', script_path, PJHOME, APNAME]
            process = subprocess.Popen(command, stdout=subprocess.PIPE)
            process.wait()
            output = process.communicate()[0].decode().strip()
            print('\nCompleted running U2LTool_Analysis.sh')
            redis_client.rpush(redis_queue_name, '4/40::Completed running U2LTool_Analysis.sh')
    
    except Exception as e:
        return jsonify({'error': str(e)})
    
    if(tool_analysis_type == 'javaanalysis'):
        try:

            script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/U2LTool_Analysis.sh'

            redis_client.rpush(redis_queue_name, '5/40::Executing U2LTool_Analysis.sh script')
            command = ['sh', script_path, PJHOME, APNAME]
            process = subprocess.Popen(command, stdout=subprocess.PIPE)
            process.wait()
            output = process.communicate()[0].decode().strip()

            logger.info('COMPLETED RUNNING U2LTool_Analysis.sh')
            redis_client.rpush(redis_queue_name, '6/40::Completed executing U2LTool_Analysis.sh')
        
        except Exception as e:
            return jsonify({'error': str(e)})

        # To run the javaanalysis scripts
        # Export 'PJHOME' variable
        PJHOME = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java'
        os.environ['PJHOME'] = PJHOME

        # Run javaSourceDiscovery.sh script
        try:
            redis_client.rpush(redis_queue_name, '7/40::Executing javaSourceDiscovery.sh script')
            script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaSourceDiscovery.sh'
            code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
            command = ['sh', script_path, code_path]
            process = subprocess.Popen(command, stdout=subprocess.PIPE)
            output = process.communicate()[0].decode().strip()
            lines = output.split('\n')
            logger.info('COMPLETED RUNNING javaSourceDiscovery.sh')
            redis_client.rpush(redis_queue_name, '8/40::Completed executing javaSourceDiscovery.sh')
        except Exception as e:
            return jsonify({'error': str(e)})    

        # Run javaRulesScan.sh script
        try:
            redis_client.rpush(redis_queue_name, '9/40::Executing javaRulesScan.sh script')
            script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaRulesScan.sh'
            code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
            command = ['sh', script_path, code_path]
            process = subprocess.Popen(command, stdout=subprocess.PIPE)
            output = process.communicate()[0].decode().strip()
            lines = output.split('\n')
            logger.info('COMPLETED RUNNING javaRulesScan.sh')
            redis_client.rpush(redis_queue_name, '10/40::Completed running javaRulesScan.sh')
        except Exception as e:
            return jsonify({'error': str(e)})   

        # Run javaRulesRemedy.sh script
        try:
            redis_client.rpush(redis_queue_name, '11/40::Executing javaRulesRemedy.sh script')
            script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaRulesRemedy.sh'
            code_path = '/usr/u2l/u2l_backend/projects/' + file_name + 'U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
            command = subprocess.Popen(['bash', script_path, code_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output = command.communicate()[0].decode().strip()
            lines = output.split('\n')
            logger.info('COMPLETED RUNNING javaRulesRemedy.sh')
            redis_client.rpush(redis_queue_name, '12/40::Completed running javaRulesRemedy.sh')
        except Exception as e:
            return jsonify({'error': str(e)})
        
        # Run javaFrameworkScan.sh script
        if (framework_type != ''):
            try:
                redis_client.rpush(redis_queue_name, '13/40::Executing javaFrameworkScan.sh script')
                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaFrameworkScan.sh'
                code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
                if(framework_type == 'JSF'):
                    framework_type = 'Jsf'
                command = subprocess.Popen(['bash', script_path, code_path, framework_type], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                output = command.communicate()[0].decode().strip()
                lines = output.split('\n')
                logger.info('COMPLETED RUNNING javaFrameworkScan.sh')
                redis_client.rpush(redis_queue_name, '14/40::Completed executing javaFrameworkScan.sh')
            except Exception as e:
                return jsonify({'error': str(e)})

        logger.info("JAVA SCRIPTS EXECUTED SUCCESSFULLY")
        redis_client.rpush(redis_queue_name, '15/40::Java Scripts Executed Successfully')

        try:

            logger.info("MOVING TEXT FILES")
            redis_client.rpush(redis_queue_name, '16/40::Transferring Logs files')
            source_dir = '/usr/u2l/u2l_backend/'
            destination_dir = '/usr/u2l/u2l_backend/projects/' + file_name + '/'

            if not os.path.exists(source_dir):
                return 'Source directory does not exist', 400
            if not os.path.exists(destination_dir):
                return 'Destination directory does not exist', 400

            files_to_transfer = os.listdir(source_dir)

            for filename in files_to_transfer:
                if os.path.isfile(os.path.join(source_dir, filename)):
                    file_extension = os.path.splitext(filename)[1]
                    if file_extension == '.txt' or not file_extension:
                        try:
                            os.rename(os.path.join(source_dir, filename), os.path.join(destination_dir, filename))
                        except FileNotFoundError:
                            continue
        except Exception as e:
            return jsonify({'error': str(e)})

        extracted_file_name = file_name.split('_')[0]

        redis_client.rpush(redis_queue_name, '17/40::Logs Transferred Successfully')
        
        #Performance
        redis_client.rpush(redis_queue_name, '18/40::Executing pmd script : Performance')
        run_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/bin/run.sh'
        pmd = 'pmd'
        d = '-d'
        src_filepath = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/' + extracted_file_name
        f = '-f'
        report_format = 'csv'
        r = '-R'
        rule_set = 'rulesets/java/u2l_performance.xml'
        use_version = '--use-version'
        version = 'java-1.8'
        report_file = '--report-file'
        report_file_name = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/Performance_Report.csv'

        command = ['sh', run_path, pmd, d, src_filepath, f, report_format, r, rule_set, use_version, version, report_file, report_file_name]

        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output, error = process.communicate()
        except:
            return 'PMD failed'

        #Best Practices
        redis_client.rpush(redis_queue_name, '19/40::Executing pmd script : Best Practice')
        run_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/bin/run.sh'
        pmd = 'pmd'
        d = '-d'
        src_filepath = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/' + extracted_file_name
        f = '-f'
        report_format = 'csv'
        r = '-R'
        rule_set = 'rulesets/java/u2l_bestpractices.xml'
        use_version = '--use-version'
        version = 'java-1.8'
        report_file = '--report-file'
        report_file_name = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/BestPractices_Report.csv'

        command = ['sh', run_path, pmd, d, src_filepath, f, report_format, r, rule_set, use_version, version, report_file, report_file_name]

        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output, error = process.communicate()
        except:
            return 'PMD failed'

        #Multithreading
        redis_client.rpush(redis_queue_name, '20/40::Executing pmd script : Multithreading')
        run_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/bin/run.sh'
        pmd = 'pmd'
        d = '-d'
        src_filepath = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/' + extracted_file_name
        f = '-f'
        report_format = 'csv'
        r = '-R'
        rule_set = 'rulesets/java/u2l_multithread.xml'
        use_version = '--use-version'
        version = 'java-1.8'
        report_file = '--report-file'
        report_file_name = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/MultiThreading_Report.csv'

        command = ['sh', run_path, pmd, d, src_filepath, f, report_format, r, rule_set, use_version, version, report_file, report_file_name]

        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output, error = process.communicate()
        except:
            return 'PMD failed'

        #Security
        redis_client.rpush(redis_queue_name, '21/40::Executing pmd script : Security')
        run_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/bin/run.sh'
        pmd = 'pmd'
        d = '-d'
        src_filepath = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/' + extracted_file_name
        f = '-f'
        report_format = 'csv'
        r = '-R'
        rule_set = 'rulesets/java/u2l_security.xml'
        use_version = '--use-version'
        version = 'java-1.8'
        report_file = '--report-file'
        report_file_name = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/Security_Report.csv'

        command = ['sh', run_path, pmd, d, src_filepath, f, report_format, r, rule_set, use_version, version, report_file, report_file_name]

        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            output, error = process.communicate()
        except:
            return 'PMD failed'
        
    elif(tool_analysis_type == 'canalysis'):
        c_report_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'
        c_remediation_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_remediation_report.xlsx'
        c_inventory_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_inventory_report.xlsx'

        form_source_ver = 'HP-UX'
        redis_client.rpush(redis_queue_name, '7/40::Executing Scan Source Script')
        source_versions = ["AIX", "Solaris", "HP-UX"]
        if form_source_ver not in source_versions:
            error_message = {'error' :"Invalid Source Vesion"}
            return jsonify(error_message), 404
        
        # file_name = "UNIX_HPE_20231128_054027"
        file_path = re.sub(r'_\d{8}_\d{6}', '', file_name)
        print(file_path)
        source_path = '/usr/u2l/u2l_backend/projects/'+ file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/' + file_path
        print(source_path)
        script_path = '/usr/u2l/u2l_backend/scan_source.sh'
        source_ver = form_source_ver
        target_ver = 'RHEL'
        command = ['sh', script_path, source_path, source_ver, target_ver]
        process = subprocess.Popen(command, stdout=subprocess.PIPE)
        output = process.communicate()[0].decode().strip()
        lines = output.split('\n')
        print(lines)
        redis_client.rpush(redis_queue_name, '8/40::Completed Executing Scan Source Script')
        redis_client.rpush(redis_queue_name, '9/40::Generating Final Report')

        ic_lines = []
        with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
            for line in file:
                if line.strip().startswith('/usr/'):
                    ic_lines.append(line)

        with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
            file.write(''.join(ic_lines))

        ic_lines = []
        with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
            for line in file:
                modified_line = line.replace(':', ',', 1)
                ic_lines.append(modified_line)

        with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
            file.write(''.join(ic_lines))

        final_report_path = '/usr/u2l/u2l_backend/final_report.csv'
        if os.path.isfile(final_report_path) and os.path.getsize(final_report_path) > 0:
            df = pd.read_csv(final_report_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ["FILE PATH","LINE NUMBER","SOURCE CODE","ERROR MESSAGE","REMEDIATION"]
            df = df.drop(df.index[0])
            O28 = df['FILE PATH'].nunique()
            P28 = df['FILE PATH'].count()
            print(df)

        generate_workbook(df, c_report_path,'Source Scan Remedy Detail')
        redis_client.rpush(redis_queue_name, '10/40::Reports almost ready')

        redis_client.rpush(redis_queue_name, '11/40::Performing cppcheck !!')

        source_path = '/usr/u2l/u2l_backend/projects/'+ file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/' + file_path
        # source_path = '/usr/u2l/u2l_backend/projects/' + file_name +'/'+ file_name + '.zip'
        stdout, stderr = run_cppcheck(source_path)

        with open("cppcheck_output.txt", "w") as file:
            file.write(stderr)
        
        with open('cppcheck_output.txt', 'r') as file:
            content = file.readlines()

        file_paths = []
        line_numbers = []
        error_messages = []
        error_lines = []

        pattern = re.compile(r'^(.*?):(\d+):\d+: (.*?): (.*)$')

        current_error_lines = []

        for line in content:
            match = pattern.match(line)
            if match:
                if current_error_lines:
                    error_lines.append(current_error_lines)
                    current_error_lines = []
                
                file_paths.append(match.group(1))
                line_numbers.append(int(match.group(2)))
                error_messages.append(match.group(4))
            else:
                current_error_lines.append(line)

        if current_error_lines:
            error_lines.append(current_error_lines)

        data = {
            'File Path': file_paths,
            'Line Number': line_numbers,
            'Error Message': error_messages,
            'Error Lines': [''.join(lines) for lines in error_lines]
        }

        df = pd.DataFrame(data)
        df.columns = ['FILE PATH', 'LINE NUMBER', 'ERROR MESSAGE', 'ERROR LINES']
        O29 = df['FILE PATH'].nunique()
        P29 = df['FILE PATH'].count()

        df.reset_index(drop=True, inplace=True)
        df.index = df.index + 1

        book = load_workbook(c_report_path)
        writer = pd.ExcelWriter(c_report_path, engine='openpyxl')
        writer.book = book
        sheet_name = 'Compilation Warning Detail'
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, header=False)
        writer.save()
        redis_client.rpush(redis_queue_name, '12/40::Compilation reports are ready')

        sheet_transfer('Source Scan Remedy Detail', '1. Source Scan Remedy Detail', 6, 1, c_report_path, c_remediation_path)
        sheet_transfer('Compilation Warning Detail', '2. Compilation Warning Detail', 6, 1, c_report_path, c_remediation_path)
        # sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, c_report_path, c_inventory_path)
        # sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, c_report_path, c_remediation_path)

        print("Start 13 : Source Code Inventory")
        redis_client.rpush(redis_queue_name, '35/40::Generating logs for Inventory')
        regex = r'^.*$'
        source_code_c_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
        print(source_code_c_path)
        if os.path.isfile(source_code_c_path) and os.path.getsize(source_code_c_path) > 0:
            
            with open(source_code_c_path, 'r') as f:
                data = f.read()
            lines = data.split('\n')
            log_entries = []
            for line in lines:
                match = re.match(regex, line)
                if match:
                    log_entries.append(match.group())
            df = pd.DataFrame(log_entries, columns=['log_message'])

            start_str = '###########  Execution 0001_stepcounter Starts  ###########'
            end_str = '###########  Execution 0001_stepcounter Ends ###########'
            start_df = df[df['log_message'].str.contains(start_str)]
            end_df = df[df['log_message'].str.contains(end_str)]

            start_index = start_df.index[0]
            end_index = end_df.index[0]
            extracted_df = df.iloc[start_index+1:end_index]

            result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
            result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
            result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
            result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
            result_df = result_df.drop(['log_message', 'col2'], axis=1)
            result_df = result_df.reset_index(drop=True)
            result_df.index += 1

            artefact_type_dict = result_df['Type'].unique()                                     # ['C/C++' 'CShell' 'KShell']
            type_counts = result_df['Type'].value_counts()
            type_counts_dict = type_counts.to_dict()                                            # {'C/C++': 690, 'KShell': 107, 'CShell': 1}
            F6 = sum(type_counts_dict.values())
            result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
            autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
            autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
            O6 = sum(autual_by_type_dict.values())
            result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
            total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
            total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'C/C++': 958392, 'CShell': 9, 'KShell': 9689}

            generate_workbook(result_df, c_report_path,'Source Code Inventory')

            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, c_report_path, c_inventory_path)
            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, c_report_path, c_remediation_path)
            print('Transfer Completed !!')
            redis_client.rpush(redis_queue_name, '36/40::Inventory reports are ready')
        
        wb = load_workbook(c_remediation_path)
        print('1')
        ws = wb['Environment Information']
        print('2')
        start_row = 22
        start_col = 2
        col_actual = 2
        col_total = 3

        for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
            col_idx = start_col
            ws.cell(row=row_idx, column=col_idx, value=artefact_type)

            value = type_counts_dict.get(artefact_type, 0)
            col_idx = start_col + 1
            ws.cell(row=row_idx, column=col_idx, value=value)

            actual_value = autual_by_type_dict.get(artefact_type, 0)
            col_idx = start_col + col_actual
            ws.cell(row=row_idx, column=col_idx, value=actual_value)

            actual_value = total_by_type_dict.get(artefact_type, 0)
            col_idx = start_col + col_total
            ws.cell(row=row_idx, column=col_idx, value=actual_value)

        wb.save(c_remediation_path)

        wb = load_workbook(c_inventory_path)
        print('1')
        ws = wb['Scope of Artefacts']
        print('2')
        start_row = 2
        start_col = 1
        col_total = 2

        for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
            col_idx = start_col
            ws.cell(row=row_idx, column=col_idx, value=artefact_type)

            value = type_counts_dict.get(artefact_type, 0)
            col_idx = start_col + 1
            ws.cell(row=row_idx, column=col_idx, value=value)

            actual_value = total_by_type_dict.get(artefact_type, 0)
            col_idx = start_col + col_total
            ws.cell(row=row_idx, column=col_idx, value=actual_value)

        wb.save(c_inventory_path)

        summary_sheet_generation(c_remediation_path, '2. Compilation Warning Detail', '2. Compilation Warning Summary', 5, 3, c_remediation_path, 7, 2)
        summary_sheet_generation(c_remediation_path, '1. Source Scan Remedy Detail', '1. Source Scan Remedy Summary', 5, 5, c_remediation_path, 7, 2)

        print("Sheet : Source Code Inventory : CREATED")
        print("End 13 : Source Code Inventory")

        print("Start 17 : Other inventory")
        redis_client.rpush(redis_queue_name, '37/40::Generating logs for other inventory if present')
        yyyy_mmdd = yMD[:4] + '-' + yMD[4:]

        # Convert check_files.out
        check_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/check_files.out'
        other_inventory_df1 = None
        if os.path.isfile(check_files_dir_path) and os.path.getsize(check_files_dir_path) > 0:
            other_inventory_df1 = pd.read_csv(check_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df1.index = other_inventory_df1.index + 1
            print(other_inventory_df1)

        # Convert exclude_files.out
        exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/exclude_files.out'
        other_inventory_df2 = None
        if os.path.isfile(exclude_files_dir_path) and os.path.getsize(exclude_files_dir_path) > 0:
            other_inventory_df2 = pd.read_csv(exclude_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df2.index = other_inventory_df2.index + 1
            print(other_inventory_df2)

        # Convert c_files.out
        c_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/c_files.out'
        other_inventory_df3 = None
        if os.path.isfile(c_files_dir_path) and os.path.getsize(c_files_dir_path) > 0:
            other_inventory_df3 = pd.read_csv(c_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df3.index = other_inventory_df3.index + 1
            print(other_inventory_df3)

        # Convert csh_list.out
        csh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/csh_list.out'
        other_inventory_df4 = None
        if os.path.isfile(csh_list_dir_path) and os.path.getsize(csh_list_dir_path) > 0:
            other_inventory_df4 = pd.read_csv(csh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df4.index = other_inventory_df4.index + 1
            print(other_inventory_df4)    

        # Convert ksh_list.out
        ksh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ksh_list.out'
        other_inventory_df5 = None
        if os.path.isfile(ksh_list_dir_path) and os.path.getsize(ksh_list_dir_path) > 0:
            other_inventory_df5 = pd.read_csv(ksh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df5.index = other_inventory_df5.index + 1
            print(other_inventory_df5)

        # Convert ot2_list.out
        ot2_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ot2_list.out'
        other_inventory_df6 = None
        if os.path.isfile(ot2_list_dir_path) and os.path.getsize(ot2_list_dir_path) > 0:
            other_inventory_df6 = pd.read_csv(ot2_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df6.index = other_inventory_df6.index + 1
            print(other_inventory_df6)   

        
        other_inventory_combined_df = pd.concat([other_inventory_df1, other_inventory_df2, other_inventory_df3, other_inventory_df4, other_inventory_df5, other_inventory_df6], ignore_index=True)
        other_inventory_combined_df = other_inventory_combined_df.drop_duplicates()
        print(other_inventory_combined_df)

        writer = pd.ExcelWriter(c_inventory_path, engine='openpyxl', mode='a')
        book = load_workbook(c_inventory_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
        writer.save()
        print('c_inventory other inventory added')

        writer = pd.ExcelWriter(c_remediation_path, engine='openpyxl', mode='a')
        book = load_workbook(c_remediation_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
        writer.save()
        print('c_remediation other inventory added')

        print("Sheet : Other Inventory : CREATED")
        print("End 17 : Other inventory\n")
        redis_client.rpush(redis_queue_name, '38/40::Other Inventory check is complete')

        middleware_type = 'Tuxedo'
        if(middleware_type != 'Tuxedo'):
            filename = c_remediation_path
            sheet_name = '6. Middleware-Tuxedo'
            workbook = load_workbook(filename)

            if sheet_name in workbook.sheetnames:
                sheet_to_delete = workbook[sheet_name]
                workbook.remove(sheet_to_delete)
                workbook.save(filename)
            print('Tuxedo sheet removed !!!')

        print('Migration Summary')
        wb = load_workbook(c_remediation_path)
        ws = wb['Migration Summary']
        start_col_1 = 15
        start_col_2 = 16
        start_row_1 = 28
        start_row_2 = 29
        if O28 != 0:
            ws.cell(row=start_row_1, column=start_col_1, value=O28)
        if O29 != 0:
            ws.cell(row=start_row_2, column=start_col_1, value=O29)
        if P28 != 0:
            ws.cell(row=start_row_1, column=start_col_2, value=P28)
        if P29 != 0:
            ws.cell(row=start_row_2, column=start_col_2, value=P29)

        O7 = P28 + P29
        F7 = O28 + O29

        if O7 != 0:
            ws.cell(row=7, column=15, value=O7)
        if O6 != 0:
            ws.cell(row=6, column=15, value=O6)
        if F7 != 0:
            ws.cell(row=7, column=6, value=F7)
        if F6 != 0:
            ws.cell(row=6, column=6, value=F6)

        wb.save(c_remediation_path)
        redis_client.rpush(redis_queue_name, '39/40::Reports are ready')
        redis_client.rpush(redis_queue_name, '40/40::CODE ASSESSMENT SUCCESSFUL')

    # Converting logs into dataframes and generating reports
    try:

        if(tool_analysis_type == 'javaanalysis'):

            yyyy_mmdd = yMD[:4] + '-' + yMD[4:]
            java_remediation_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
            java_high_level_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
            java_inventory_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_inventory_report.xlsx'
            java_report_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            O29=0
            P29=0
            O32=0
            P32=0
            O35=0
            P35=0
            O36=0
            P36=0
            
            # Convert javaSourceScanRemedy
            logger.info("START 1 : javaSourceScanRemedy")
            redis_client.rpush(redis_queue_name, '22/40::Getting report from logs generated : javaSourceScanRemedy')
            javaSourceScanRemedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/javaSourceScanRemedy'
            if os.path.isfile(javaSourceScanRemedy) and os.path.getsize(javaSourceScanRemedy) > 0:
                df = pd.read_csv(javaSourceScanRemedy, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
                df['FILE NAME'] = df['FILE NAME'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                generate_workbook(df, java_report_path,'javaSourceScanRemedy')

                df1 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='javaSourceScanRemedy', skiprows=1, header=None)
                df1 = df1.iloc[:, [4]]
            logger.info("END 1 : javaSourceScanRemedy")

            # Convert javaSourceCode2Remedy
            logger.info("START 2 : javaSourceCode2Remedy")
            redis_client.rpush(redis_queue_name, '23/40::Getting report from logs generated : javaSourceCode2Remedy')
            javaSourceCode2Remedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/javaSourceCode2Remedy'
            if os.path.isfile(javaSourceCode2Remedy) and os.path.getsize(javaSourceCode2Remedy) > 0:
                with open(javaSourceCode2Remedy, 'r') as file:
                    lines = file.readlines()

                with open(javaSourceCode2Remedy, 'w') as file:
                    for line in lines:
                        colon_count = line.count(':')
                        if colon_count == 4:
                            last_colon_index = line.rfind(':')
                            line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                        else:
                            line_without_last_colon = line

                        if line.startswith('\t'):
                            line_without_first_tab = line[1:]
                        else:
                            line_without_first_tab = line

                        file.write(line_without_last_colon)

                with open(javaSourceCode2Remedy, 'r') as file:
                    lines = file.readlines()

                with open(javaSourceCode2Remedy, 'w') as file:
                    for line in lines:
                        first_tab_index = line.find('\t')
                        if first_tab_index != -1:
                            line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                            line_without_tabs = line_with_colon.replace('\t', '')
                        else:
                            line_without_tabs = line

                        file.write(line_without_tabs)

                df = pd.read_csv(javaSourceCode2Remedy, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
                df['FILE NAME'] = df['FILE NAME'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                O33 = df['FILE NAME'].nunique()
                P33 = df['FILE NAME'].count()
                
                generate_workbook(df, java_report_path,'javaSourceCode2Remedy')

                book = load_workbook(java_report_path)
                writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
                writer.book = book
                sheet_name = 'javaSourceCode2Remedy'
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
                writer.save()

                df2 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='javaSourceCode2Remedy', skiprows=0)
                df2 = df2.iloc[:, [1,2,3,4,5]]

                book = load_workbook(java_report_path)
                writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
                writer.book = book
                sheet_name = 'javaSourceCode2Remedy_1'
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
                writer.save()

                sheet_transfer('javaSourceCode2Remedy_1', '6. Java Source Scan Remedy', 6, 1, java_report_path, java_remediation_path)
                summary_sheet_generation(java_remediation_path, '6. Java Source Scan Remedy', '6. Java Source Scan Summary', 5, 4, java_remediation_path, 7, 2)
            logger.info("SHEET : 6. Java Source Scan Remedy : CREATED")
            logger.info("SHEET : 6. Java Source Scan Summary : CREATED")
            logger.info("END 2 : javaSourceCode2Remedy")

            # Convert sourceScanRemedy
            logger.info("START 3 : sourceScanRemedy")
            redis_client.rpush(redis_queue_name, '24/40::Getting report from logs generated : sourceScanRemedy')
            sourceScanRemedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/sourceScanRemedy'
            if os.path.isfile(sourceScanRemedy_dir_path) and os.path.getsize(sourceScanRemedy_dir_path) > 0:
                ic_lines = []
                def remove_trailing_tabs(line):
                    return re.sub(r'\t*$', '', line)
                with open(sourceScanRemedy_dir_path, 'r') as file:
                    for line in file:
                        cleaned_line = remove_trailing_tabs(line)
                        if cleaned_line.strip().startswith('IC_'):
                            ic_lines.append(cleaned_line)

                with open(sourceScanRemedy_dir_path, 'w') as file:
                    file.write(''.join(ic_lines))

                df = pd.read_csv(sourceScanRemedy_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.index = df.index + 1
                df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
                df['FILE NAME'] = df['FILE NAME'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                generate_workbook(df, java_report_path,'sourceScanRemedy')

                df1 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='sourceScanRemedy', skiprows=1, header=None)
                df1 = df1.iloc[:, [4]]
            logger.info("END 3 : sourceScanRemedy")

            # Convert sourceCode2Remedy
            logger.info("START 4 : sourceCode2Remedy")
            redis_client.rpush(redis_queue_name, '25/40::Getting report from logs generated : sourceCode2Remedy')
            sourceCode2Remedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/sourceCode2Remedy'
            if os.path.isfile(sourceCode2Remedy_dir_path) and os.path.getsize(sourceCode2Remedy_dir_path) > 0:
                with open(sourceCode2Remedy_dir_path, 'r') as file:
                    lines = file.readlines()

                with open(sourceCode2Remedy_dir_path, 'w') as file:
                    for line in lines:
                        colon_count = line.count(':')
                        if colon_count == 4:
                            last_colon_index = line.rfind(':')
                            line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                        else:
                            line_without_last_colon = line

                        if line.startswith('\t'):
                            line_without_first_tab = line[1:]
                        else:
                            line_without_first_tab = line

                        file.write(line_without_last_colon)

                with open(sourceCode2Remedy_dir_path, 'r') as file:
                    lines = file.readlines()

                with open(sourceCode2Remedy_dir_path, 'w') as file:
                    for line in lines:
                        first_tab_index = line.find('\t')
                        if first_tab_index != -1:
                            line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                            line_without_tabs = line_with_colon.replace('\t', '')
                        else:
                            line_without_tabs = line

                        file.write(line_without_tabs)
                df = pd.read_csv(sourceCode2Remedy_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.index = df.index + 1
                df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
                df['FILE NAME'] = df['FILE NAME'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                O35 = df['FILE NAME'].nunique()
                P35 = df['FILE NAME'].count()
                generate_workbook(df, java_report_path,'sourceCode2Remedy')

                book = load_workbook(java_report_path)
                writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
                writer.book = book
                sheet_name = 'sourceCode2Remedy'
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
                writer.save()

                df2 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='sourceCode2Remedy', skiprows=0)
                df2 = df2.iloc[:, [1,2,3,4,5]]

                book = load_workbook(java_report_path)
                writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
                writer.book = book
                sheet_name = 'sourceCode2Remedy_1'
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
                writer.save()

                sheet_transfer('sourceCode2Remedy_1', '8. Affected Framework Remedy', 6, 1, java_report_path, java_remediation_path)
                summary_sheet_generation(java_remediation_path, '8. Affected Framework Remedy', '8. Affected Framework Summary', 5, 4, java_remediation_path, 7, 2)
            logger.info("SHEET : 8. Affected Framework Remedy : CREATED")
            logger.info("SHEET : 8. Affected Framework Summary : CREATED")
            logger.info("END 4 : sourceCode2Remedy")

            # Convert javaUniqRulesGrepped.txt
            # logger.info("START 5 : javaUniqRulesGrepped.txt")
            # redis_client.rpush(redis_queue_name, '26/40::Getting report from logs generated : javaUniqRulesGrepped')
            # javaUniqRulesGrepped_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/javaUniqRulesGrepped.txt'
            # if os.path.isfile(javaUniqRulesGrepped_dir_path) and os.path.getsize(javaUniqRulesGrepped_dir_path) > 0:
            #     ic_lines = []
            #     with open(javaUniqRulesGrepped_dir_path, 'r') as file:
            #         for line in file:
            #             cleaned_line = re.sub(r'\t*$', '', line)
            #             if cleaned_line.strip().startswith('IC_'):
            #                 ic_lines.append(cleaned_line)

            #     with open(javaUniqRulesGrepped_dir_path, 'w') as file:
            #         file.write('\n'.join(ic_lines))

            #     with open(javaUniqRulesGrepped_dir_path, 'r') as file:
            #         lines = [re.sub(r'\t*$', '', line) for line in file]

            #     with open(javaUniqRulesGrepped_dir_path, 'w') as file:
            #         file.write('\n'.join(lines))
            #     df = pd.read_csv(javaUniqRulesGrepped_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            #     df.index = df.index + 1
            #     df.columns = ['RULE ID', 'ENTITY', 'PACKAGE', 'OBJECT', 'REMEDY']

            #     generate_workbook(df, java_report_path,'Possible Import Remediations')
            #     sheet_transfer('Possible Import Remediations', '7. Possible Import Remediation', 6, 1, java_report_path, java_remediation_path)
            #     summary_sheet_generation(java_remediation_path, '7. Possible Import Remediation', '7. Import Remediation Summary', 5, 6, java_remediation_path, 8, 2)
            # logger.info("SHEET : 7. Possible Import Remediation : CREATED")
            # logger.info("SHEET : 7. Import Remediation Summary : CREATED")
            # logger.info("END 5 : javaUniqRulesGrepped.txt")

            # Convert javacWarnings.list
            logger.info("Start 6 : javacWarnings.list")
            redis_client.rpush(redis_queue_name, '27/40::Getting report from logs generated : javacWarnings')
            javacWarnings_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/javacLog/' + yyyy_mmdd + '/javacWarnings.list'
            if os.path.isfile(javacWarnings_dir_path) and os.path.getsize(javacWarnings_dir_path) > 0:
                df = pd.read_csv(javacWarnings_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.index = df.index + 1
                df.columns = ['WARNING MESSAGE WITH TYPE', 'FILE PATH', 'FILE NAME', 'LINE', 'Column5']
                df = df.drop(columns=['Column5'])
                O29 = df['FILE NAME'].nunique()
                P29 = df['FILE NAME'].count()

                generate_workbook(df, java_report_path,'Compilation Warning Report')
                sheet_transfer('Compilation Warning Report', '2. Compilation Warning Detail', 6, 1, java_report_path, java_remediation_path)
                summary_sheet_generation(java_remediation_path, '2. Compilation Warning Detail', '2. Compilation Warning Summary', 5, 2, java_remediation_path, 9, 2)
            logger.info("Sheet : 2. Compilation Warning Detail : CREATED")
            logger.info("Sheet : 2. Compilation Warning Summary : CREATED")
            logger.info("End 6 : javacWarnings.list")

            # Convert import_lines.out
            logger.info("Start 7 : import_lines.out")
            redis_client.rpush(redis_queue_name, '28/40::Getting report from logs generated : import_lines')
            importLines_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines.out'
            if os.path.isfile(importLines_dir_path) and os.path.getsize(importLines_dir_path) > 0:
                df = pd.read_csv(importLines_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['IMPORTS', 'FILE NAME', 'LINE NUMBER']
                df.index = df.index + 1

                generate_workbook(df, java_report_path,'Import Class Report')
                sheet_transfer('Import Class Report', 'Import Class Detail', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : Import Class Detail : CREATED")
            logger.info("End 7 : import_lines.out")

            # Convert import_lines_jsp.out
            logger.info("Start 8 : import_lines_jsp.out")
            redis_client.rpush(redis_queue_name, '29/40::Getting report from logs generated : import_lines_jsp')
            importLinesJsp_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines_jsp.out'
            if os.path.isfile(importLinesJsp_dir_path) and os.path.getsize(importLinesJsp_dir_path) > 0:
                df = pd.read_csv(importLinesJsp_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['FILE NAME', 'LINE NUMBER', 'IMPORTED JSP']
                df['FILE NAME'] = './' + df['FILE NAME']
                df.index = df.index + 1

                generate_workbook(df, java_report_path,'Import JSP Report')
                sheet_transfer('Import JSP Report', 'Import JSP Detail', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : Import JSP Detail : CREATED")
            logger.info("End 8 : import_lines_jsp.out")
            
            # Best Practices
            logger.info("Start 9 : Best Practices")
            redis_client.rpush(redis_queue_name, '30/40::Getting report from logs generated : Best Practices')
            best_pratices_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/BestPractices_Report.csv'
            if os.path.isfile(best_pratices_path) and os.path.getsize(best_pratices_path) > 0:
                df = pd.read_csv(best_pratices_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
                df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                df = df.drop(df.index[0])

                generate_workbook(df, java_report_path,'Best_Practices')
                sheet_transfer('Best_Practices', '9. Best Practices', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : 9. Best Practices : CREATED")
            logger.info("End 9 : Best Practices")

            # MultiThreading
            logger.info("Start 10 : MultiThreading")
            redis_client.rpush(redis_queue_name, '31/40::Getting report from logs generated : Multithreading')
            multithreading_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/MultiThreading_Report.csv'
            if os.path.isfile(multithreading_path) and os.path.getsize(multithreading_path) > 0:
                df = pd.read_csv(multithreading_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
                df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                df = df.drop(df.index[0])

                generate_workbook(df, java_report_path,'Multithreading')
                sheet_transfer('Multithreading', '10. MultiThreading', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : 10. MultiThreading : CREATED")
            logger.info("End 10 : MultiThreading")

            # Performance
            logger.info("Start 11 : Performance")
            redis_client.rpush(redis_queue_name, '32/40::Getting report from logs generated : Perfromance')
            performance_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/Performance_Report.csv'
            if os.path.isfile(performance_path) and os.path.getsize(performance_path) > 0:
                df = pd.read_csv(performance_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
                df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                df = df.drop(df.index[0])

                generate_workbook(df, java_report_path,'Performance')
                sheet_transfer('Performance', '11. Performance', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : 11. Performance : CREATED")
            logger.info("End 11 : Performance")

            # Security
            logger.info("Start 12 : Security")
            redis_client.rpush(redis_queue_name, '33/40::Getting report from logs generated : Security')
            security_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/pmd-bin-6.55.0/Reports_1/Security_Report.csv'
            if os.path.isfile(security_path) and os.path.getsize(security_path) > 0:
                df = pd.read_csv(security_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
                df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
                df = df.drop(df.index[0])

                generate_workbook(df, java_report_path,'Security')
                sheet_transfer('Security', '12. Security', 6, 1, java_report_path, java_remediation_path)
            logger.info("Sheet : 12. Security : CREATED")
            logger.info("End 12 : Security")
            
            #Source Code Inventory
            logger.info("Start 13 : Source Code Inventory")
            redis_client.rpush(redis_queue_name, '34/40::Getting report from logs generated : Source Code Inventory')
            regex = r'^.*$'
            source_code_java_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
            print(source_code_java_path)
            if os.path.isfile(source_code_java_path) and os.path.getsize(source_code_java_path) > 0:
                
                with open(source_code_java_path, 'r') as f:
                    data = f.read()
                lines = data.split('\n')
                log_entries = []
                for line in lines:
                    match = re.match(regex, line)
                    if match:
                        log_entries.append(match.group())
                df = pd.DataFrame(log_entries, columns=['log_message'])

                start_str = '###########  Execution 0001_stepcounter Starts  ###########'
                end_str = '###########  Execution 0001_stepcounter Ends ###########'
                start_df = df[df['log_message'].str.contains(start_str)]
                end_df = df[df['log_message'].str.contains(end_str)]

                start_index = start_df.index[0]
                end_index = end_df.index[0]
                extracted_df = df.iloc[start_index+1:end_index]

                result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
                result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
                result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
                result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
                result_df = result_df.drop(['log_message', 'col2'], axis=1)
                result_df = result_df.reset_index(drop=True)
                result_df.index += 1

                artefact_type_dict = result_df['Type'].unique()                                     # ['Java', 'KShell']
                type_counts = result_df['Type'].value_counts()
                type_counts_dict = type_counts.to_dict()                                            # {'KShell': 23, 'Java': 2}
                F6 = sum(type_counts_dict.values())
                result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
                autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
                autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
                O6 = sum(autual_by_type_dict.values())
                result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
                total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
                total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'Java': 424, 'KShell': 5841}

                generate_workbook(result_df, java_report_path,'Source Code Inventory')

                sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, java_report_path, java_inventory_path)
                sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, java_report_path, java_remediation_path)
                sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, java_report_path, java_high_level_path)
                print('Transfer Completed !!')
                redis_client.rpush(redis_queue_name, '35/40::Transferring generated sheets to excel workbook')

                wb = load_workbook(java_remediation_path)
                print('1')
                ws = wb['Environment Information']
                print('2')
                start_row = 22
                start_col = 2
                col_actual = 2
                col_total = 3

                for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                    col_idx = start_col
                    ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                    value = type_counts_dict.get(artefact_type, 0)
                    col_idx = start_col + 1
                    ws.cell(row=row_idx, column=col_idx, value=value)

                    actual_value = autual_by_type_dict.get(artefact_type, 0)
                    col_idx = start_col + col_actual
                    ws.cell(row=row_idx, column=col_idx, value=actual_value)

                    actual_value = total_by_type_dict.get(artefact_type, 0)
                    col_idx = start_col + col_total
                    ws.cell(row=row_idx, column=col_idx, value=actual_value)

                wb.save(java_remediation_path)

                wb = load_workbook(java_high_level_path)
                ws = wb['Environment Information']
                start_row = 22
                start_col = 2
                col_actual = 2
                col_total = 3

                for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                    col_idx = start_col
                    ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                    value = type_counts_dict.get(artefact_type, 0)
                    col_idx = start_col + 1
                    ws.cell(row=row_idx, column=col_idx, value=value)

                    actual_value = autual_by_type_dict.get(artefact_type, 0)
                    col_idx = start_col + col_actual
                    ws.cell(row=row_idx, column=col_idx, value=actual_value)

                    actual_value = total_by_type_dict.get(artefact_type, 0)
                    col_idx = start_col + col_total
                    ws.cell(row=row_idx, column=col_idx, value=actual_value)

                wb.save(java_high_level_path)
            print("Sheet : Source Code Inventory : CREATED")
            print("End 13 : Source Code Inventory")

            # Convert OS_diff.out
            print("Start 14 : OS_diff.out")
            OSDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/OS_diff.out'
            if os.path.isfile(OSDiff_dir_path) and os.path.getsize(OSDiff_dir_path) > 0:
                print("inside 12")

                with open(OSDiff_dir_path, 'r') as file:
                    lines = file.readlines()

                modified_lines = []
                for line in lines:
                    if not line.startswith('OS-IC'):
                        line = line.replace('\t', '')
                        modified_lines[-1] = modified_lines[-1].rstrip() + line.strip()
                    else:
                        modified_lines.append(line)

                with open(OSDiff_dir_path, 'w') as file:
                    file.writelines(modified_lines)

                with open(OSDiff_dir_path, 'r') as file:
                    lines = file.readlines()

                modified_lines = []
                for line in lines:
                    parts = line.split()
                    parts = [part.strip() for part in parts]
                    modified_line = '@'.join(parts)
                    modified_lines.append(modified_line + '\n')

                with open(OSDiff_dir_path, 'w') as file:
                    file.writelines(modified_lines)

                with open(OSDiff_dir_path, 'r') as file:
                    lines = file.readlines()

                modified_lines = []
                for line in lines:
                    modified_line = re.sub(r'@', lambda match: ' ' if match.start() > line.find('@', line.find('@', line.find('@') + 1) + 1) else match.group(), line.strip())
                    modified_lines.append(modified_line + '\n')

                with open(OSDiff_dir_path, 'w') as file:
                    file.writelines(modified_lines)

                df = pd.read_csv(OSDiff_dir_path, sep='@', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)

                df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                O32 = df['FILE NAME'].nunique()
                P32 = df['FILE NAME'].count()

                generate_workbook(df, java_report_path,'OS Analysis Details')
                sheet_transfer('OS Analysis Details', '5. OS Analysis Detail', 6, 1, java_report_path, java_remediation_path)
                summary_sheet_generation(java_remediation_path, '5. OS Analysis Detail', '5. OS Analysis Summary', 5, 5, java_remediation_path, 8, 2)
            print("Sheet : 5. OS Analysis Detail : CREATED")
            print("Sheet : 5. OS Analysis Summary : CREATED")
            print("End 14 : OS_diff.out\n")

            # Convert jdk_diff.out
            print("Start 15 : jdk_diff.out")
            jdkDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/jdk_diff.out'
            if os.path.isfile(jdkDiff_dir_path) and os.path.getsize(jdkDiff_dir_path) > 0:
                df = pd.read_csv(jdkDiff_dir_path, sep='\s+', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df = df.drop(columns=['Column1'])
                df.index = df.index + 1

                generate_workbook(df, java_report_path,'JDK Anaylsis Details')
                sheet_transfer('JDK Anaylsis Details', '3. JDK Analysis Detail', 6, 1, java_report_path, java_remediation_path)
                summary_sheet_generation(java_remediation_path, '3. JDK Analysis Detail', '3. JDK Analysis Summary', 5, 6, java_remediation_path, 8, 2)
            print("Sheet : 3. JDK Analysis Detail : CREATED")
            print("Sheet : 3. JDK Analysis Summary : CREATED")
            print("End 15 : jdk_diff.out\n")

            #java_high_level_report.xlsx - Source Code Impacted
            print("Start 16 : Source Code Impacted")
            sheet_configs = {
                '1. Compilation Error Report': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '2. Compilation Warning Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '3. JDK Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '4. J2EE Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '5. OS Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '6. Java Source Scan Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1},
                # '7. Possible Import Remediation': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
                '8. Affected Framework Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1}
            }

            dfs = []

            for sheet_name, config in sheet_configs.items():
                df = pd.read_excel(java_remediation_path, engine='openpyxl', sheet_name=sheet_name, usecols=config['usecols'], skiprows=config['skiprows'])
                df = df.drop(df.columns[0], axis=1)
                df = df.iloc[:, config['column_index']]

                if not df.empty and not df.isnull().all().all():
                    dfs.append(df)

            df_combined = pd.concat(dfs, ignore_index=True)
            df_combined = df_combined.drop_duplicates()
            df_combined = df_combined.reset_index(drop=True)
            df_combined = df_combined.dropna()
            df_count = len(df_combined)
            F7 = df_count
            print(df_combined)

            writer = pd.ExcelWriter(java_high_level_path, engine='openpyxl', mode='a')
            book = load_workbook(java_high_level_path)
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df_combined.to_excel(writer, sheet_name='Source Code Impacted', index=False, startrow=6, startcol=1, header=False)
            writer.save()

            wb = load_workbook(java_remediation_path)
            ws = wb['Migration Summary']
            start_row = 7
            start_col = 6
            ws.cell(row=start_row, column=start_col, value=df_count)

            wb.save(java_remediation_path)

            wb = load_workbook(java_high_level_path)
            ws = wb['Analysis Summary']
            start_row = 7
            start_col = 6
            ws.cell(row=start_row, column=start_col, value=df_count)

            wb.save(java_high_level_path)

            # fetch_store_java_data(form_project_name, form_application_name, java_report_path)

            # Migration Summary : java_remediation_report
            print('Migration Summary')
            wb = load_workbook(java_remediation_path)
            ws = wb['Migration Summary']
            start_col_1 = 15
            start_row_1 = 29
            start_row_2 = 32
            start_row_3 = 33
            start_row_4 = 35
            start_col_2 = 16
            start_row_5 = 29
            start_row_6 = 32
            start_row_7 = 33
            start_row_8 = 35
            if O29 != 0:
                ws.cell(row=start_row_1, column=start_col_1, value=O29)
            if O32 != 0:
                ws.cell(row=start_row_2, column=start_col_1, value=O32)
            if O33 != 0:
                ws.cell(row=start_row_3, column=start_col_1, value=O33)
            if O35 != 0:
                ws.cell(row=start_row_4, column=start_col_1, value=O35)
            if P29 != 0:
                ws.cell(row=start_row_5, column=start_col_2, value=P29)
            if P32 != 0:
                ws.cell(row=start_row_6, column=start_col_2, value=P32)
            if P33 != 0:
                ws.cell(row=start_row_7, column=start_col_2, value=P33)
            if P35 != 0:
                ws.cell(row=start_row_8, column=start_col_2, value=P35)

            O7 = P29 + P32 + P33 + P35
            wb.save(java_remediation_path)

            print('Storing values to database')
            redis_client.rpush(redis_queue_name, '36/40::Storing required values to db!')
            # db_analysis_summary_java = analysis_summary_java(form_project_name, form_application_name, O29, P29, O32, P32, O33, P33, O35, P35, F6, F7, O6, O7)
            # db.session.add(db_analysis_summary_java)
            # try:
            #     db.session.commit()
            # except Exception as e:
            #     db.session.rollback()

            print('Analysis Summary')

            # Analysis Summary : java_high_level_report
            wb = load_workbook(java_high_level_path)
            ws = wb['Analysis Summary']
            start_col_1 = 15
            start_row_1 = 29
            start_row_2 = 32
            start_row_3 = 33
            start_row_4 = 35
            start_col_2 = 16
            start_row_5 = 29
            start_row_6 = 32
            start_row_7 = 33
            start_row_8 = 35
            if O29 != 0:
                ws.cell(row=start_row_1, column=start_col_1, value=O29)
            if O32 != 0:
                ws.cell(row=start_row_2, column=start_col_1, value=O32)
            if O33 != 0:
                ws.cell(row=start_row_3, column=start_col_1, value=O33)
            if O35 != 0:
                ws.cell(row=start_row_4, column=start_col_1, value=O35)
            if P29 != 0:
                ws.cell(row=start_row_5, column=start_col_2, value=P29)
            if P32 != 0:
                ws.cell(row=start_row_6, column=start_col_2, value=P32)
            if P33 != 0:
                ws.cell(row=start_row_7, column=start_col_2, value=P33)
            if P35 != 0:
                ws.cell(row=start_row_8, column=start_col_2, value=P35)

            wb.save(java_high_level_path)

            print("Sheet : Source Code Impacted : CREATED")
            print("End 16 : Source Code Impacted\n")
            redis_client.rpush(redis_queue_name, '37/40::Source Code Impacted Sheet Added')

            print("Start 17 : Other inventory")
            # Convert check_files.out
            check_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/check_files.out'
            other_inventory_df1 = None
            if os.path.isfile(check_files_dir_path) and os.path.getsize(check_files_dir_path) > 0:
                other_inventory_df1 = pd.read_csv(check_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                other_inventory_df1.index = other_inventory_df1.index + 1
                print(other_inventory_df1)

            # Convert exclude_files.out
            exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/exclude_files.out'
            other_inventory_df2 = None
            if os.path.isfile(exclude_files_dir_path) and os.path.getsize(exclude_files_dir_path) > 0:
                other_inventory_df2 = pd.read_csv(exclude_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                other_inventory_df2.index = other_inventory_df2.index + 1
                print(other_inventory_df2)

            if other_inventory_df1 is not None and other_inventory_df2 is not None:
                other_inventory_combined_df = pd.concat([other_inventory_df1, other_inventory_df2], ignore_index=True)
                other_inventory_combined_df = other_inventory_combined_df.drop_duplicates()
                print(other_inventory_combined_df)

                writer = pd.ExcelWriter(java_high_level_path, engine='openpyxl', mode='a')
                book = load_workbook(java_high_level_path)
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
                writer.save()
                print('java_high_level other inventory added')

                writer = pd.ExcelWriter(java_inventory_path, engine='openpyxl', mode='a')
                book = load_workbook(java_inventory_path)
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
                writer.save()
                print('java_inventory other inventory added')

                writer = pd.ExcelWriter(java_remediation_path, engine='openpyxl', mode='a')
                book = load_workbook(java_remediation_path)
                writer.book = book
                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
                writer.save()
                print('java_remediation other inventory added')

            print("Sheet : Other Inventory : CREATED")
            print("End 17 : Other inventory\n")
            redis_client.rpush(redis_queue_name, '38/40::Other Inventory Sheet Added')
            redis_client.rpush(redis_queue_name, '39/40::Code Assessment Successful')
            redis_client.rpush(redis_queue_name, '40/40::CODE ASSESSMENT ENDED')

        # elif(tool_analysis_type == 'canalysis'):
        #     # Old C Report
        #     yyyy_mmdd = yMD[:4] + '-' + yMD[4:]
        #     # Convert stk-YYYY-MMDD-detail.log
        #     stk_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/stk/' + yyyy_mmdd + '/stk-' + yyyy_mmdd + '-detail.log'
        #     if os.path.isfile(stk_dir_path) and os.path.getsize(stk_dir_path) > 0:
        #         df = pd.read_csv(stk_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        #         df.index = df.index + 1
        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'

        #         generate_workbook(df, save_location,'STK Analysis Detail')

        #     # Convert 64bit_C_reportlog
        #     bitC_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/checkanalysis/64bit_C_reportlog'
        #     if os.path.isfile(bitC_dir_path) and os.path.getsize(bitC_dir_path) > 0:
        #         df = pd.read_csv(bitC_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        #         df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE', 'Column4', 'Column5']
        #         df = df.drop(columns=['Column4', 'Column5'])
        #         df.index = df.index + 1
        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'

        #         generate_workbook(df, save_location,'64-bit Analysis Detail')

        #     # Convert endian_C_reportlog
        #     endian_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/checkanalysis/endian_C_reportlog'
        #     if os.path.isfile(endian_dir_path) and os.path.getsize(endian_dir_path) > 0:
        #         df = pd.read_csv(endian_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        #         df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
        #         df.index = df.index + 1
        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'

        #         generate_workbook(df, save_location,'Endianness Analysis Detail')

        #     # Convert char-array.out
        #     charArray_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/charinv/' + yyyy_mmdd + '/char-array.out'
        #     if os.path.isfile(charArray_dir_path) and os.path.getsize(charArray_dir_path) > 0:
        #         df = pd.read_csv(charArray_dir_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        #         df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
        #         df.index = df.index + 1
        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'

        #         generate_workbook(df, save_location,'String-Mem Analysis Detail')

        #     # Convert Log_Envcheck_CFiles_YYYY-MMDD.log
        #     envCheck_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Envcheck_CFiles_' + yyyy_mmdd + '.log'
        #     if os.path.isfile(envCheck_dir_path) and os.path.getsize(envCheck_dir_path) > 0:
        #         df = pd.read_csv(envCheck_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        #         df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
        #         df.index = df.index + 1
        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'

        #         generate_workbook(df, save_location,'ExecutionEnv Analysis Detail')


        #     #Source Code Inventory
        #     regex = r'^.*$'
        #     source_code_c_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
        #     if os.path.isfile(source_code_c_path) and os.path.getsize(source_code_c_path) > 0:
        #         with open(source_code_c_path, 'r') as f:
        #             data = f.read()
        #         lines = data.split('\n')
        #         log_entries = []
        #         for line in lines:
        #             match = re.match(regex, line)
        #             if match:
        #                 log_entries.append(match.group())
        #         df = pd.DataFrame(log_entries, columns=['log_message'])

        #         start_str = '###########  Execution 0001_stepcounter Starts  ###########'
        #         end_str = '###########  Execution 0001_stepcounter Ends ###########'
        #         start_df = df[df['log_message'].str.contains(start_str)]
        #         end_df = df[df['log_message'].str.contains(end_str)]

        #         start_index = start_df.index[0]
        #         end_index = end_df.index[0]
        #         extracted_df = df.iloc[start_index+1:end_index]

        #         result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
        #         result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
        #         result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
        #         result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
        #         result_df = result_df.drop(['log_message', 'col2'], axis=1)
        #         result_df = result_df.reset_index(drop=True)
        #         result_df.index += 1

        #         save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report_old.xlsx'
        #         generate_workbook(result_df, save_location,'Source Code Inventory')

        elif(tool_analysis_type == 'shellanalysis'):
        
            yyyy_mmdd = yMD[:4] + '-' + yMD[4:]
            redis_client.rpush(redis_queue_name, '20/40::Converting logs to reports')

            # Convert ksh_env_common.txt
            ksh_env_common_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/env/' + yyyy_mmdd + '/ksh_env_common.txt'
            if os.path.isfile(ksh_env_common_path) and os.path.getsize(ksh_env_common_path) > 0:
                df = pd.read_csv(ksh_env_common_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                df = df.drop(columns=['Column1'])
                df.index = df.index + 1
                save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                generate_workbook(df, save_location,'OS Specific ENV Var Details')

            # Convert ksh_filepath_common.txt
            ksh_filepath_common_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/env/' + yyyy_mmdd + '/ksh_filepath_common.txt'
            if os.path.isfile(ksh_filepath_common_path) and os.path.getsize(ksh_filepath_common_path) > 0:
                df = pd.read_csv(ksh_filepath_common_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                df = df.drop(columns=['Column1'])
                df.index = df.index + 1
                save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                generate_workbook(df, save_location,'OS Specific File Path Details')

            # Convert ksh_parsed_uniq.out
            ksh_parsed_uniq_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/parse/' + yyyy_mmdd + '/ksh_parsed_uniq.out'
            if os.path.isfile(ksh_parsed_uniq_path) and os.path.getsize(ksh_parsed_uniq_path) > 0:
                df = pd.read_csv(ksh_parsed_uniq_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Column1', 'Column2', 'COMMAND', 'Column4', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                df = df.drop(columns=['Column1', 'Column2', 'Column4'])
                df.index = df.index + 1
                save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                generate_workbook(df, save_location,'Command and option Details')


            # Convert ksh_heredocument.txt
            ksh_heredocument_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/syntax/' + yyyy_mmdd + '/ksh_heredocument.txt'
            if os.path.isfile(ksh_heredocument_path) and os.path.getsize(ksh_heredocument_path) > 0:
                df = pd.read_csv(ksh_heredocument_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                df = df.drop(columns=['Column1'])
                df.index = df.index + 1
                save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                generate_workbook(df, save_location,'Heredoc Analysis')
            
            redis_client.rpush(redis_queue_name, '25/40::Logs converted to reports')
            redis_client.rpush(redis_queue_name, '30/40::Generating Inventory report')

            # Source Code Inventory
            regex = r'^.*$'
            source_code_shell_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
            if os.path.isfile(source_code_shell_path) and os.path.getsize(source_code_shell_path) > 0:
                with open(source_code_shell_path, 'r') as f:
                    data = f.read()
                lines = data.split('\n')
                log_entries = []
                for line in lines:
                    match = re.match(regex, line)
                    if match:
                        log_entries.append(match.group())
                df = pd.DataFrame(log_entries, columns=['log_message'])

                start_str = '###########  Execution 0001_stepcounter Starts  ###########'
                end_str = '###########  Execution 0001_stepcounter Ends ###########'
                start_df = df[df['log_message'].str.contains(start_str)]
                end_df = df[df['log_message'].str.contains(end_str)]

                start_index = start_df.index[0]
                end_index = end_df.index[0]
                extracted_df = df.iloc[start_index+1:end_index]

                result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
                result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
                result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
                result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
                result_df = result_df.drop(['log_message', 'col2'], axis=1)
                result_df = result_df.reset_index(drop=True)
                result_df.index += 1

                shell_report_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                generate_workbook(result_df, shell_report_path,'Source Code Inventory')

                shell_inventory_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_inventory_report.xlsx'

                sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, shell_report_path, shell_inventory_path)
                redis_client.rpush(redis_queue_name, '35/40::All reports ready to be published')

                redis_client.rpush(redis_queue_name, '40/40::CODE ASSESSMENT SUCCESSFUL')

            # # Convert ksh_shebang.txt
            # ksh_shebang_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/shebang/' + yyyy_mmdd + '/ksh_shebang.txt'
            # if os.path.isfile(ksh_shebang_path) and os.path.getsize(ksh_shebang_path) > 0:
            #     with open(ksh_shebang_path, 'r') as f:
            #         lines = f.readlines()
            #     data = []
            #     prev_column1 = None
            #     for line in lines:
            #         split_line = line.strip().split('#!')
            #         if split_line[0].strip():
            #             prev_column1 = split_line[0].strip()
            #         data.append([prev_column1, '#!' + split_line[1].strip() if len(split_line) > 1 else ''])
            #     df = pd.DataFrame(data, columns=['column1', 'column2']).dropna()
            #     df['column2'] = df.groupby(df.index // 2)['column2'].transform(lambda x: x.iloc[1])
            #     df.columns = ['FILE NAME', 'SOURCE CODE']
            #     df.index = df.index + 1
            #     df = df.drop(df[df.index % 2 == 0].index).reset_index(drop=True)

            #     save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

            #     generate_workbook(df, save_location,'Shebang Line Analysis')
    except Exception as e:
        return jsonify({'error': str(e)})

    # except Exception as e:
    #     return jsonify({'message': 'analysis failed', 'error': str(e)}), 401

    return True

@celery.task
def testing_tasks(message):

    logger.info('STARTED Execution')
    redis_client.rpush(redis_queue_name, 'Connection 1 established successfully!!')
    time.sleep(30)
    redis_client.rpush(redis_queue_name, 'Connection 2 established successfully!!')
    return True

def run_cppcheck(source_path):
    command = ["cppcheck", "--force", source_path]
    result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    return result.stdout.decode(), result.stderr.decode()

def process_content(lines):
    modified_lines = lines[0]
    print(modified_lines)
    return modified_lines

def convert_to_dataframe(package_names, package_descriptions, version, package_name):
    data = {'Deprecated Item Name': [], 'Deprecation Comment': []}

    for item_name, comment in zip(package_names, package_descriptions):
        data['Deprecated Item Name'].append(item_name)
        
        cleaned_comment = " ".join(comment.split())
        data['Deprecation Comment'].append(cleaned_comment)

    # df = pd.DataFrame(data)
    df = pd.DataFrame(data).reset_index(drop=True)

    print(df)
    df[['Package', 'Class']] = df['Deprecated Item Name'].str.rsplit('.', n=1, expand=True)
    df = df[['Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']]

    df.insert(0, 'Custom ID', f'{package_name}' + version + '_' + df.index.map(lambda x: f'{x + 1:04d}'))
    df.columns = ['Custom ID', 'Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']
    df.columns = ['RuleId', 'Deprecated List', 'Library to Search', 'Entity to Search', 'Remedy']
    print(df)
    return df

def convert_to_dataframe_move_to_excel(package_name, package_description, version, package):
                
    data = {'Deprecated Item Name': [], 'Deprecation Comment': []}

    for deprecated_item, deprecation_comment in zip(package_name, package_description):
        item_name = deprecated_item.text.strip()
        comment = deprecation_comment.text.strip()
        comment = deprecation_comment.text.strip().replace('\t', ' ')
        comment = deprecation_comment.text.strip().replace('\n', ' ')

        comment = re.sub(r'\s+', ' ', deprecation_comment.text.strip())

        if(comment == ''):
            comment = openai(item_name, version)

        data['Deprecated Item Name'].append(item_name)
        data['Deprecation Comment'].append(comment)

    df = pd.DataFrame(data).reset_index(drop=True)
    df[['Package', 'Class']] = df['Deprecated Item Name'].str.rsplit('.', n=1, expand=True)
    df = df[['Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']]

    df.insert(0, 'Custom ID', f'{package}'+ '_' + version + '_' + df.index.map(lambda x: f'{x + 1:04d}'))
    df.columns = ['Custom ID', 'Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']
    df.columns = ['RuleId', 'Deprecated List', 'Library to Search', 'Entity to Search', 'Remedy']

    return df

def createDf(deprecated_item, version, package_name):
    df = pd.DataFrame({'Fullname': deprecated_item}).reset_index(drop=True)
    df[['Deprecated Item Name', 'Deprecation Comment']] = df['Fullname'].str.split('\n', 1, expand=True)
    df.drop('Fullname', axis=1, inplace=True)
    df[['Package', 'Class']] = df['Deprecated Item Name'].str.rsplit('.', n=1, expand=True)
    df['Deprecation Comment'] = df['Deprecation Comment'].str.replace(r'[\t\n]+', ' ')
    df['Deprecation Comment'] = df['Deprecation Comment'].str.replace(r'\s+', ' ')
    df = df[['Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']]
    
    df.insert(0, 'Custom ID', f'{package_name}' + '_' + version + '_' + df.index.map(lambda x: f'{x + 1:04d}'))
    df.columns = ['Custom ID', 'Deprecated Item Name', 'Package', 'Class', 'Deprecation Comment']
    df.columns = ['RuleId', 'Deprecated List', 'Library to Search', 'Entity to Search', 'Remedy']
    return df

# def java_doc_sheet(file_path, sheet_name, df):
#     book = load_workbook(file_path)
#     writer = pd.ExcelWriter(file_path, engine='openpyxl')
#     writer.book = book
#     sheet_name = sheet_name
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#     df.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=0, index=False)
#     writer.save()

def java_doc_sheet(file_path, sheet_name, df):
    book = load_workbook(file_path)
    if sheet_name not in book.sheetnames:
        book.create_sheet(sheet_name)
    sheet = book[sheet_name]
    data_to_append = [df.columns.tolist()] + df.values.tolist()
    next_row = sheet.max_row + 1
    for row in data_to_append:
        sheet.append(row)
    book.save(file_path) 

def extract_versions(package_name):
    package_versions = {
        'Mchange commons java': ['0.2.19', '0.2.20'],
        'Apache Tomcat': ['11.0', '10.0','9.0','8.0','7.0'],
        'Geronimo': ['2.0.1', '2.1.3'],
        'Apache OpenJPA': ['1.0.0','1.0.1','1.0.2','1.0.3','1.0.4','1.1.0','1.2.0','1.2.1','1.2.2','1.2.3','2.0.0','2.0.1','2.1.0','2.1.1','2.2.0','2.2.1','2.2.2','2.3.0','2.4.0','2.4.1','2.4.2','2.4.3','3.0.0','3.1.0','3.1.2','3.2.0','3.2.1','3.2.2'],
        'POI': ['3.17','4.0','4.1','5.0'],
        'Quartz scheduler': ['2.5.0-SNAPSHOT','2.3.1-SNAPSHOT','2.3.0','2.2.2','2.1.7','2.0.2','1.8.6'],
        'XMLBeans': ['1.0.4','2.0.0.','2.1.0','2.2.0','2.4.0','2.6.0','3.0.0','3.1.0','4.0.0','5.0.0']
    }
    return package_versions[package_name]      

def deprecated_data_sheet(url, package_name, version):
    try:
        response = requests.get(url, verify=False)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        deprecated_table = None
        deprecated_item_names = []
        deprecation_comments = []
        deprecated_table = soup.find_all('td', class_='colOne')
        if not deprecated_table:
            deprecated_table = soup.find('table', {'border': '1', 'width': '100%', 'cellpadding': '3', 'cellspacing': '0'})
            if not deprecated_table:
                deprecated_table = soup.find('table', class_='summary-table')
                if not deprecated_table:
                    deprecated_table = soup.find('table', class_='deprecatedSummary')
                    if not deprecated_table:
                        deprecated_table = soup.find('div', class_='two-column-summary')
                        if not deprecated_table:
                            return f'Error: Unable to find deprecated classes table on {url}'
                        else:
                            interface_names = deprecated_table.find_all('div', class_='col-summary-item-name')
                            deprecation_comments = deprecated_table.find_all('div', class_='col-last')

                            deprecated_item_names = [item_name.find('a').text.strip() for item_name in interface_names]
                            deprecation_comments = [comment.text.strip().replace('\t', ' ') for comment in deprecation_comments]
                    else:
                        deprecated_items = deprecated_table.find_all('tr', class_='altColor' or 'rowColor')
                        for deprecated_item in deprecated_items:
                            item_name = deprecated_item.find('a').text.strip()

                            comment_div = deprecated_item.find('div', class_='deprecationComment')
                            comment = comment_div.text.strip().replace('\t', ' ') if comment_div else ''
                            # comment = comment_div.text.strip() if comment_div else ''

                            deprecated_item_names.append(item_name)
                            deprecation_comments.append(comment)
                else:
                    deprecated_items = deprecated_table.find_all('tr', class_='alt-color' or 'row-color')
                    for deprecated_item in deprecated_items:
                        item_name = deprecated_item.find('a').text.strip()

                        comment_div = deprecated_item.find('div', class_='deprecation-comment')
                        comment = comment_div.text.strip().replace('\t', ' ') if comment_div else ''
                        # comment = comment_div.text.strip() if comment_div else ''

                        deprecated_item_names.append(item_name)
                        deprecation_comments.append(comment)
            else:
                deprecated_items = deprecated_table.find_all('tr', class_='TableRowColor')
                for deprecated_item in deprecated_items:
                    item_name = deprecated_item.find('a').text.strip()
                
                    comment = deprecated_item.find('i').text.strip().replace('\t', ' ') if deprecated_item.find('i') else ''

                    deprecated_item_names.append(item_name)
                    deprecation_comments.append(comment)
        else:
            for deprecated_item in deprecated_table:
                item_name = deprecated_item.find('a').text.strip()

                comment_div = deprecated_item.find_next('div', class_='block')
                comment = comment_div.text.strip().replace('\t', ' ') if comment_div else ''
                # comment = comment_div.text.strip() if comment_div else ''

                deprecated_item_names.append(item_name)
                deprecation_comments.append(comment)

        df = convert_to_dataframe(deprecated_item_names, deprecation_comments, version, package_name)
        df = df.reset_index(drop=True)

        output_folder = 'u2l_total_deprecated_data'
        os.makedirs(output_folder, exist_ok=True)
        # file_path = os.path.join(output_folder, 'total_deprecated_data.xlsx')

        file_path = os.path.join(output_folder, f'{package_name}_deprecated_data.xlsx')
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='DeprecatedItems', startcol=0, startrow=0, index=False)

        # if os.path.exists(file_path):
        #     # Check if the sheet exists in the file
        #     with pd.ExcelFile(file_path) as xls:
        #         sheet_exists = 'total_deprecated_data' in xls.sheet_names

        #     if sheet_exists:
        #         existing_data = pd.read_excel(file_path, sheet_name='total_deprecated_data')
        #         # Concatenate existing data with the new data
        #         combined_data = pd.concat([existing_data, df], ignore_index=True)
        #         # Write the combined data back to the sheet
        #         with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        #             combined_data.to_excel(writer, sheet_name='total_deprecated_data', startrow=len(existing_data), index=False)
        #     else:
        #         # If the sheet doesn't exist, write the new data to it
        #         with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
        #             df.to_excel(writer, sheet_name='total_deprecated_data', startrow=0, index=False)
        # else:
        #     with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        #         df.to_excel(writer, sheet_name='total_deprecated_data', startrow=0, index=False)    

        return f'Data collected and saved successfully from {url}'  

    except requests.exceptions.RequestException as e:
        return f'Error: {str(e)} while processing {url}'

def generate_custom_id(version, index):
    return f'java_{version:02d}_{index + 1:04d}'

def openai(package_name, version):
    
    import openai

    openai.api_type = "azure"
    openai.api_base = "https://azure-openai-demo1.openai.azure.com/"
    openai.api_version = "2023-07-01-preview"
    openai.api_key = "c2c49c25ad2946b98d9e7ca876f7f7cb"

    prompt = "Give me one liner deprecation reason for " + package_name + " present in java docs version " + version

    message_text = [
    {"role": "system", "content": "You are an expert middleware architect."},
    {"role": "user", "content": prompt}
    ]
    completion = openai.ChatCompletion.create(
        engine="gpt-4-32k",
        messages = message_text,
        temperature=0.0,
        max_tokens=800,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None
        )
    result = completion.choices[0].message['content']
    print(result)
    return result

def insert_middleware_component_to_db(key, middleware_component, stack, stack_current_version, stack_upgraded_version, recommendation):
    logger.info(f"Inserting middleware component with key: {key}")
    middleware_component_entry = MiddlewareComponent.query.filter_by(key=key).first()
    if middleware_component_entry:
        logger.warning("Middleware component already exists. Returning existing recommendation.")
        return middleware_component_entry.recommendation
    new_middleware_component = MiddlewareComponent(
        key=key,
        middleware_component=middleware_component,
        stack=stack,
        stack_current_version=stack_current_version,
        stack_upgraded_version=stack_upgraded_version,
        recommendation=recommendation
    )
    db.session.add(new_middleware_component)
    db.session.commit()
    logger.info("Middleware component inserted successfully.")
    return recommendation

def get_recommendation(middleware_components, stack, stack_current_version, stack_upgraded_version):

    import openai

    print(f"Prompting IN OPENAI GPT!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
    openai.api_type = "azure"
    openai.api_base = "https://azure-openai-demo1.openai.azure.com/"
    openai.api_version = "2023-07-01-preview"
    openai.api_key = "c2c49c25ad2946b98d9e7ca876f7f7cb"
    recommendations = {}

    for component, details in middleware_components.items():
        lines = details.get("lines", [])
        message_text = [
            {"role": "system", "content": "You are an expert middleware architect."},
            {"role": "user", "content": f"I have the middleware component {component} and now I want to upgrade the stack {stack} from {stack_current_version} to {stack_upgraded_version}. Please provide recommendations, configuration, and version for the middleware component."},
        ]
        completion = openai.ChatCompletion.create(
            engine="gpt-4-32k",
            messages=message_text,
            temperature=0.0,
            max_tokens=800,
            top_p=0.95,
            frequency_penalty=0,
            presence_penalty=0,
            timeout=1200,
            stop=None
        )

        recommendation = completion.choices[0].message['content']
        details["recommendation"] = recommendation 
        recommendations[component] = details

    return recommendations

def java_middleware_extraction(zip_file_path, output_folder, middleware_json_path):

    logging.info(f'Starting middleware extraction for {zip_file_path}')
    def search_middleware(content, patterns, stack_current_version, stack_upgraded_version):
        found_components = {}
        lines = content.split('\n')
        for component, search_patterns in patterns.items():
            line_numbers = []
            for i, line in enumerate(lines, start=1):
                if any(re.search(r'\b{}\b'.format(pattern), line) for pattern in search_patterns):
                    line_numbers.append(i)
            if line_numbers:
                key = f"{stack_current_version}_{component}_{stack_upgraded_version}"
                recommendation = insert_middleware_component_to_db(
                    key=key,
                    middleware_component=component,
                    stack="java",
                    stack_current_version=stack_current_version,
                    stack_upgraded_version=stack_upgraded_version,
                    recommendation=""
                )
                found_components[key] = {
                    "component_name": component,
                    "lines": line_numbers,
                    "recommendation": recommendation
                }
        logger.info(f'Extracting the line number for each middleware component from {zip_file_path}')
        return found_components

    
    def search_in_zip(zip_file_path, patterns, stack_current_version, stack_upgraded_version):
        logging.info(f'Extracting the middleware components from {zip_file_path}')
        results = []

        with zipfile.ZipFile(zip_file_path, 'r') as zip_file:
            for file_info in zip_file.infolist():
                if file_info.filename.endswith('.java'):
                    with zip_file.open(file_info) as java_file:
                        content = java_file.read().decode('utf-8')
                        found_components = search_middleware(content, patterns, stack_current_version, stack_upgraded_version)
                        for key, details in found_components.items():
                            recommendation = details.get("recommendation", None)
                            if (recommendation == ''):
                                recommendation = get_recommendation(
                                    {key: details},
                                    stack="java",
                                    stack_current_version=stack_current_version,
                                    stack_upgraded_version=stack_upgraded_version,
                                ).get(key, {}).get("recommendation", "")
                                insert_middleware_component_to_db(
                                    key=key,
                                    middleware_component=details["component_name"],
                                    stack="java",
                                    stack_current_version=stack_current_version,
                                    stack_upgraded_version=stack_upgraded_version,
                                    recommendation=recommendation
                                )
                            details["recommendation"] = recommendation

                        result = {
                            "file_name": file_info.filename,
                            "middleware_components": found_components,
                        }
                        results.append(result)
                        create_dependencies_json(zip_file_path, result, output_folder)

        return results

    def create_dependencies_json(zip_file_path, result, output_folder):
        logging.info('Writing the middleware components into a json')
        zip_file_name = os.path.basename(zip_file_path)
        json_file_name = os.path.splitext(zip_file_name)[0] + "_middleware.json"
        json_file_path = os.path.join(output_folder, json_file_name)
        if os.path.exists(json_file_path):
            with open(json_file_path, 'r') as json_file:
                try:
                    existing_data = json.load(json_file)
                except json.decoder.JSONDecodeError:
                    print("Existing data is not a valid JSON. Creating a new dictionary.")
                    existing_data = []
            existing_data.append(result)
            with open(json_file_path, 'w') as json_file:
                json.dump(existing_data, json_file, indent=2)
            convert_json_to_excel(json_file_path, zip_file_name, output_folder)
        else:
            with open(json_file_path, 'w') as json_file:
                json.dump([result], json_file, indent=2)

    def convert_json_to_excel(json_file_path, zip_file_name, output_folder):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        with open(json_file_path, 'r') as json_file:
            data = json.load(json_file)
        sheet.cell(row=1, column=1, value="File Name")
        sheet.cell(row=1, column=2, value="Middleware Component")
        sheet.cell(row=1, column=3, value="Lines")
        sheet.cell(row=1, column=4, value="Recommendation")
        for row, item in enumerate(data, start=2):
            file_name = item['file_name']
            for component, details in item['middleware_components'].items():
                lines = ', '.join(map(str, details.get('lines', [])))
                recommendation = details.get('recommendation', '')
                sheet.cell(row=row, column=1, value=file_name)
                sheet.cell(row=row, column=2, value=component)
                sheet.cell(row=row, column=3, value=lines)
                sheet.cell(row=row, column=4, value=recommendation)
                row += 1
        excel_file_name = os.path.splitext(zip_file_name)[0] + "_middleware.xlsx"
        excel_file_path = os.path.join(output_folder, excel_file_name)
        workbook.save(excel_file_path)
        logging.info(f'Exporting the JSON into an excel file {excel_file_name}')
    with open(middleware_json_path, 'r') as middleware_file:
        middleware_patterns = json.load(middleware_file)
    results = search_in_zip(zip_file_path, middleware_patterns, stack_current_version="11", stack_upgraded_version="17")

    if results:
        print(f"Results found. JSON and Excel files created for each middleware component.")
        return 'success'
    else:
        print("No results found.")