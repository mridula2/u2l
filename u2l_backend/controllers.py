from flask import Blueprint, jsonify, request, Response, Flask, stream_with_context
from utils import validate_email, generate_workbook, sheet_transfer, fetch_store_java_data, send_email, summary_sheet_generation, testing_task, run_cppcheck, get_task_queue_name, convert_to_dataframe , convert_to_dataframe_move_to_excel, createDf, java_doc_sheet, extract_versions, deprecated_data_sheet, insert_middleware_component_to_db, get_recommendation, java_middleware_extraction, populate_database
from models import *
import datetime
import os
import stat
import subprocess
import shutil   
import pandas as pd
from openpyxl import load_workbook
import re
import numpy as np
import zipfile
import openpyxl
import logging
from celery_setup import celery
import requests
from bs4 import BeautifulSoup

from flask_jwt_extended import create_access_token #, get_jwt_claims
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt

import json
import time
from celery.result import AsyncResult

from flask_sse import sse
import redis

redis_client = redis.StrictRedis(host='redis', port=6379, db=0)

logging.basicConfig(filename='middleware_extraction.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

authentication_controller = Blueprint('authentication', __name__)
analysis_bp = Blueprint('analysis', __name__)

tool_analysis_type = ''
PJHOME = ''
APNAME = ''
file_name = ''
yMD = ''
hMS = ''
hM = ''
db_username = ''
db_password_encrypt = ''

@authentication_controller.route('/signup', methods=['POST'])
def signup():
    json_data = request.get_json()
    form_email = json_data['email']
    form_password = json_data['password']
    form_first_name = json_data['first_name']
    form_last_name = json_data['last_name']
    form_role = json_data['user_role']
    roles = ["delivery", "pursuit", "customer", "admin"]

    if form_role.lower() not in roles:
      error_message = {'error' :"Invalid role type"}
      return jsonify(error_message), 404

    if form_email and validate_email(form_email):
        user_data_validate = user_details.query.filter_by(email=form_email).first()
        if user_data_validate:
            return jsonify({'message': 'Email Already Exsists !!'}), 401
        else:
            cur_date = datetime.datetime.now()
            new_user = user_details(form_email, form_password, form_first_name, form_last_name, form_role, cur_date, cur_date)
            db.session.add(new_user)
            db.session.commit()
            return jsonify({'message': 'SignUp Successfull'}), 200
    else:
        return jsonify({'message': 'Invalid Email Id!'}), 401

@authentication_controller.route('/login', methods=['POST'])
def authentication():
    json_data = request.get_json()
    form_email = json_data['email']
    form_password = json_data['password']

    if form_email and validate_email(form_email):
        user_data_validate = user_details.query.filter_by(email=form_email).first()
        print(user_data_validate)
        if user_data_validate:
            if user_data_validate.password == form_password:
                additional_claims = {
                    "id": user_data_validate.id,
                    "email": user_data_validate.email,
                    "first_name": user_data_validate.first_name,
                    "last_name": user_data_validate.last_name,
                    "user_role": user_data_validate.user_role
                }
            access_token = create_access_token(identity= user_data_validate.id, additional_claims= additional_claims)
            return jsonify({'token': access_token,'message': 'authentication success'}), 200
        else:
            return jsonify({'message': 'authentication failed'}), 401
    else:
        return jsonify({'message': 'Invalid Email Id!'}), 401

@authentication_controller.route('/jwt_create_token', methods=['POST'])
def jwt_create_token():
    user_data_validate = user_details.query.filter_by(email='admin@hpe.com', password='UGFzc3dvcmRAMTIz').first()
    additional_claims = {
        "id": user_data_validate.id,
        "email": user_data_validate.email,
        "first_name": user_data_validate.first_name,
        "last_name": user_data_validate.last_name,
        "user_role": user_data_validate.user_role
    }
    access_token = create_access_token(identity= user_data_validate.id, additional_claims= additional_claims)
    return jsonify({'token': access_token,'message': 'authentication success' }), 200

@authentication_controller.route('/jwt_access_token', methods=['POST'])
@jwt_required()
def jwt_access_token():
    current_user_id = get_jwt_identity()
    print("using jwt : current_user_id : " + str(current_user_id))
    user = user_details.query.get(current_user_id)
    print("using jwt : user : " + str(user))
    user_data = {
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "user_role": user.user_role
    }

    user_data_jwt = get_jwt()
    user_data_claims = {
        "id": user_data_jwt.get('id', 'default_id'),
        "email": user_data_jwt.get('email', 'default_name'),
        "first_name": user_data_jwt.get('first_name', 'default_first_name'),
        "last_name": user_data_jwt.get('last_name', 'default_last_name'),
        "user_role": user_data_jwt.get('user_role', 'default_user_role')
    }
    return jsonify(user_data = user_data,user_data_claims = user_data_claims, jwt_claims = user_data_jwt), 200

@authentication_controller.route('/jwt_access_token2', methods=['POST'])
@jwt_required()
def jwt_access_token2():
    current_user_id = get_jwt_identity()
    print("using jwt : current_user_id : " + str(current_user_id))
    user = user_details.query.get(current_user_id)
    print("using jwt : user : " + str(user))
    user_data = {
        "id": user.id,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "user_role": user.user_role
    }
    return jsonify(user_data = user_data), 200

@authentication_controller.route('/analysis', methods=['POST'])
@jwt_required()
def upload():
    
    print('Analysis Started !!')
    file = request.files['file_name']

    form_project_name = request.form['project_name']
    form_application_name = request.form['application_name']
    form_project_client = request.form['project_client']
    form_project_manager = request.form['project_manager']
    global tool_analysis_type
    tool_analysis_type = request.form['analysis_type']

    form_source_os = request.form['source_os']
    form_source_os_version = request.form['source_os_version']
    form_target_os = request.form['target_os']
    form_target_os_version = request.form['target_os_version']
    c_analysis_types = ['C', 'C++', 'Pro*C', 'C/Pro*C', 'C++/Pro*C']
    # framework_type = request.form['framework']
    framework_type = ''

    form_source_jdk = ''
    form_target_jdk = ''
    form_source_jsp = ''
    form_target_jsp = ''
    form_source_servlet = ''
    form_target_servlet = ''

    form_source_pre_compiler = ''
    form_target_pre_compiler = ''
    form_source_pre_compiler_version = ''
    form_target_pre_compiler_version = ''
    form_source_compiler = ''
    form_target_compiler = ''
    form_source_compiler_version = ''
    form_target_compiler_version = ''

    form_source_compiler = ''
    form_target_compiler = ''
    form_source_compiler_version = ''
    form_target_compiler_version = ''
    form_source_pre_compiler = ''
    form_target_pre_compiler = ''
    form_source_pre_compiler_version = ''
    form_target_pre_compiler_version = ''

    form_source_pre_compiler = ''
    form_target_pre_compiler = ''
    form_source_pre_compiler_version = ''
    form_target_pre_compiler_version = ''
    form_source_compiler = ''
    form_target_compiler = ''
    form_source_compiler_version = ''
    form_target_compiler_version = ''

    form_source_shell = ''
    form_target_shell = ''
    form_source_shell_version = ''
    form_target_shell_version = ''

    if(tool_analysis_type == 'Java'):
        tool_analysis_type = 'javaanalysis'
        form_source_jdk = request.form['source_jdk']
        form_target_jdk = request.form['target_jdk']
        form_source_jsp = request.form['source_jsp']
        form_target_jsp = request.form['target_jsp']
        form_source_servlet = request.form['source_servlet']
        form_target_servlet = request.form['target_servlet']
        framework_type = request.form['framework']
    elif(tool_analysis_type in c_analysis_types):
        if(tool_analysis_type == 'Pro*C'):
            tool_analysis_type = 'canalysis'
            form_source_pre_compiler = request.form['source_pre_compiler']
            form_target_pre_compiler = request.form['target_pre_compiler']
            form_source_pre_compiler_version = request.form['source_pre_compiler_version']
            form_target_pre_compiler_version = request.form['target_pre_compiler_version']
            form_source_compiler = ''
            form_target_compiler = ''
            form_source_compiler_version = ''
            form_target_compiler_version = ''
        elif(tool_analysis_type == 'C' or tool_analysis_type == 'C++'):
            print('Accepting C Parameters!')
            tool_analysis_type = 'canalysis'
            form_source_compiler = request.form['source_compiler']
            form_target_compiler = request.form['target_compiler']
            form_source_compiler_version = request.form['source_compiler_version']
            form_target_compiler_version = request.form['target_compiler_version']
            form_source_pre_compiler = ''
            form_target_pre_compiler = ''
            form_source_pre_compiler_version = ''
            form_target_pre_compiler_version = ''
            form_middleware_type = request.form['middleware_type']
            print(form_middleware_type)
        elif(tool_analysis_type == 'C/Pro*C' or tool_analysis_type == 'C++/Pro*C'):
            tool_analysis_type = 'canalysis'
            form_source_pre_compiler = request.form['source_pre_compiler']
            form_target_pre_compiler = request.form['target_pre_compiler']
            form_source_pre_compiler_version = request.form['source_pre_compiler_version']
            form_target_pre_compiler_version = request.form['target_pre_compiler_version']  
            form_source_compiler = request.form['source_compiler']
            form_target_compiler = request.form['target_compiler']
            form_source_compiler_version = request.form['source_compiler_version']
            form_target_compiler_version = request.form['target_compiler_version']
    elif(tool_analysis_type == 'Shell'):  
        tool_analysis_type = 'shellanalysis'
        form_source_shell = request.form['source_shell']
        form_target_shell = request.form['target_shell']
        form_source_shell_version = request.form['source_shell_version']
        form_target_shell_version = request.form['target_shell_version']

    analysis_types = ['javaanalysis', 'canalysis', 'shellanalysis']
    if tool_analysis_type not in analysis_types:
        error_message = {'error' :"Invalid analysis type"}
        return jsonify(error_message), 404
    
    cur_date = datetime.datetime.now()

    global file_name
    file_name = file.filename
    temp_tuple = os.path.splitext(file_name)
    if temp_tuple[1] != '.zip':
        error_message = {'error' :"Upload .zip files only"}
        return jsonify(error_message), 404

    file_name = temp_tuple[0]

    year_month_day = str(cur_date).split(" ")
    global yMD
    yMD = year_month_day[0].replace("-", "")

    hour_min_sec = os.path.splitext(year_month_day[1])
    global hMS
    hMS = hour_min_sec[0].replace(":", "")

    file_name = file_name + "_" + yMD + "_" + hMS 
    directory = '/usr/u2l/u2l_backend/projects/' + file_name
    os.makedirs(directory)
    os.chmod(directory, stat.S_IRWXU | stat.S_IRWXG | stat.S_IRWXO) # Giving all read, write, execute permissions
    print('directory created')
    file.save(os.path.join(directory, file_name +'.zip'))

    file_size_bytes = os.stat(directory +'/'+ file_name + '.zip').st_size
    file_size_mb = round(file_size_bytes/(1024), 3)

    # # to copy Java, U2L and report_templates folder
    try:
        source_dir = '/usr/u2l/u2l_backend/'
        source_dir_report = '/usr/u2l/u2l_backend/report_templates/'
        dest_dir = directory
        source_path = os.path.join(source_dir, 'U2L')
        dest_path = os.path.join(dest_dir, 'U2L')
        shutil.copytree(source_path, dest_path)  
        
        if(tool_analysis_type == 'javaanalysis'):
            source_path = os.path.join(source_dir, 'Java')
            dest_path = os.path.join(dest_dir, 'Java')
            shutil.copytree(source_path, dest_path)

            source_path = os.path.join(source_dir, 'pmd-bin-6.55.0')
            dest_path = os.path.join(dest_dir, 'pmd-bin-6.55.0')
            shutil.copytree(source_path, dest_path)

            make_directory = dest_dir + '/pmd-bin-6.55.0/Reports_1'
            os.makedirs(make_directory)

            source_path = os.path.join(source_dir_report, 'java_report.xlsx')
            dest_path = os.path.join(dest_dir, 'java_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'java_inventory_report.xlsx')
            dest_path = os.path.join(dest_dir, 'java_inventory_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'java_remediation_report.xlsx')
            dest_path = os.path.join(dest_dir, 'java_remediation_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'java_high_level_report.xlsx')
            dest_path = os.path.join(dest_dir, 'java_high_level_report.xlsx')
            shutil.copy(source_path, dest_path)

        elif(tool_analysis_type == 'shellanalysis'):

            source_path = os.path.join(source_dir, 'pmd-bin-6.55.0')
            dest_path = os.path.join(dest_dir, 'pmd-bin-6.55.0')
            shutil.copytree(source_path, dest_path)

            make_directory = dest_dir + '/pmd-bin-6.55.0/Reports_1'
            os.makedirs(make_directory)

            source_path = os.path.join(source_dir_report, 'shell_report.xlsx')
            dest_path = os.path.join(dest_dir, 'shell_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'shell_inventory_report.xlsx')
            dest_path = os.path.join(dest_dir, 'shell_inventory_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'shell_remediation_report.xlsx')
            dest_path = os.path.join(dest_dir, 'shell_remediation_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'shell_high_level_report.xlsx')
            dest_path = os.path.join(dest_dir, 'shell_high_level_report.xlsx')
            shutil.copy(source_path, dest_path)

        elif(tool_analysis_type == 'canalysis'):
            source_path = os.path.join(source_dir_report, 'c_report.xlsx')
            dest_path = os.path.join(dest_dir, 'c_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'c_report_old.xlsx')
            dest_path = os.path.join(dest_dir, 'c_report_old.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'c_inventory_report.xlsx')
            dest_path = os.path.join(dest_dir, 'c_inventory_report.xlsx')
            shutil.copy(source_path, dest_path)

            source_path = os.path.join(source_dir_report, 'c_remediation_report.xlsx')
            dest_path = os.path.join(dest_dir, 'c_remediation_report.xlsx')
            shutil.copy(source_path, dest_path)
    
    except Exception as e:
        response = jsonify({'error': str(e)})
        response.status_code = 500
        return response

    #storing values in db analysis_type table
    db_analysis_type = analysis_type(form_project_name, form_application_name, tool_analysis_type)
    db.session.add(db_analysis_type)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'analysis_type : Project name already exists'}), 409

    #storing values in db os_details table
    db_os_details = os_details(form_project_name, form_application_name, form_source_os, form_source_os_version, form_target_os, form_target_os_version)
    db.session.add(db_os_details)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'os_details : Project name already exists'}), 409
    
    #storing values in db analysis_status table
    analysis_start_time = datetime.datetime.now()
    analysis_end_time = datetime.datetime.now()
    db_analysis_status = analysis_status(form_project_name, form_application_name, db_analysis_type.id, 'admin@hpe.com', 'analysis started', file_name, file_size_mb, analysis_start_time, analysis_end_time)
    db.session.add(db_analysis_status)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'analysis_status : Project name already exists'}), 409

    script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/U2LTool_Install.sh'

    if(tool_analysis_type == 'javaanalysis'):
        db_analysis_java = analysis_java(form_project_name, form_application_name, db_analysis_type.id, form_source_jdk, form_target_jdk, form_source_jsp, form_target_jsp, form_source_servlet, form_target_servlet)
        db.session.add(db_analysis_java)
        db.session.commit()
        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_' + yMD + '_' + hMS
    elif(tool_analysis_type == 'canalysis'):
        db_analysis_c = analysis_c(form_project_name, form_application_name, db_analysis_type.id, form_source_compiler, form_target_compiler, form_source_compiler_version, form_target_compiler_version, form_source_pre_compiler, form_target_pre_compiler, form_source_pre_compiler_version, form_target_pre_compiler_version)
        db.session.add(db_analysis_c)
        db.session.commit()
        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS
    elif(tool_analysis_type == 'shellanalysis'):
        db_analysis_shell = analysis_shell(form_project_name, form_application_name, db_analysis_type.id, form_source_shell, form_target_shell, form_source_shell_version, form_target_shell_version)
        db.session.add(db_analysis_shell)
        db.session.commit()
        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS

    task = testing_task.apply_async(args=[yMD, hMS, tool_analysis_type, framework_type, file_name, script_path, file_path, form_project_name, form_application_name, form_project_client, form_project_manager, form_source_os, form_source_os_version, form_target_os, form_target_os_version, form_source_jdk, form_target_jdk, form_source_jsp, form_target_jsp, form_source_servlet, form_target_servlet, form_source_compiler, form_target_compiler, form_source_compiler_version, form_target_compiler_version, form_source_pre_compiler, form_target_pre_compiler, form_source_pre_compiler_version, form_target_pre_compiler_version, form_source_shell, form_target_shell, form_source_shell_version, form_target_shell_version])

    task_id = task.id
    task_status = task.status
    #storing values in db project_details table
    db_project_details = project_details(form_project_name, form_application_name, task_id, 'admin@hpe.com', 1, form_project_client, form_project_manager, file_name, file_size_mb, task_status, cur_date)
    db.session.add(db_project_details)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(e)
        return jsonify({'message': 'project_details : Project name already exists'}), 409

    print('Async call in progress !!')

    return jsonify({'message': 'Request processing started', 'task_id': task.id}), 202

@authentication_controller.route('/stream_logs/<task_id>', methods=['GET'])
def stream_logs(task_id):
    query = project_details.query.filter_by(task_id=task_id)
    stream_task_status = query.first().analysis_status  # {'PENDING', 'SUCCESS'}

    total_logs = celery_logs.query.filter_by(task_id=task_id).count()
    redis_queue_name = get_task_queue_name(task_id)  # {'logs : {task_id}'}

    def event_stream():
        if stream_task_status == 'SUCCESS':
            print('Status: SUCCESS')
            logs = celery_logs.query.filter_by(task_id=task_id).all()
            log_message = []
            for log in logs:
                log = log.task_log
                log_message.append(log)
            x = {
                "logs" : log_message,
                "progress" : 100
            }
            y = json.dumps(x)
            yield f"data: {y}\n\n"

        while stream_task_status == 'PENDING':
            print('Status: PENDING')
            log_message = redis_client.lpop(redis_queue_name)
            if log_message:
                progress_info, message = log_message.decode('utf-8').split('::', 1)
                progress = int(progress_info.split('/')[0])
                total_progress = int(progress_info.split('/')[1])
                print(progress)
                print(total_progress)
                print(message)

                db_celery_logs = celery_logs(task_id, stream_task_status, message)
                db.session.add(db_celery_logs)
                try:
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    return jsonify({'message': 'Celery logs db error'}), 409
                
                # prog = (progress / total_progress) * 100
                # print('prog={}'.format(prog))

                # x = {
                #     "logs" : message,
                #     "progress" : prog
                # }
                # y = json.dumps(x)
                # yield f"data: {y}\n\n"

                logs = celery_logs.query.filter_by(task_id=task_id).all()
                log_message = []
                for log in logs:
                    log = log.task_log
                    log_message.append(log)
                prog = (progress / total_progress) * 100    
                x = {
                    "logs" : log_message,
                    "progress" : prog
                }
                y = json.dumps(x)
                yield f"data: {y}\n\n"
            time.sleep(20)

        print('event end !')
        yield "event: end\n\n"

        # response = Response(stream_with_context(event_stream()), content_type='text/event-stream')
        # response.headers['X-Accel-Buffering'] = 'no'
        # return response
    return Response(stream_with_context(event_stream()), content_type='text/event-stream')

@authentication_controller.route('/stream_logs1/<task_id>', methods=['GET'])
def stream_logs1(task_id):
    query = project_details.query.filter_by(task_id=task_id)
    stream_task_status = query.first().analysis_status      # {'PENDING', 'SUCCESS'}

    redis_queue_name = get_task_queue_name(task_id)         # {'logs : {task_id}'}
    
    def event_stream():
        if stream_task_status == 'SUCCESS':
            print('Status: SUCCESS')
            logs = celery_logs.query.filter_by(task_id=task_id).all()
            log_message = []
            for log in logs:
                log = log.task_log
                log_message.append(log)

            yield f"data: {log_message}\n\n"

        while stream_task_status == 'PENDING':
            print('Status: PENDING')
            log_message = redis_client.lpop(redis_queue_name)
            if log_message:
                db_celery_logs = celery_logs(task_id, stream_task_status, log_message)
                db.session.add(db_celery_logs)
                try:
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    return jsonify({'message':'Celery logs db error'}), 409

            logs = celery_logs.query.filter_by(task_id=task_id).all()
            log_message = []
            for log in logs:
                log = log.task_log
                log_message.append(log)

            yield f"data: {log_message}\n\n"
            
            time.sleep(20)

        print('event end !')    
        yield f"event: end"

    return Response(stream_with_context(event_stream()), content_type='text/event-stream')

@authentication_controller.route('/projects/<query_email>', methods=['GET']) #{email}
@jwt_required()
def projects(query_email):
    
    query = project_details.query.with_entities(project_details.task_id).filter_by(email=query_email)

    task_status_details = [task_id[0] for task_id in query.all()]

    task_statuses = []

    for task_id in task_status_details:
        result = AsyncResult(task_id, app=celery)
        task_statuses.append({
            'task_id': task_id,
            'status': result.state
        })

    for task_analysis_status in task_statuses:
        if(task_analysis_status['status'] == 'SUCCESS'):
            query = project_details.query.filter_by(task_id=task_analysis_status['task_id'])
            query.update({"analysis_status": "SUCCESS"})
            db.session.commit()

    query = project_details.query.filter_by(email=query_email)
    json_project_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_project_details.append(row_dict)

    return jsonify({'project_details': json_project_details}), 200

@authentication_controller.route('/report/<query_project_name>/<query_application_name>/<query_email>', methods=['GET']) #{project_name} excel reports zip
@jwt_required()
def report(query_project_name, query_application_name, query_email):

    query = project_details.query.filter_by(project_name=query_project_name, application_name=query_application_name, email=query_email)
    print(query.first().file_name)
    fileName = query.first().file_name

    query = analysis_type.query.filter_by(project_name=query_project_name, application_name=query_application_name)
    analysisType = query.first().analysis_type
    print(analysisType)

    if analysisType == 'javaanalysis':
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_remediation_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_inventory_report.xlsx'
        file3_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_high_level_report.xlsx'
    elif analysisType == 'canalysis':
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_inventory_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_report.xlsx'
        file3_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_report_old.xlsx'
    elif analysisType == 'shellanalysis':
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_inventory_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_report.xlsx'

    if not os.path.exists(file1_path):
        return 'File 1 not found', 404
    if not os.path.exists(file2_path):
        return 'File 2 not found', 404
    if not os.path.exists(file3_path):
        return 'File 3 not found', 404
    
    zip_buffer = zipfile.ZipFile('files.zip', 'w', zipfile.ZIP_DEFLATED)
    zip_buffer.write(file1_path, os.path.basename(file1_path))
    zip_buffer.write(file2_path, os.path.basename(file2_path))
    zip_buffer.write(file3_path, os.path.basename(file3_path))
    zip_buffer.close()

    os.chmod('files.zip', 0o777)

    return Response(
        open('files.zip', 'rb').read(),
        mimetype='application/zip',
        headers={
            'Content-Disposition': 'attachment;filename=files.zip',
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/zip',
            'Cache-Control': 'no-cache, no-store, must-revalidate'
        }
    )

@authentication_controller.route('/pdf/<query_project_name>/<query_application_name>/<query_email>', methods=['GET']) #{project_name} pdf generation
@jwt_required()
def pdf(query_project_name, query_application_name, query_email):

    query = project_details.query.filter_by(project_name=query_project_name, application_name=query_application_name, email=query_email)
    json_project_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_project_details.append(row_dict)

    fileName = query.first().file_name

    query = os_details.query.filter_by(project_name=query_project_name, application_name=query_application_name)
    json_os_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_os_details.append(row_dict)

    query = analysis_type.query.filter_by(project_name=query_project_name, application_name=query_application_name)
    json_analysis_type = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_analysis_type.append(row_dict)

    analysisType = query.first().analysis_type
    print(analysisType)
    sheetName = 'Source Code Inventory'
    selected_cols = ['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']

    if analysisType == 'javaanalysis':      
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_report.xlsx'

        sheet_names = ['OS Analysis Details', 'Compilation Error Report', 'Import Class Report', 'JDK Anaylsis Details', 'Import JSP Report', 'Compilation Warning Report']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)

        no_of_artefacts = []
        count_of_ids = 0
        total_count_of_ids = 0
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            count_of_ids = 0
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1
                    total_count_of_ids +=1

            no_of_artefacts.append({sheet_name: count_of_ids})
        print(no_of_artefacts)

    elif analysisType == 'canalysis':
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_inventory_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_report.xlsx'

        sheet_names = ['STK Analysis Detail', '64-bit Analysis Detail', 'Endianness Analysis Detail', 'String-Mem Analysis Detail', 'ExecutionEnv Analysis Detail']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)
        sheets = [wb[sheet_name] for sheet_name in sheet_names]
        count_of_ids = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1

    elif analysisType == 'shellanalysis':
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_inventory_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_report.xlsx'

        sheet_names = ['OS Specific ENV Var Details', 'OS Specific File Path Details', 'Command and option Details', 'Shebang Line Analysis', 'Heredoc Analysis']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)
        sheets = [wb[sheet_name] for sheet_name in sheet_names]
        count_of_ids = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1

    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    df = df.apply(pd.to_numeric, errors='coerce')
    column_sums = df.sum()


    sheetName = 'Source Code Inventory'
    selected_cols = ['Type']
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    type_counts = df['Type'].value_counts().reset_index()
    counts = type_counts.rename(columns={'index': 'Type', 'Type': 'number'})
    counts_list = counts.to_dict('records')

    total_no_of_lines = column_sums['Total Nr LoC']
    percentage = (total_count_of_ids/total_no_of_lines) * 100
    rounded_percentage = round(percentage, 2)

    return jsonify({'project_details': json_project_details, 'os_details': json_os_details, 'analysis_type': json_analysis_type, 'chart1': column_sums.to_dict(), 'chart2' : counts_list, 'chart3' :  no_of_artefacts, 'percent' : rounded_percentage}), 200    

@authentication_controller.route('/documentation/<query_file_name>', methods=['GET'])
@jwt_required()
def export_configuration_file(query_file_name):
    if query_file_name == 'Code Delivery Guidelines-V0.4':      
        file_path = '/usr/u2l/u2l_backend/media/'+ query_file_name +'.docx'
    if query_file_name == 'HPE U2L Questionnaire-light':      
        file_path = '/usr/u2l/u2l_backend/media/'+ query_file_name +'.xls'
    if not os.path.exists(file_path):
        return 'File not found', 404
    zip_buffer = zipfile.ZipFile('files.zip', 'w', zipfile.ZIP_DEFLATED)
    zip_buffer.write(file_path, os.path.basename(file_path))
    zip_buffer.close()
    os.chmod('files.zip', 0o777)
    return Response(
        open('files.zip', 'rb').read(),
        headers={
            'Content-Disposition': 'attachment;filename=files.zip',
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/zip',
            'Cache-Control': 'no-cache, no-store, must-revalidate'
        }
    )

@authentication_controller.route('/contact', methods=['POST'])
@jwt_required()
def contactus():
    json_data = request.get_json()
    print(json_data)
    first_name = json_data['first_name']
    last_name = json_data['last_name']
    email = json_data['email']
    contact_number = json_data['contact_number']
    message = json_data['message']

    user_info = {
        'first_name': first_name,
        'last_name': last_name,
        'email': email,
        'message': message,
        'contact_number': contact_number
    }
    print(f"User Info: {user_info}")

    # Save the user information in the "contact-us" table in the database
    created_at = datetime.datetime.now()
    db_contact_us = contact_us(user_info['first_name'],user_info['last_name'], user_info['email'], user_info['contact_number'], user_info['message'], created_at)
    db.session.add(db_contact_us)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'failed : ' + str(e)}), 409

    send_email(user_info)

    return jsonify({'message': 'Message send successfully. We will contact you !'}), 200

@authentication_controller.route('/generate_logs', methods=['POST'])
@jwt_required()
def generate_logs():

    file_name = request.form['file_name']
    tool_analysis_type = request.form['analysis_type']

    if(tool_analysis_type == 'javaanalysis'):

        yyyy_mmdd = yMD[:4] + '-' + yMD[4:]
        java_remediation_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
        java_high_level_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
        O29=0
        P29=0
        O32=0
        P32=0
        O35=0
        P35=0
        O36=0
        P36=0
        
        # Convert javaSourceScanRemedy
        print("Start 1")
        javaSourceScanRemedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaSourceScanRemedy'
        if os.path.isfile(javaSourceScanRemedy) and os.path.getsize(javaSourceScanRemedy) > 0:
            df = pd.read_csv(javaSourceScanRemedy, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'javaSourceScanRemedy')

            df1 = pd.read_excel(save_location, engine='openpyxl', sheet_name='javaSourceScanRemedy', skiprows=1, header=None)
            df1 = df1.iloc[:, [4]]
        print("End 1")

        # Convert javaSourceCode2Remedy
        print("Start 2")
        javaSourceCode2Remedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaSourceCode2Remedy'
        if os.path.isfile(javaSourceCode2Remedy) and os.path.getsize(javaSourceCode2Remedy) > 0:
            with open(javaSourceCode2Remedy, 'r') as file:
                lines = file.readlines()

            with open(javaSourceCode2Remedy, 'w') as file:
                for line in lines:
                    colon_count = line.count(':')
                    if colon_count == 4:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 5:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 6:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 7:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 8:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 9:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    else:
                        line_without_last_colon = line

                    if line.startswith('\t'):
                        line_without_first_tab = line[1:]
                    else:
                        line_without_first_tab = line

                    file.write(line_without_last_colon)

            with open(javaSourceCode2Remedy, 'r') as file:
                lines = file.readlines()

            with open(javaSourceCode2Remedy, 'w') as file:
                for line in lines:
                    first_tab_index = line.find('\t')
                    if first_tab_index != -1:
                        line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                        line_without_tabs = line_with_colon.replace('\t', '')
                    else:
                        line_without_tabs = line

                    file.write(line_without_tabs)

            df = pd.read_csv(javaSourceCode2Remedy, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            O35 = df['FILE NAME'].nunique()
            P35 = df['FILE NAME'].count()
            
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'javaSourceCode2Remedy')

            book = load_workbook(save_location)
            writer = pd.ExcelWriter(save_location, engine='openpyxl')
            writer.book = book
            sheet_name = 'javaSourceCode2Remedy'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
            writer.save()

            df2 = pd.read_excel(save_location, engine='openpyxl', sheet_name='javaSourceCode2Remedy', skiprows=0)
            df2 = df2.iloc[:, [1,2,3,4,5]]

            book = load_workbook(save_location)
            writer = pd.ExcelWriter(save_location, engine='openpyxl')
            writer.book = book
            sheet_name = 'javaSourceCode2Remedy_1'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
            writer.save()

            sheet_transfer('javaSourceCode2Remedy_1', '10. Java Source Scan Remedy', 6, 1, save_location, java_remediation_path)
        print("End 2")

        # Convert sourceScanRemedy
        print("Start 3.1")
        sourceScanRemedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/sourceScanRemedy'
        if os.path.isfile(sourceScanRemedy_dir_path) and os.path.getsize(sourceScanRemedy_dir_path) > 0:
            ic_lines = []
            def remove_trailing_tabs(line):
                return re.sub(r'\t*$', '', line)
            with open(sourceScanRemedy_dir_path, 'r') as file:
                for line in file:
                    cleaned_line = remove_trailing_tabs(line)
                    if cleaned_line.strip().startswith('IC_'):
                        ic_lines.append(cleaned_line)

            with open(sourceScanRemedy_dir_path, 'w') as file:
                file.write(''.join(ic_lines))
            df = pd.read_csv(sourceScanRemedy_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'sourceScanRemedy')

            df1 = pd.read_excel(save_location, engine='openpyxl', sheet_name='sourceScanRemedy', skiprows=1, header=None)
            df1 = df1.iloc[:, [4]]
        print("End 3.1")

        # Convert sourceCode2Remedy
        print("Start 3.2")
        sourceCode2Remedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/sourceCode2Remedy'
        if os.path.isfile(sourceCode2Remedy_dir_path) and os.path.getsize(sourceCode2Remedy_dir_path) > 0:
            with open(sourceCode2Remedy_dir_path, 'r') as file:
                lines = file.readlines()

            with open(sourceCode2Remedy_dir_path, 'w') as file:
                for line in lines:
                    colon_count = line.count(':')
                    if colon_count == 4:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 5:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 6:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 7:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 8:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 9:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 10:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 11:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 12:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 13:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    elif colon_count == 14:
                        last_colon_index = line.rfind(':')
                        line_without_last_colon = line[:last_colon_index] + line[last_colon_index + 1:]
                    else:
                        line_without_last_colon = line

                    if line.startswith('\t'):
                        line_without_first_tab = line[1:]
                    else:
                        line_without_first_tab = line

                    file.write(line_without_last_colon)

            with open(sourceCode2Remedy_dir_path, 'r') as file:
                lines = file.readlines()

            with open(sourceCode2Remedy_dir_path, 'w') as file:
                for line in lines:
                    first_tab_index = line.find('\t')
                    if first_tab_index != -1:
                        line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                        line_without_tabs = line_with_colon.replace('\t', '')
                    else:
                        line_without_tabs = line

                    file.write(line_without_tabs)
            df = pd.read_csv(sourceCode2Remedy_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            O36 = df['FILE NAME'].nunique()
            P36 = df['FILE NAME'].count()
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'sourceCode2Remedy')

            book = load_workbook(save_location)
            writer = pd.ExcelWriter(save_location, engine='openpyxl')
            writer.book = book
            sheet_name = 'sourceCode2Remedy'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
            writer.save()

            df2 = pd.read_excel(save_location, engine='openpyxl', sheet_name='sourceCode2Remedy', skiprows=0)
            df2 = df2.iloc[:, [1,2,3,4,5]]

            book = load_workbook(save_location)
            writer = pd.ExcelWriter(save_location, engine='openpyxl')
            writer.book = book
            sheet_name = 'sourceCode2Remedy_1'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
            writer.save()

            sheet_transfer('sourceCode2Remedy_1', '13. Affected Framework Remedy', 6, 1, save_location, java_remediation_path)
        print("End 3.2")

        # Convert sourceScanImports.txt
        print("Start 4")
        sourceScanImports_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/sourceScanImports.txt'
        if os.path.isfile(sourceScanImports_dir_path) and os.path.getsize(sourceScanImports_dir_path) > 0:
            df = pd.read_csv(sourceScanImports_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['FILE NAME', 'LINE NUMBER', 'IMPORTS']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'Collection of Imports')
            sheet_transfer('Collection of Imports', '11. Collection of Imports', 6, 1, save_location, java_remediation_path)
        print("End 4")

        # Convert javaUniqRulesGrepped.txt
        print("Start 5")
        javaUniqRulesGrepped_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaUniqRulesGrepped.txt'
        if os.path.isfile(javaUniqRulesGrepped_dir_path) and os.path.getsize(javaUniqRulesGrepped_dir_path) > 0:
            ic_lines = []
            with open(javaUniqRulesGrepped_dir_path, 'r') as file:
                for line in file:
                    cleaned_line = re.sub(r'\t*$', '', line)
                    if cleaned_line.strip().startswith('IC_'):
                        ic_lines.append(cleaned_line)

            with open(javaUniqRulesGrepped_dir_path, 'w') as file:
                file.write('\n'.join(ic_lines))

            with open(javaUniqRulesGrepped_dir_path, 'r') as file:
                lines = [re.sub(r'\t*$', '', line) for line in file]

            with open(javaUniqRulesGrepped_dir_path, 'w') as file:
                file.write('\n'.join(lines))
            print('inside 5')
            df = pd.read_csv(javaUniqRulesGrepped_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'ENTITY', 'PACKAGE', 'OBJECT', 'REMEDY']

            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
            generate_workbook(df, save_location,'Possible Import Remediations')
            sheet_transfer('Possible Import Remediations', '12. Possible Import Remediation', 6, 1, save_location, java_remediation_path)
        print("End 5")

        # Convert javacWarnings.list
        print("Start 6")
        javacWarnings_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/java/javacLog/2023-0801/javacWarnings.list'
        if os.path.isfile(javacWarnings_dir_path) and os.path.getsize(javacWarnings_dir_path) > 0:
            df = pd.read_csv(javacWarnings_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['WARNING MESSAGE WITH TYPE', 'FILE PATH', 'FILE NAME', 'LINE', 'Column5']
            df = df.drop(columns=['Column5'])
            O29 = df['FILE NAME'].nunique()
            P29 = df['FILE NAME'].count()
            print(O29)
            print(P29)
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Compilation Warning Report')
            sheet_transfer('Compilation Warning Report', '2. Compilation Warning Detail', 6, 1, save_location, java_remediation_path)
        print("End 6")

        # Convert import_lines.out
        print("Start 7")
        importLines_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/java/import/2023-0801/import_lines.out'
        if os.path.isfile(importLines_dir_path) and os.path.getsize(importLines_dir_path) > 0:
            df = pd.read_csv(importLines_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['IMPORTS', 'FILE NAME', 'LINE NUMBER']
            df.index = df.index + 1
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Import Class Report')
            sheet_transfer('Import Class Report', '8. Import Class Detail', 6, 1, save_location, java_remediation_path)
        print("End 7")

        # Convert import_lines_jsp.out
        print("Start 8")
        importLinesJsp_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/java/import/2023-0801/import_lines_jsp.out'
        if os.path.isfile(importLinesJsp_dir_path) and os.path.getsize(importLinesJsp_dir_path) > 0:
            df = pd.read_csv(importLinesJsp_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['FILE NAME', 'LINE NUMBER', 'IMPORTED JSP']
            df['FILE NAME'] = './' + df['FILE NAME']
            df.index = df.index + 1
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Import JSP Report')
            sheet_transfer('Import JSP Report', '9. Import JSP Detail', 6, 1, save_location,java_remediation_path)
        print("End 8")
        
        # Best Practices
        print("Start 9")
        best_pratices_path = '/usr/u2l/u2l_backend/pmd/' + file_name + '/BestPractices_Report.csv'
        if os.path.isfile(best_pratices_path) and os.path.getsize(best_pratices_path) > 0:
            df = pd.read_csv(best_pratices_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
            df['File'] = df['File'].str.replace(r'/home/hpsa/ap/ALL/Java/', './', regex=True)
            df = df.drop(df.index[0])
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Best_Practices')
            sheet_transfer('Best_Practices', '14. Best Practices', 6, 1, save_location, java_remediation_path)
        print("End 9")

        # MultiThreading
        print("Start 10")
        multithreading_path = '/usr/u2l/u2l_backend/pmd/' + file_name + '/MultiThreading_Report.csv'
        if os.path.isfile(multithreading_path) and os.path.getsize(multithreading_path) > 0:
            df = pd.read_csv(multithreading_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
            df['File'] = df['File'].str.replace(r'/home/hpsa/ap/ALL/Java/', './', regex=True)
            df = df.drop(df.index[0])
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Multithreading')
            sheet_transfer('Multithreading', '15. MultiThreading', 6, 1, save_location, java_remediation_path)
        print("End 10")

        # Performance
        print("Start 11")
        performance_path = '/usr/u2l/u2l_backend/pmd/' + file_name + '/Performance_Report.csv'
        if os.path.isfile(performance_path) and os.path.getsize(performance_path) > 0:
            df = pd.read_csv(performance_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
            df['File'] = df['File'].str.replace(r'/home/hpsa/ap/ALL/Java/', './', regex=True)
            df = df.drop(df.index[0])
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Performance')
            sheet_transfer('Performance', '16. Performance', 6, 1, save_location, java_remediation_path)
        print("End 11")

        # Security
        print("Start 12")
        security_path = '/usr/u2l/u2l_backend/pmd/' + file_name + '/Security_Report.csv'
        if os.path.isfile(security_path) and os.path.getsize(security_path) > 0:
            df = pd.read_csv(security_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
            df['File'] = df['File'].str.replace(r'/home/hpsa/ap/ALL/Java/', './', regex=True)
            df = df.drop(df.index[0])
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'Security')
            sheet_transfer('Security', '17. Security', 6, 1, save_location, java_remediation_path)
        print("End 12")
        
        #Source Code Inventory
        print("Start 13")
        regex = r'^.*$'
        source_code_java_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/Log_Source_Analysis_20230801.log'
        print(source_code_java_path)
        if os.path.isfile(source_code_java_path) and os.path.getsize(source_code_java_path) > 0:
            
            with open(source_code_java_path, 'r') as f:
                data = f.read()
            lines = data.split('\n')
            log_entries = []
            for line in lines:
                match = re.match(regex, line)
                if match:
                    log_entries.append(match.group())
            df = pd.DataFrame(log_entries, columns=['log_message'])

            start_str = '###########  Execution 0001_stepcounter Starts  ###########'
            end_str = '###########  Execution 0001_stepcounter Ends ###########'
            start_df = df[df['log_message'].str.contains(start_str)]
            end_df = df[df['log_message'].str.contains(end_str)]

            start_index = start_df.index[0]
            end_index = end_df.index[0]
            extracted_df = df.iloc[start_index+1:end_index]

            result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
            result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
            result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
            result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
            result_df = result_df.drop(['log_message', 'col2'], axis=1)
            result_df = result_df.reset_index(drop=True)
            result_df.index += 1

            artefact_type_dict = result_df['Type'].unique()                                     # ['Java', 'KShell']
            type_counts = result_df['Type'].value_counts()
            type_counts_dict = type_counts.to_dict()                                            # {'KShell': 23, 'Java': 2}
            F6 = sum(type_counts_dict.values())
            result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
            autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
            autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
            O6 = sum(autual_by_type_dict.values())
            result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
            total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
            total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'Java': 424, 'KShell': 5841}

            save_location_1 = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(result_df, save_location_1,'Source Code Inventory')

            save_location_2 = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_inventory_report.xlsx'

            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, save_location_1, save_location_2)
            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, save_location_1, java_remediation_path)
            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, save_location_1, java_high_level_path)
            print('Transfer Completed !!')

            file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
            wb = load_workbook(file_path)
            ws = wb['Environment Information']
            start_row = 22
            start_col = 2
            col_actual = 2
            col_total = 3

            for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                col_idx = start_col
                ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                value = type_counts_dict.get(artefact_type, 0)
                col_idx = start_col + 1
                ws.cell(row=row_idx, column=col_idx, value=value)

                actual_value = autual_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_actual
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

                actual_value = total_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_total
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

            wb.save(file_path)

            file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
            wb = load_workbook(file_path)
            ws = wb['Environment Information']
            start_row = 22
            start_col = 2
            col_actual = 2
            col_total = 3

            for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                col_idx = start_col
                ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                value = type_counts_dict.get(artefact_type, 0)
                col_idx = start_col + 1
                ws.cell(row=row_idx, column=col_idx, value=value)

                actual_value = autual_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_actual
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

                actual_value = total_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_total
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

            wb.save(file_path)
        print("End 13")

        # Convert OS_diff.out
        print("Start 14")
        OSDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/java/versiondiff/2023-0801/OS_diff.out'
        if os.path.isfile(OSDiff_dir_path) and os.path.getsize(OSDiff_dir_path) > 0:
            print("inside 12")

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                if not line.startswith('OS-IC'):
                    line = line.replace('\t', '')
                    modified_lines[-1] = modified_lines[-1].rstrip() + line.strip()
                else:
                    modified_lines.append(line)

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                parts = line.split()
                parts = [part.strip() for part in parts]
                modified_line = '@'.join(parts)
                modified_lines.append(modified_line + '\n')

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                modified_line = re.sub(r'@', lambda match: ' ' if match.start() > line.find('@', line.find('@', line.find('@') + 1) + 1) else match.group(), line.strip())
                modified_lines.append(modified_line + '\n')

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            df = pd.read_csv(OSDiff_dir_path, sep='@', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)

            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
            O32 = df['FILE NAME'].nunique()
            P32 = df['FILE NAME'].count()

            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'OS Analysis Details')
            sheet_transfer('OS Analysis Details', '5. OS Analysis Detail', 6, 1, save_location, java_remediation_path)
        print("End 14")

        # Convert jdk_diff.out
        print("Start 15")
        jdkDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/JavaAnalysis/java/versiondiff/2023-0801/jdk_diff.out'
        if os.path.isfile(jdkDiff_dir_path) and os.path.getsize(jdkDiff_dir_path) > 0:
            df = pd.read_csv(jdkDiff_dir_path, sep='\s+', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df = df.drop(columns=['Column1'])
            df.index = df.index + 1
            save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

            generate_workbook(df, save_location,'JDK Anaylsis Details')
            sheet_transfer('JDK Anaylsis Details', '3. JDK Analysis Detail', 6, 1, save_location, java_remediation_path)
        print("End 15")

        #java_high_level_report.xlsx - Source Code Impacted
        print("Start 16")
        excel_file = java_remediation_path
        sheet_configs = {
            '1. Compilation Error Report': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '2. Compilation Warning Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '3. JDK Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '4. J2EE Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '5. OS Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '6. App Server Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '7. Database Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            # '8. Import Class Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            # '9. Import JSP Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '10. Java Source Scan Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1},
            # '11. Collection of Imports': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1},
            # '12. Possible Import Remediation': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '13. Affected Framework Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1},
            # '14. Best Practices': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 3},
            # '15. MultiThreading': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 3},
            # '16. Performance': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 3},
            # '17. Security': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 3},
        }

        dfs = []

        for sheet_name, config in sheet_configs.items():
            df = pd.read_excel(excel_file, engine='openpyxl', sheet_name=sheet_name, usecols=config['usecols'], skiprows=config['skiprows'])
            df = df.drop(df.columns[0], axis=1)
            df = df.iloc[:, config['column_index']]

            if not df.empty and not df.isnull().all().all():
                dfs.append(df)

        df_combined = pd.concat(dfs, ignore_index=True)
        df_combined = df_combined.drop_duplicates()
        df_combined = df_combined.reset_index(drop=True)
        df_combined = df_combined.dropna()
        df_count = len(df_combined)
        F7 = df_count
        print(df_combined)

        save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
        writer = pd.ExcelWriter(save_location, engine='openpyxl', mode='a')
        book = load_workbook(save_location)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_combined.to_excel(writer, sheet_name='Source Code Impacted', index=False, startrow=6, startcol=1, header=False)
        writer.save()

        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
        wb = load_workbook(file_path)
        ws = wb['Migration Summary']
        start_row = 7
        start_col = 6
        ws.cell(row=start_row, column=start_col, value=df_count)

        wb.save(file_path)

        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
        wb = load_workbook(file_path)
        ws = wb['Analysis Summary']
        start_row = 7
        start_col = 6
        ws.cell(row=start_row, column=start_col, value=df_count)

        wb.save(file_path)

        filepath = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
        # fetch_store_java_data(form_project_name, form_application_name, filepath, filepath_percentage)

        # Migration Summary : java_remediation_report
        print('Migration Summary')
        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
        wb = load_workbook(file_path)
        ws = wb['Migration Summary']
        start_col_1 = 15
        start_row_1 = 29
        start_row_2 = 32
        start_row_3 = 35
        start_row_4 = 36
        start_col_2 = 16
        start_row_5 = 29
        start_row_6 = 32
        start_row_7 = 35
        start_row_8 = 36
        if O29 != 0:
            ws.cell(row=start_row_1, column=start_col_1, value=O29)
        if O32 != 0:
            ws.cell(row=start_row_2, column=start_col_1, value=O32)
        if O35 != 0:
            ws.cell(row=start_row_3, column=start_col_1, value=O35)
        if O36 != 0:
            ws.cell(row=start_row_4, column=start_col_1, value=O36)
        if P29 != 0:
            ws.cell(row=start_row_5, column=start_col_2, value=P29)
        if P32 != 0:
            ws.cell(row=start_row_6, column=start_col_2, value=P32)
        if P35 != 0:
            ws.cell(row=start_row_7, column=start_col_2, value=P35)
        if P36 != 0:
            ws.cell(row=start_row_8, column=start_col_2, value=P36)

        O7 = P29 + P32 + P35 + P36
        wb.save(file_path)

        # print('Storing values to database')
        # db_analysis_summary_java = analysis_summary_java(form_project_name, form_application_name, O29, P29, O32, P32, O35, P35, O36, P36, F6, F7, O6, O7)
        # db.session.add(db_analysis_summary_java)
        # try:
        #     db.session.commit()
        # except Exception as e:
        #     db.session.rollback()

        print('Analysis Summary')
        # Analysis Summary : java_high_level_report
        file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
        wb = load_workbook(file_path)
        ws = wb['Analysis Summary']
        start_col_1 = 15
        start_row_1 = 29
        start_row_2 = 32
        start_row_3 = 35
        start_row_4 = 36
        start_col_2 = 16
        start_row_5 = 29
        start_row_6 = 32
        start_row_7 = 35
        start_row_8 = 36
        if O29 != 0:
            ws.cell(row=start_row_1, column=start_col_1, value=O29)
        if O32 != 0:
            ws.cell(row=start_row_2, column=start_col_1, value=O32)
        if O35 != 0:
            ws.cell(row=start_row_3, column=start_col_1, value=O35)
        if O36 != 0:
            ws.cell(row=start_row_4, column=start_col_1, value=O36)

        # Check and assign values for the second column
        if P29 != 0:
            ws.cell(row=start_row_5, column=start_col_2, value=P29)
        if P32 != 0:
            ws.cell(row=start_row_6, column=start_col_2, value=P32)
        if P35 != 0:
            ws.cell(row=start_row_7, column=start_col_2, value=P35)
        if P36 != 0:
            ws.cell(row=start_row_8, column=start_col_2, value=P36)

        wb.save(file_path)

        print("End 16")

    return '----------X Report Generated X----------'

@authentication_controller.route('/generate_logs_latest', methods=['POST'])
@jwt_required()
def generate_logs_latest():
    file_name = request.form['file_name']
    tool_analysis_type = request.form['analysis_type']
    yyyy = request.form['yyyy']
    mm = request.form['mm']
    dd = request.form['dd']
    yyyy_mmdd = yyyy + '-' + mm + dd
    yyyymmdd = yyyy + mm + dd
    
    source_dir_report = '/usr/u2l/u2l_backend/report_templates/'
    dest_dir = '/usr/u2l/u2l_backend/projects/' + file_name

    if(tool_analysis_type == 'javaanalysis'):

        source_path = os.path.join(source_dir_report, 'java_report.xlsx')
        dest_path = os.path.join(dest_dir, 'java_report.xlsx')
        shutil.copy(source_path, dest_path)

        source_path = os.path.join(source_dir_report, 'java_inventory_report.xlsx')
        dest_path = os.path.join(dest_dir, 'java_inventory_report.xlsx')
        shutil.copy(source_path, dest_path)

        source_path = os.path.join(source_dir_report, 'java_remediation_report.xlsx')
        dest_path = os.path.join(dest_dir, 'java_remediation_report.xlsx')
        shutil.copy(source_path, dest_path)

        source_path = os.path.join(source_dir_report, 'java_high_level_report.xlsx')
        dest_path = os.path.join(dest_dir, 'java_high_level_report.xlsx')
        shutil.copy(source_path, dest_path)
    
    if(tool_analysis_type == 'javaanalysis'):
        tool_analysis_type = 'JavaAnalysis'
        
        java_remediation_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_remediation_report.xlsx'
        java_high_level_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_high_level_report.xlsx'
        java_inventory_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_inventory_report.xlsx'
        java_report_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'
        O29=0
        P29=0
        O32=0
        P32=0
        O35=0
        P35=0
        O36=0
        P36=0
        
        # Convert javaSourceScanRemedy
        print("Start 1 : javaSourceScanRemedy")
        javaSourceScanRemedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaSourceScanRemedy'
        if os.path.isfile(javaSourceScanRemedy) and os.path.getsize(javaSourceScanRemedy) > 0:
            df = pd.read_csv(javaSourceScanRemedy, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            generate_workbook(df, java_report_path,'javaSourceScanRemedy')

            df1 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='javaSourceScanRemedy', skiprows=1, header=None)
            df1 = df1.iloc[:, [4]]
        print("End 1 : javaSourceScanRemedy")

        # Convert javaSourceCode2Remedy
        print("Start 2 : javaSourceCode2Remedy")
        javaSourceCode2Remedy = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaSourceCode2Remedy'
        if os.path.isfile(javaSourceCode2Remedy) and os.path.getsize(javaSourceCode2Remedy) > 0:
            with open(javaSourceCode2Remedy, 'r') as file:
                lines = file.readlines()

            with open(javaSourceCode2Remedy, 'w') as file:
                for line in lines:
                    first_tab_index = line.find('\t')
                    if first_tab_index != -1:
                        line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                        line_without_tabs = line_with_colon.replace('\t', '')
                    else:
                        line_without_tabs = line

                    file.write(line_without_tabs)
            
            with open(javaSourceCode2Remedy, 'r') as file:
                lines = file.readlines()

            with open(javaSourceCode2Remedy, 'w') as file:
                for line in lines:
                    parts = line.split(':')
                    modified_line = ':'.join(parts[:4]) + ''.join(parts[4:])
                    file.write(modified_line)

            df = pd.read_csv(javaSourceCode2Remedy, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            O33 = df['FILE NAME'].nunique()
            P33 = df['FILE NAME'].count()
            
            generate_workbook(df, java_report_path,'javaSourceCode2Remedy')

            book = load_workbook(java_report_path)
            writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
            writer.book = book
            sheet_name = 'javaSourceCode2Remedy'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
            writer.save()

            df2 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='javaSourceCode2Remedy', skiprows=0)
            df2 = df2.iloc[:, [1,2,3,4,5]]

            book = load_workbook(java_report_path)
            writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
            writer.book = book
            sheet_name = 'javaSourceCode2Remedy_1'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
            writer.save()

            sheet_transfer('javaSourceCode2Remedy_1', '6. Java Source Scan Remedy', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '6. Java Source Scan Remedy', '6. Java Source Scan Summary', 5, 4, java_remediation_path, 7, 2)
        print("Sheet : 6. Java Source Scan Remedy : CREATED")
        print("Sheet : 6. Java Source Scan Summary : CREATED")
        print("End 2 : javaSourceCode2Remedy\n")

        # Convert sourceScanRemedy
        print("Start 3 : sourceScanRemedy")
        sourceScanRemedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/sourceScanRemedy'
        if os.path.isfile(sourceScanRemedy_dir_path) and os.path.getsize(sourceScanRemedy_dir_path) > 0:
            ic_lines = []
            def remove_trailing_tabs(line):
                return re.sub(r'\t*$', '', line)
            with open(sourceScanRemedy_dir_path, 'r') as file:
                for line in file:
                    cleaned_line = remove_trailing_tabs(line)
                    if cleaned_line.strip().startswith('IC_'):
                        ic_lines.append(cleaned_line)

            with open(sourceScanRemedy_dir_path, 'w') as file:
                file.write(''.join(ic_lines))

            df = pd.read_csv(sourceScanRemedy_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'REMEDY']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            generate_workbook(df, java_report_path,'sourceScanRemedy')

            df1 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='sourceScanRemedy', skiprows=1, header=None)
            df1 = df1.iloc[:, [4]]
        print("End 3 : sourceScanRemedy")

        # Convert sourceCode2Remedy
        print("Start 4 : sourceCode2Remedy")
        sourceCode2Remedy_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/sourceCode2Remedy'
        if os.path.isfile(sourceCode2Remedy_dir_path) and os.path.getsize(sourceCode2Remedy_dir_path) > 0:
            with open(sourceCode2Remedy_dir_path, 'r') as file:
                lines = file.readlines()

            with open(sourceCode2Remedy_dir_path, 'w') as file:
                for line in lines:
                    first_tab_index = line.find('\t')
                    if first_tab_index != -1:
                        line_with_colon = line[:first_tab_index] + ':' + line[first_tab_index + 1:]
                        line_without_tabs = line_with_colon.replace('\t', '')
                    else:
                        line_without_tabs = line

                    file.write(line_without_tabs)
            
            with open(sourceCode2Remedy_dir_path, 'r') as file:
                lines = file.readlines()

            with open(sourceCode2Remedy_dir_path, 'w') as file:
                for line in lines:
                    parts = line.split(':')
                    modified_line = ':'.join(parts[:4]) + ''.join(parts[4:])
                    file.write(modified_line)

            df = pd.read_csv(sourceCode2Remedy_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'AFFECTED CODE']
            df['FILE NAME'] = df['FILE NAME'].str.replace(r'/home/hpsa/u2l_local/'+ file_name +'/work/JavaAnalysis/', './', regex=True)
            O35 = df['FILE NAME'].nunique()
            P35 = df['FILE NAME'].count()
            generate_workbook(df, java_report_path,'sourceCode2Remedy')

            book = load_workbook(java_report_path)
            writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
            writer.book = book
            sheet_name = 'sourceCode2Remedy'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df1.to_excel(writer, sheet_name=sheet_name, startcol=5, startrow=1, index=False, header=False)
            writer.save()

            df2 = pd.read_excel(java_report_path, engine='openpyxl', sheet_name='sourceCode2Remedy', skiprows=0)
            df2 = df2.iloc[:, [1,2,3,4,5]]

            book = load_workbook(java_report_path)
            writer = pd.ExcelWriter(java_report_path, engine='openpyxl')
            writer.book = book
            sheet_name = 'sourceCode2Remedy_1'
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            df2.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, index=False, header=False)
            writer.save()

            sheet_transfer('sourceCode2Remedy_1', '8. Affected Framework Remedy', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '8. Affected Framework Remedy', '8. Affected Framework Summary', 5, 4, java_remediation_path, 7, 2)
        print("Sheet : 8. Affected Framework Remedy : CREATED")
        print("Sheet : 8. Affected Framework Summary : CREATED")
        print("End 4 : sourceCode2Remedy\n")

        # Convert javaUniqRulesGrepped.txt
        print("Start 5 : javaUniqRulesGrepped.txt")
        javaUniqRulesGrepped_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_logs/javaUniqRulesGrepped.txt'
        if os.path.isfile(javaUniqRulesGrepped_dir_path) and os.path.getsize(javaUniqRulesGrepped_dir_path) > 0:
            ic_lines = []
            with open(javaUniqRulesGrepped_dir_path, 'r') as file:
                for line in file:
                    cleaned_line = re.sub(r'\t*$', '', line)
                    if cleaned_line.strip().startswith('IC_'):
                        ic_lines.append(cleaned_line)

            with open(javaUniqRulesGrepped_dir_path, 'w') as file:
                file.write('\n'.join(ic_lines))

            with open(javaUniqRulesGrepped_dir_path, 'r') as file:
                lines = [re.sub(r'\t*$', '', line) for line in file]

            with open(javaUniqRulesGrepped_dir_path, 'w') as file:
                file.write('\n'.join(lines))
            print('inside 5')
            df = pd.read_csv(javaUniqRulesGrepped_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['RULE ID', 'ENTITY', 'PACKAGE', 'OBJECT', 'REMEDY']

            generate_workbook(df, java_report_path,'Possible Import Remediations')
            sheet_transfer('Possible Import Remediations', '7. Possible Import Remediation', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '7. Possible Import Remediation', '7. Import Remediation Summary', 5, 5, java_remediation_path, 7, 2)
        print("Sheet : 7. Possible Import Remediation : CREATED")
        print("Sheet : 7. Import Remediation Summary : CREATED")
        print("End 5 : javaUniqRulesGrepped.txt\n")

        # Convert javacWarnings.list
        print("Start 6 : javacWarnings.list")
        javacWarnings_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/java/javacLog/' + yyyy_mmdd + '/javacWarnings.list'
        if os.path.isfile(javacWarnings_dir_path) and os.path.getsize(javacWarnings_dir_path) > 0:
            df = pd.read_csv(javacWarnings_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.index = df.index + 1
            df.columns = ['WARNING MESSAGE WITH TYPE', 'FILE PATH', 'FILE NAME', 'LINE', 'Column5']
            df = df.drop(columns=['Column5'])
            O29 = df['FILE NAME'].nunique()
            P29 = df['FILE NAME'].count()

            generate_workbook(df, java_report_path,'Compilation Warning Report')
            sheet_transfer('Compilation Warning Report', '2. Compilation Warning Detail', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '2. Compilation Warning Detail', '2. Compilation Warning Summary', 5, 1, java_remediation_path, 8, 2)
        print("Sheet : 2. Compilation Warning Detail : CREATED")
        print("Sheet : 2. Compilation Warning Summary : CREATED")
        print("End 6 : javacWarnings.list\n")

        # Convert import_lines.out
        print("Start 7 : import_lines.out")
        importLines_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines.out'
        if os.path.isfile(importLines_dir_path) and os.path.getsize(importLines_dir_path) > 0:
            df = pd.read_csv(importLines_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['IMPORTS', 'FILE NAME', 'LINE NUMBER']
            df.index = df.index + 1

            generate_workbook(df, java_report_path,'Import Class Report')
            sheet_transfer('Import Class Report', 'Import Class Detail', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : Import Class Detail : CREATED")
        print("End 7 : import_lines.out\n")

        # Convert import_lines_jsp.out
        print("Start 8 : import_lines_jsp.out")
        importLinesJsp_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines_jsp.out'
        if os.path.isfile(importLinesJsp_dir_path) and os.path.getsize(importLinesJsp_dir_path) > 0:
            df = pd.read_csv(importLinesJsp_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['FILE NAME', 'LINE NUMBER', 'IMPORTED JSP']
            df['FILE NAME'] = './' + df['FILE NAME']
            df.index = df.index + 1

            generate_workbook(df, java_report_path,'Import JSP Report')
            sheet_transfer('Import JSP Report', 'Import JSP Detail', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : Import JSP Detail : CREATED")
        print("End 8 : import_lines_jsp.out\n")
        
        # Best Practices
        print("Start 9 : Best Practices")
        best_pratices_path = '/usr/u2l/u2l_backend/pmd/'+ file_name +'/BestPractices_Report.csv'
        if os.path.isfile(best_pratices_path) and os.path.getsize(best_pratices_path) > 0:
            df = pd.read_csv(best_pratices_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
            df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
            df = df.drop(df.index[0])

            generate_workbook(df, java_report_path,'Best_Practices')
            sheet_transfer('Best_Practices', '9. Best Practices', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : 9. Best Practices : CREATED")
        print("End 9 : Best Practices\n")

        # MultiThreading
        print("Start 10 : MultiThreading")
        multithreading_path = '/usr/u2l/u2l_backend/pmd/'+ file_name +'/MultiThreading_Report.csv'
        if os.path.isfile(multithreading_path) and os.path.getsize(multithreading_path) > 0:
            df = pd.read_csv(multithreading_path, sep=',', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ['Problem','Package','File','Priority','Line','Description','Rule Set','Rule']
            df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
            df = df.drop(df.index[0])

            generate_workbook(df, java_report_path,'Multithreading')
            sheet_transfer('Multithreading', '10. MultiThreading', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : 10. MultiThreading : CREATED")
        print("End 10 : MultiThreading\n")

        # Performance
        print("Start 11 : Performance")
        performance_path = '/usr/u2l/u2l_backend/pmd/'+ file_name +'/Performance_Report.csv'
        if os.path.isfile(performance_path) and os.path.getsize(performance_path) > 0:
            df = pd.read_csv(performance_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
            df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
            df = df.drop(df.index[0])

            generate_workbook(df, java_report_path,'Performance')
            sheet_transfer('Performance', '11. Performance', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : 11. Performance : CREATED")
        print("End 11 : Performance")

        # Security
        print("Start 12 : Security")
        security_path = '/usr/u2l/u2l_backend/u2l_backend/pmd/'+ file_name +'/Security_Report.csv'
        if os.path.isfile(security_path) and os.path.getsize(security_path) > 0:
            df = pd.read_csv(security_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df.columns = ["Problem","Package","File","Priority","Line","Description","Rule set","Rule"]
            df['File'] = df['File'].str.replace(r'/usr/u2l/u2l_backend/projects/'+ file_name +'/U2L/java-vfunction_' + yMD + '_' + hMS + '/work/javaanalysis(/new)?/', './', regex=True)
            df = df.drop(df.index[0])

            generate_workbook(df, java_report_path,'Security')
            sheet_transfer('Security', '12. Security', 6, 1, java_report_path, java_remediation_path)
        print("Sheet : 12. Security : CREATED")
        print("End 12 : Security")
        
        #Source Code Inventory
        print("Start 13 : Source Code Inventory")
        regex = r'^.*$'
        source_code_java_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yyyymmdd + '.log'
        print(source_code_java_path)
        if os.path.isfile(source_code_java_path) and os.path.getsize(source_code_java_path) > 0:
            
            with open(source_code_java_path, 'r') as f:
                data = f.read()
            lines = data.split('\n')
            log_entries = []
            for line in lines:
                match = re.match(regex, line)
                if match:
                    log_entries.append(match.group())
            df = pd.DataFrame(log_entries, columns=['log_message'])

            start_str = '###########  Execution 0001_stepcounter Starts  ###########'
            end_str = '###########  Execution 0001_stepcounter Ends ###########'
            start_df = df[df['log_message'].str.contains(start_str)]
            end_df = df[df['log_message'].str.contains(end_str)]

            start_index = start_df.index[0]
            end_index = end_df.index[0]
            extracted_df = df.iloc[start_index+1:end_index]

            result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
            result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
            result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
            result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
            result_df = result_df.drop(['log_message', 'col2'], axis=1)
            result_df = result_df.reset_index(drop=True)
            result_df.index += 1

            artefact_type_dict = result_df['Type'].unique()                                     # ['Java', 'KShell']
            type_counts = result_df['Type'].value_counts()
            type_counts_dict = type_counts.to_dict()                                            # {'KShell': 23, 'Java': 2}
            F6 = sum(type_counts_dict.values())
            result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
            autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
            autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
            O6 = sum(autual_by_type_dict.values())
            result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
            total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
            total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'Java': 424, 'KShell': 5841}

            generate_workbook(result_df, java_report_path,'Source Code Inventory')

            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, java_report_path, java_inventory_path)
            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, java_report_path, java_remediation_path)
            sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, java_report_path, java_high_level_path)
            print('Transfer Completed !!')

            wb = load_workbook(java_remediation_path)
            print('1')
            ws = wb['Environment Information']
            print('2')
            start_row = 22
            start_col = 2
            col_actual = 2
            col_total = 3

            for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                col_idx = start_col
                ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                value = type_counts_dict.get(artefact_type, 0)
                col_idx = start_col + 1
                ws.cell(row=row_idx, column=col_idx, value=value)

                actual_value = autual_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_actual
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

                actual_value = total_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_total
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

            wb.save(java_remediation_path)

            wb = load_workbook(java_high_level_path)
            ws = wb['Environment Information']
            start_row = 22
            start_col = 2
            col_actual = 2
            col_total = 3

            for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
                col_idx = start_col
                ws.cell(row=row_idx, column=col_idx, value=artefact_type)

                value = type_counts_dict.get(artefact_type, 0)
                col_idx = start_col + 1
                ws.cell(row=row_idx, column=col_idx, value=value)

                actual_value = autual_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_actual
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

                actual_value = total_by_type_dict.get(artefact_type, 0)
                col_idx = start_col + col_total
                ws.cell(row=row_idx, column=col_idx, value=actual_value)

            wb.save(java_high_level_path)
        print("Sheet : Source Code Inventory : CREATED")
        print("End 13 : Source Code Inventory")

        # Convert OS_diff.out
        print("Start 14 : OS_diff.out")
        OSDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/OS_diff.out'
        if os.path.isfile(OSDiff_dir_path) and os.path.getsize(OSDiff_dir_path) > 0:
            print("inside 12")

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                if not line.startswith('OS-IC'):
                    line = line.replace('\t', '')
                    modified_lines[-1] = modified_lines[-1].rstrip() + line.strip()
                else:
                    modified_lines.append(line)

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                parts = line.split()
                parts = [part.strip() for part in parts]
                modified_line = '@'.join(parts)
                modified_lines.append(modified_line + '\n')

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            with open(OSDiff_dir_path, 'r') as file:
                lines = file.readlines()

            modified_lines = []
            for line in lines:
                modified_line = re.sub(r'@', lambda match: ' ' if match.start() > line.find('@', line.find('@', line.find('@') + 1) + 1) else match.group(), line.strip())
                modified_lines.append(modified_line + '\n')

            with open(OSDiff_dir_path, 'w') as file:
                file.writelines(modified_lines)

            df = pd.read_csv(OSDiff_dir_path, sep='@', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)

            df.columns = ['RULE ID', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
            O32 = df['FILE NAME'].nunique()
            P32 = df['FILE NAME'].count()

            generate_workbook(df, java_report_path,'OS Analysis Details')
            sheet_transfer('OS Analysis Details', '5. OS Analysis Detail', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '5. OS Analysis Detail', '5. OS Analysis Summary', 5, 5, java_remediation_path, 8, 2)
        print("Sheet : 5. OS Analysis Detail : CREATED")
        print("Sheet : 5. OS Analysis Summary : CREATED")
        print("End 14 : OS_diff.out\n")

        # Convert jdk_diff.out
        print("Start 15 : jdk_diff.out")
        jdkDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/jdk_diff.out'
        if os.path.isfile(jdkDiff_dir_path) and os.path.getsize(jdkDiff_dir_path) > 0:
            df = pd.read_csv(jdkDiff_dir_path, sep='\s+', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            df = df.drop(columns=['Column1'])
            df.index = df.index + 1

            generate_workbook(df, java_report_path,'JDK Anaylsis Details')
            sheet_transfer('JDK Anaylsis Details', '3. JDK Analysis Detail', 6, 1, java_report_path, java_remediation_path)
            summary_sheet_generation(java_remediation_path, '3. JDK Analysis Detail', '3. JDK Analysis Summary', 5, 6, java_remediation_path, 8, 2)
        print("Sheet : 3. JDK Analysis Detail : CREATED")
        print("Sheet : 3. JDK Analysis Summary : CREATED")
        print("End 15 : jdk_diff.out\n")

        #java_high_level_report.xlsx - Source Code Impacted
        print("Start 16 : Source Code Impacted")
        sheet_configs = {
            '1. Compilation Error Report': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '2. Compilation Warning Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '3. JDK Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '4. J2EE Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '5. OS Analysis Detail': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '6. Java Source Scan Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1},
            # '7. Possible Import Remediation': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 2},
            '8. Affected Framework Remedy': {'usecols': lambda x: x != 1, 'skiprows': 5, 'column_index': 1}
        }

        dfs = []

        for sheet_name, config in sheet_configs.items():
            df = pd.read_excel(java_remediation_path, engine='openpyxl', sheet_name=sheet_name, usecols=config['usecols'], skiprows=config['skiprows'])
            df = df.drop(df.columns[0], axis=1)
            df = df.iloc[:, config['column_index']]

            if not df.empty and not df.isnull().all().all():
                dfs.append(df)

        df_combined = pd.concat(dfs, ignore_index=True)
        df_combined = df_combined.drop_duplicates()
        df_combined = df_combined.reset_index(drop=True)
        df_combined = df_combined.dropna()
        df_count = len(df_combined)
        F7 = df_count
        print(df_combined)

        writer = pd.ExcelWriter(java_high_level_path, engine='openpyxl', mode='a')
        book = load_workbook(java_high_level_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df_combined.to_excel(writer, sheet_name='Source Code Impacted', index=False, startrow=6, startcol=1, header=False)
        writer.save()

        wb = load_workbook(java_remediation_path)
        ws = wb['Migration Summary']
        start_row = 7
        start_col = 6
        ws.cell(row=start_row, column=start_col, value=df_count)

        wb.save(java_remediation_path)

        wb = load_workbook(java_high_level_path)
        ws = wb['Analysis Summary']
        start_row = 7
        start_col = 6
        ws.cell(row=start_row, column=start_col, value=df_count)

        wb.save(java_high_level_path)

        # fetch_store_java_data(form_project_name, form_application_name, java_report_path)

        # Migration Summary : java_remediation_report
        print('Migration Summary')
        wb = load_workbook(java_remediation_path)
        ws = wb['Migration Summary']
        start_col_1 = 15
        start_row_1 = 29
        start_row_2 = 32
        start_row_3 = 33
        start_row_4 = 35
        start_col_2 = 16
        start_row_5 = 29
        start_row_6 = 32
        start_row_7 = 33
        start_row_8 = 35
        if O29 != 0:
            ws.cell(row=start_row_1, column=start_col_1, value=O29)
        if O32 != 0:
            ws.cell(row=start_row_2, column=start_col_1, value=O32)
        if O33 != 0:
            ws.cell(row=start_row_3, column=start_col_1, value=O33)
        if O35 != 0:
            ws.cell(row=start_row_4, column=start_col_1, value=O35)
        if P29 != 0:
            ws.cell(row=start_row_5, column=start_col_2, value=P29)
        if P32 != 0:
            ws.cell(row=start_row_6, column=start_col_2, value=P32)
        if P33 != 0:
            ws.cell(row=start_row_7, column=start_col_2, value=P33)
        if P35 != 0:
            ws.cell(row=start_row_8, column=start_col_2, value=P35)

        O7 = P29 + P32 + P33 + P35

        if O7 != 0:
            ws.cell(row=7, column=15, value=O7)
        if O6 != 0:
            ws.cell(row=6, column=15, value=O6)
        if F6 != 0:
            ws.cell(row=6, column=6, value=F6)

        wb.save(java_remediation_path)

        print('Analysis Summary')

        # Analysis Summary : java_high_level_report
        wb = load_workbook(java_high_level_path)
        ws = wb['Analysis Summary']
        start_col_1 = 15
        start_row_1 = 29
        start_row_2 = 32
        start_row_3 = 33
        start_row_4 = 35
        start_col_2 = 16
        start_row_5 = 29
        start_row_6 = 32
        start_row_7 = 33
        start_row_8 = 35
        if O29 != 0:
            ws.cell(row=start_row_1, column=start_col_1, value=O29)
        if O32 != 0:
            ws.cell(row=start_row_2, column=start_col_1, value=O32)
        if O33 != 0:
            ws.cell(row=start_row_3, column=start_col_1, value=O33)
        if O35 != 0:
            ws.cell(row=start_row_4, column=start_col_1, value=O35)
        if P29 != 0:
            ws.cell(row=start_row_5, column=start_col_2, value=P29)
        if P32 != 0:
            ws.cell(row=start_row_6, column=start_col_2, value=P32)
        if P33 != 0:
            ws.cell(row=start_row_7, column=start_col_2, value=P33)
        if P35 != 0:
            ws.cell(row=start_row_8, column=start_col_2, value=P35)

        O7 = P29 + P32 + P33 + P35

        if O7 != 0:
            ws.cell(row=7, column=15, value=O7)
        if O6 != 0:
            ws.cell(row=6, column=15, value=O6)
        if F6 != 0:
            ws.cell(row=6, column=6, value=F6)

        wb.save(java_high_level_path)

        print("Sheet : Source Code Impacted : CREATED")
        print("End 16 : Source Code Impacted\n")

        print("Start 17 : Other inventory")
        other_inventory_combined_df = pd.DataFrame()
        # Convert check_files.out
        check_files_dir_path = '/usr/u2l/u2l_backend/projects/'+ file_name +'/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/check_files.out'
        other_inventory_df1 = None
        if os.path.isfile(check_files_dir_path) and os.path.getsize(check_files_dir_path) > 0:
            other_inventory_df1 = pd.read_csv(check_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df1.index = other_inventory_df1.index + 1
            unique_df1 = other_inventory_df1[~other_inventory_df1.isin(other_inventory_combined_df)].dropna()
            other_inventory_combined_df = other_inventory_combined_df.append(unique_df1)
            print(other_inventory_df1)

        # Convert exclude_files.out
        exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/exclude_files.out'
        other_inventory_df2 = None
        if os.path.isfile(exclude_files_dir_path) and os.path.getsize(exclude_files_dir_path) > 0:
            other_inventory_df2 = pd.read_csv(exclude_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
            other_inventory_df2.index = other_inventory_df2.index + 1
            unique_df2 = other_inventory_df2[~other_inventory_df2.isin(other_inventory_combined_df)].dropna()
            other_inventory_combined_df = other_inventory_combined_df.append(unique_df2)
            print(other_inventory_df2)

        other_inventory_combined_df.drop_duplicates(inplace=True)
        print(other_inventory_combined_df)

        writer = pd.ExcelWriter(java_high_level_path, engine='openpyxl', mode='a')
        book = load_workbook(java_high_level_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
        writer.save()
        print('java_high_level other inventory added')

        writer = pd.ExcelWriter(java_inventory_path, engine='openpyxl', mode='a')
        book = load_workbook(java_inventory_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
        writer.save()
        print('java_inventory other inventory added')

        writer = pd.ExcelWriter(java_remediation_path, engine='openpyxl', mode='a')
        book = load_workbook(java_remediation_path)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
        writer.save()
        print('java_remediation other inventory added')

        print("Sheet : Other Inventory : CREATED")
        print("End 17 : Other inventory\n")

    return 'success'

@authentication_controller.route('/clogs', methods=['POST'])
# @jwt_required()
def clogs():

    c_report_path = '/usr/u2l/u2l_backend/projects/c_report.xlsx'
    c_remediation_path = '/usr/u2l/u2l_backend/projects/c_remediation_report.xlsx'
    c_inventory_path = '/usr/u2l/u2l_backend/projects/c_inventory_report.xlsx'

    form_source_ver = 'AIX'

    source_versions = ["AIX", "Solaris", "HP-UX"]
    if form_source_ver not in source_versions:
      error_message = {'error' :"Invalid Source Vesion"}
      return jsonify(error_message), 404

    script_path = '/usr/u2l/u2l_backend/scan_source.sh'
    file_path = '/usr/u2l/u2l_backend/UNIX_HPE'
    source_ver = form_source_ver
    target_ver = 'RHEL'
    command = ['sh', script_path, file_path, source_ver, target_ver]
    process = subprocess.Popen(command, stdout=subprocess.PIPE)
    output = process.communicate()[0].decode().strip()
    lines = output.split('\n')
    print(lines)

    ic_lines = []
    with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
        for line in file:
            if line.strip().startswith('/usr/'):
                ic_lines.append(line)

    with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
        file.write(''.join(ic_lines))

    ic_lines = []
    with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
        for line in file:
            modified_line = line.replace(':', ',', 1)
            ic_lines.append(modified_line)

    with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
        file.write(''.join(ic_lines))

    final_report_path = '/usr/u2l/u2l_backend/final_report.csv'
    if os.path.isfile(final_report_path) and os.path.getsize(final_report_path) > 0:
        df = pd.read_csv(final_report_path, sep=',', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        df.columns = ["FILE PATH","LINE NUMBER","SOURCE CODE","ERROR MESSAGE","REMEDIATION"]
        df = df.drop(df.index[0])
        O28 = df['FILE PATH'].nunique()
        P28 = df['FILE PATH'].count()
        print(df)

    generate_workbook(df, c_report_path,'Source Scan Remedy Detail')

    source_path = "/usr/u2l/u2l_backend/UNIX_HPE"
    stdout, stderr = run_cppcheck(source_path)

    with open("cppcheck_output.txt", "w") as file:
        file.write(stderr)
    
    with open('cppcheck_output.txt', 'r') as file:
        content = file.readlines()

    file_paths = []
    line_numbers = []
    error_messages = []
    error_lines = []

    pattern = re.compile(r'^(.*?):(\d+):\d+: (.*?): (.*)$')

    current_error_lines = []

    for line in content:
        match = pattern.match(line)
        if match:
            if current_error_lines:
                error_lines.append(current_error_lines)
                current_error_lines = []
            
            file_paths.append(match.group(1))
            line_numbers.append(int(match.group(2)))
            error_messages.append(match.group(4))
        else:
            current_error_lines.append(line)

    if current_error_lines:
        error_lines.append(current_error_lines)

    data = {
        'File Path': file_paths,
        'Line Number': line_numbers,
        'Error Message': error_messages,
        'Error Lines': [''.join(lines) for lines in error_lines]
    }

    df = pd.DataFrame(data)
    df.columns = ['FILE PATH', 'LINE NUMBER', 'ERROR MESSAGE', 'ERROR LINES']
    O29 = df['FILE PATH'].nunique()
    P29 = df['FILE PATH'].count()

    df.reset_index(drop=True, inplace=True)
    df.index = df.index + 1

    book = load_workbook(c_report_path)
    writer = pd.ExcelWriter(c_report_path, engine='openpyxl')
    writer.book = book
    sheet_name = 'Compilation Warning Detail'
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, header=False)
    writer.save()

    sheet_transfer('Source Scan Remedy Detail', '1. Source Scan Remedy Detail', 6, 1, c_report_path, c_remediation_path)
    sheet_transfer('Compilation Warning Detail', '2. Compilation Warning Detail', 6, 1, c_report_path, c_remediation_path)
    sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, c_report_path, c_inventory_path)
    sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, c_report_path, c_remediation_path)

    print("Start 13 : Source Code Inventory")
    regex = r'^.*$'
    source_code_java_path = '/usr/u2l/u2l_backend/projects//Log_Source_Analysis_20230907.log'
    print(source_code_java_path)
    if os.path.isfile(source_code_java_path) and os.path.getsize(source_code_java_path) > 0:
        
        with open(source_code_java_path, 'r') as f:
            data = f.read()
        lines = data.split('\n')
        log_entries = []
        for line in lines:
            match = re.match(regex, line)
            if match:
                log_entries.append(match.group())
        df = pd.DataFrame(log_entries, columns=['log_message'])

        start_str = '###########  Execution 0001_stepcounter Starts  ###########'
        end_str = '###########  Execution 0001_stepcounter Ends ###########'
        start_df = df[df['log_message'].str.contains(start_str)]
        end_df = df[df['log_message'].str.contains(end_str)]

        start_index = start_df.index[0]
        end_index = end_df.index[0]
        extracted_df = df.iloc[start_index+1:end_index]

        result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
        result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
        result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
        result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
        result_df = result_df.drop(['log_message', 'col2'], axis=1)
        result_df = result_df.reset_index(drop=True)
        result_df.index += 1

        artefact_type_dict = result_df['Type'].unique()                                     # ['C/C++' 'CShell' 'KShell']
        type_counts = result_df['Type'].value_counts()
        type_counts_dict = type_counts.to_dict()                                            # {'C/C++': 690, 'KShell': 107, 'CShell': 1}
        F6 = sum(type_counts_dict.values())
        result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
        autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
        autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
        O6 = sum(autual_by_type_dict.values())
        result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
        total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
        total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'C/C++': 958392, 'CShell': 9, 'KShell': 9689}

        # generate_workbook(result_df, c_report_path,'Source Code Inventory')
        print('Transfer Completed !!')
    
    wb = load_workbook(c_remediation_path)
    print('1')
    ws = wb['Environment Information']
    print('2')
    start_row = 22
    start_col = 2
    col_actual = 2
    col_total = 3

    for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
        col_idx = start_col
        ws.cell(row=row_idx, column=col_idx, value=artefact_type)

        value = type_counts_dict.get(artefact_type, 0)
        col_idx = start_col + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

        actual_value = autual_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_actual
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

        actual_value = total_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_total
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

    wb.save(c_remediation_path)

    wb = load_workbook(c_inventory_path)
    print('1')
    ws = wb['Scope of Artefacts']
    print('2')
    start_row = 2
    start_col = 1
    col_total = 2

    for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
        col_idx = start_col
        ws.cell(row=row_idx, column=col_idx, value=artefact_type)

        value = type_counts_dict.get(artefact_type, 0)
        col_idx = start_col + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

        actual_value = total_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_total
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

    wb.save(c_inventory_path)

    summary_sheet_generation(c_remediation_path, '2. Compilation Warning Detail', '2. Compilation Warning Summary', 5, 3, c_remediation_path, 7, 2)
    summary_sheet_generation(c_remediation_path, '1. Source Scan Remedy Detail', '1. Source Scan Remedy Summary', 5, 5, c_remediation_path, 7, 2)

    print("Sheet : Source Code Inventory : CREATED")
    print("End 13 : Source Code Inventory")

    print("Start 17 : Other inventory")
    # Convert check_files.out
    # check_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/check_files.out'
    check_files_dir_path = '/usr/u2l/u2l_backend/projects/check_files.out'
    other_inventory_df1 = None
    if os.path.isfile(check_files_dir_path) and os.path.getsize(check_files_dir_path) > 0:
        other_inventory_df1 = pd.read_csv(check_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df1.index = other_inventory_df1.index + 1
        print(other_inventory_df1)

    # Convert exclude_files.out
    # exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/exclude_files.out'
    exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/exclude_files.out'
    other_inventory_df2 = None
    if os.path.isfile(exclude_files_dir_path) and os.path.getsize(exclude_files_dir_path) > 0:
        other_inventory_df2 = pd.read_csv(exclude_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df2.index = other_inventory_df2.index + 1
        print(other_inventory_df2)

    # Convert c_files.out
    # c_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/c_files.out'
    c_files_dir_path = '/usr/u2l/u2l_backend/projects/c_files.out'
    other_inventory_df3 = None
    if os.path.isfile(c_files_dir_path) and os.path.getsize(c_files_dir_path) > 0:
        other_inventory_df3 = pd.read_csv(c_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df3.index = other_inventory_df3.index + 1
        print(other_inventory_df3)

    # Convert csh_list.out
    # csh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/csh_list.out'
    csh_list_dir_path = '/usr/u2l/u2l_backend/projects/csh_list.out'
    other_inventory_df4 = None
    if os.path.isfile(csh_list_dir_path) and os.path.getsize(csh_list_dir_path) > 0:
        other_inventory_df4 = pd.read_csv(csh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df4.index = other_inventory_df4.index + 1
        print(other_inventory_df4)    

    # Convert ksh_list.out
    # ksh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ksh_list.out'
    ksh_list_dir_path = '/usr/u2l/u2l_backend/projects/ksh_list.out'
    other_inventory_df5 = None
    if os.path.isfile(ksh_list_dir_path) and os.path.getsize(ksh_list_dir_path) > 0:
        other_inventory_df5 = pd.read_csv(ksh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df5.index = other_inventory_df5.index + 1
        print(other_inventory_df5)

    # Convert ot2_list.out
    # ot2_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ot2_list.out'
    ot2_list_dir_path = '/usr/u2l/u2l_backend/projects/ot2_list.out'
    other_inventory_df6 = None
    if os.path.isfile(ot2_list_dir_path) and os.path.getsize(ot2_list_dir_path) > 0:
        other_inventory_df6 = pd.read_csv(ot2_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df6.index = other_inventory_df6.index + 1
        print(other_inventory_df6)   

    
    other_inventory_combined_df = pd.concat([other_inventory_df1, other_inventory_df2, other_inventory_df3, other_inventory_df4, other_inventory_df5, other_inventory_df6], ignore_index=True)
    other_inventory_combined_df = other_inventory_combined_df.drop_duplicates()
    print(other_inventory_combined_df)

    writer = pd.ExcelWriter(c_inventory_path, engine='openpyxl', mode='a')
    book = load_workbook(c_inventory_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
    writer.save()
    print('c_inventory other inventory added')

    writer = pd.ExcelWriter(c_remediation_path, engine='openpyxl', mode='a')
    book = load_workbook(c_remediation_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
    writer.save()
    print('c_remediation other inventory added')

    print("Sheet : Other Inventory : CREATED")
    print("End 17 : Other inventory\n")

    middleware_type = 'Tuxedo'
    if(middleware_type != 'Tuxedo'):
        filename = c_remediation_path
        sheet_name = '6. Middleware-Tuxedo'
        workbook = load_workbook(filename)

        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
            workbook.save(filename)
        print('Tuxedo sheet removed !!!')

    print('Migration Summary')
    wb = load_workbook(c_remediation_path)
    ws = wb['Migration Summary']
    start_col_1 = 15
    start_col_2 = 16
    start_row_1 = 28
    start_row_2 = 29
    if O28 != 0:
        ws.cell(row=start_row_1, column=start_col_1, value=O28)
    if O29 != 0:
        ws.cell(row=start_row_2, column=start_col_1, value=O29)
    if P28 != 0:
        ws.cell(row=start_row_1, column=start_col_2, value=P28)
    if P29 != 0:
        ws.cell(row=start_row_2, column=start_col_2, value=P29)

    O7 = P28 + P29
    F7 = O28 + O29

    if O7 != 0:
        ws.cell(row=7, column=15, value=O7)
    if O6 != 0:
        ws.cell(row=6, column=15, value=O6)
    if F7 != 0:
        ws.cell(row=7, column=6, value=F7)
    if F6 != 0:
        ws.cell(row=6, column=6, value=F6)

    wb.save(c_remediation_path)

    return 'clog generated'

@authentication_controller.route('/apitest', methods=['POST'])
def apitest():

    # Convert javaUniqRulesGrepped.txt
    
    javaUniqRulesGrepped_dir_path = '/usr/u2l/u2l_backend/javaUniqRulesGrepped.txt'
    if os.path.isfile(javaUniqRulesGrepped_dir_path) and os.path.getsize(javaUniqRulesGrepped_dir_path) > 0:
        ic_lines = []
        with open(javaUniqRulesGrepped_dir_path, 'r') as file:
            for line in file:
                cleaned_line = re.sub(r'\t*$', '', line)
                if cleaned_line.strip().startswith('IC_'):
                    ic_lines.append(cleaned_line)

        with open(javaUniqRulesGrepped_dir_path, 'w') as file:
            file.write(''.join(ic_lines))

        with open(javaUniqRulesGrepped_dir_path, 'r') as file:
            lines = [re.sub(r'\t*$', '', line) for line in file]

        with open(javaUniqRulesGrepped_dir_path, 'w') as file:
            file.write(''.join(lines))
        df = pd.read_csv(javaUniqRulesGrepped_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        df.index = df.index + 1
        df.columns = ['RULE ID', 'ENTITY', 'PACKAGE', 'OBJECT', 'REMEDY']
        print(df)
    #     generate_workbook(df, java_report_path,'Possible Import Remediations')
    #     sheet_transfer('Possible Import Remediations', '7. Possible Import Remediation', 6, 1, java_report_path, java_remediation_path)
    #     summary_sheet_generation(java_remediation_path, '7. Possible Import Remediation', '7. Import Remediation Summary', 5, 6, java_remediation_path, 8, 2)
    # logger.info("SHEET : 7. Possible Import Remediation : CREATED")
    # logger.info("SHEET : 7. Import Remediation Summary : CREATED")
    # logger.info("END 5 : javaUniqRulesGrepped.txt\n")

    return 'success'

# @authentication_controller.route('/extract_middleware', methods=['POST'])
# def extract_middleware():

#     if(tool_analysis_type == 'javaanalysis'):
#         middleware_json_path = '/usr/u2l/u2l_backend/project/' + file_name + '/middleware_list/java_middleware_list.json'
#     elif(tool_analysis_type == 'shellanalysis'):
#         middleware_json_path = '/usr/u2l/u2l_backend/project/' + file_name + '/middleware_list/shell_middleware_list.json'
#     elif(tool_analysis_type == 'canalysis'):
#         middleware_json_path = '/usr/u2l/u2l_backend/project/' + file_name + '/middleware_list/c_middleware_list.json'

#     middleware_file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/' + file_name + '.zip'
#     output_folder = '/usr/u2l/u2l_backend/projects/' + file_name
#     results = java_middleware_extraction(middleware_file_path, output_folder, middleware_json_path)

#     # if (results == 'success'):
#     #     return jsonify({'message': 'Middleware extraction completed', 'results': results}), 200
#     # else:
#     #     return jsonify({'message': 'No results found.'}), 404

#     return 'Success'

# Java SE 10-19 doc for all deprecated items
@authentication_controller.route('/collectdata', methods=['POST'])
def collectdata():

    json_data = request.get_json()
    java_docs_version = json_data['java_doc_version']
    # java_docs_version = request.form['java_doc_version']
    java_docs_version_int = int(java_docs_version)
    package_name = 'java'

    if ((java_docs_version_int > 10) and (java_docs_version_int < 20)):
        # 11, 12, 13, 14, 15, 16, 17, 18, 19
        url = 'https://docs.oracle.com/en/java/javase/' + java_docs_version + '/docs/api/deprecated-list.html'
    elif((java_docs_version_int > 5) and (java_docs_version_int < 11)):
        # 6, 7, 8, 9, 10
        url = 'https://docs.oracle.com/javase/' + java_docs_version + '/docs/api/deprecated-list.html'
    else:
        return 'Please Enter Valid Java Version !!'

    try:

        deprecated_excel_file_path = '/usr/u2l/u2l_backend/deprecated_data.xlsx'
        sheet_name = 'JDK_Depricated'
        response = requests.get(url)
        response.raise_for_status()

        soup = BeautifulSoup(response.content, 'html.parser')
  
        # JDK SE 6, 7, 8, 9
        if ((java_docs_version_int > 5) and (java_docs_version_int < 10)):
            if (java_docs_version_int == 6):
                deprecated_item_names = [item.text.strip() for item in soup.find_all('tr', class_='TableRowColor')]
            elif ((java_docs_version_int == 7) or (java_docs_version_int == 8)):
                deprecated_item_names = [item.text.strip() for item in soup.find_all('td', class_='colOne')]
            elif (java_docs_version_int == 9):

                deprecated_item_names_alt = [item.text.strip() for item in soup.find_all('tr', class_='altColor')]
                df1 = createDf(deprecated_item_names_alt, java_docs_version, 'jdk')

                deprecated_item_names_row = [item.text.strip() for item in soup.find_all('tr', class_='rowColor')]
                df2 = createDf(deprecated_item_names_row, java_docs_version, 'jdk')

                df = df1.append(df2, ignore_index=True)
                java_doc_sheet(deprecated_excel_file_path, sheet_name, df)

                return f'Data collected and saved to successfully'

            df = createDf(deprecated_item_names, java_docs_version, 'jdk')
            print(df)
            java_doc_sheet(deprecated_excel_file_path, sheet_name, df)

            return f'Data collected and saved to successfully'

        # JDK SE 10, 11, 12, 13, 14, 15
        elif ((java_docs_version_int > 8) and (java_docs_version_int < 16)):

            deprecated_item_names = soup.find_all('th', class_='colDeprecatedItemName')
            deprecation_comments = soup.find_all('td', class_='colLast')
            df = convert_to_dataframe_move_to_excel(deprecated_item_names, deprecation_comments, java_docs_version, package_name)

            ## OpenAI Implementation

            java_doc_sheet(deprecated_excel_file_path, sheet_name, df)

            return f'Data collected and saved to successfully'

        # JDK SE 16, 17, 18, 19
        elif ((java_docs_version_int > 15) and (java_docs_version_int < 20)):

            if(java_docs_version_int == 16):
                deprecated_item_names = soup.find_all('div', class_=['col-deprecated-item-name', 'col-first'])
            else:
                deprecated_item_names = soup.find_all('div', class_=['col-summary-item-name', 'col-first'])

            deprecation_comments = soup.find_all('div', class_='col-last')
            df = convert_to_dataframe_move_to_excel(deprecated_item_names, deprecation_comments, java_docs_version, package_name)
            java_doc_sheet(deprecated_excel_file_path, sheet_name, df)

            return f'Data collected and saved to successfully'

        return 'Requested URL not found !!'

    except requests.exceptions.RequestException as e:
        return jsonify({'error': str(e)})

@authentication_controller.route('/javalib', methods=['GET'])
def javalib():
    populate_database()
    packages = Package.query.all()
    package_data = []
    for package in packages:
        package_versions = [version.version for version in package.versions] if package.versions else []
        package_info = {
            'package_name': package.name,
            'package_url': package.url if package.url else "",
            'package_version': package_versions
        }
        package_data.append(package_info)
    return jsonify(package_data)

@authentication_controller.route('/collect_other_libraries', methods=['POST'])
def collect_other_libraries():

    try:
        
        print('Inside collect other libraries')
        urls_dict = {
        'Angus Activation': 'https://eclipse-ee4j.github.io/angus-activation/api/deprecated-list.html',
        'ECJ': 'https://javadoc.io/doc/org.eclipse.jdt.core.compiler/ecj/latest/deprecated-list.html',
        'JavaHamcrest': 'https://hamcrest.org/JavaHamcrest/javadoc/2.2/deprecated-list.html',
        'Mchange commons java': 'https://www.mchange.com/projects/mchange-commons-java-versions/{package_version}/apidocs/deprecated-list.html',
        'Micrometer Core': 'https://javadoc.io/doc/io.micrometer/micrometer-core/latest/deprecated-list.html',
        'okhttp': 'https://square.github.io/okhttp/1.x/okhttp/deprecated-list.html',
        'zip4j': 'https://javadoc.io/doc/net.lingala.zip4j/zip4j/latest/deprecated-list.html',
        'Apache Tomcat': 'https://tomcat.apache.org/tomcat-{package_version}-doc/api/deprecated-list.html',
        'Apache ANT': 'https://ant.apache.org/manual/api/deprecated-list.html',
        'aspectjweaver': 'https://javadoc.io/doc/org.aspectj/aspectjweaver/latest/deprecated-list.html',
        'aspectjrt': 'https://javadoc.io/doc/org.aspectj/aspectjrt/latest/deprecated-list.html',
        'Axis': 'https://axis.apache.org/axis/java/apiDocs/deprecated-list.html',
        'Bouncycastle': 'https://www.bouncycastle.org/docs/pkixdocs1.5on/index.html?deprecated-list.html',
        'cglib-nodep': 'https://javadoc.io/doc/cglib/cglib-nodep/latest/deprecated-list.html',
        'Commons Beanutils': 'https://commons.apache.org/proper/commons-beanutils/apidocs/deprecated-list.html',
        'Commons Configuration': 'https://commons.apache.org/proper/commons-configuration/apidocs/deprecated-list.html',
        'Commons Dbcp': 'https://commons.apache.org/proper/commons-dbcp/apidocs/deprecated-list.html',
        'Commons Digester': 'https://commons.apache.org/proper/commons-digester/apidocs/deprecated-list.html',
        'Commons Discovery': 'https://commons.apache.org/dormant/commons-discovery/apidocs/deprecated-list.html',
        'Commons Lang': 'https://commons.apache.org/proper/commons-lang/javadocs/api-2.6/index.html?deprecated-list.html',
        'Commons Collections': 'https://commons.apache.org/proper/commons-collections/apidocs/deprecated-list.html',
        'Commons Codec': 'https://commons.apache.org/proper/commons-codec/apidocs/deprecated-list.html',
        'Apache Commons IO': 'https://javadoc.scijava.org/Apache-Commons-IO/deprecated-list.html',
        'Commons Pool': 'https://commons.apache.org/proper/commons-pool/apidocs/deprecated-list.html',
        'Commons Logging': 'https://commons.apache.org/proper/commons-logging/apidocs/deprecated-list.html',
        'Diffutils': 'https://javadoc.io/static/io.github.java-diff-utils/java-diff-utils/4.5/deprecated-list.html',
        'Geronimo': 'https://geronimo.apache.org/apidocs/{package_version}/deprecated-list.html',
        'Groovy-lang': 'https://docs.groovy-lang.org/latest/html/api/deprecated-list.html',
        'Hazelcast': 'https://docs.hazelcast.org/docs/latest/javadoc/deprecated-list.html',
        'HTTPcore': 'https://javadoc.io/doc/org.apache.httpcomponents/httpcore/latest/deprecated-list.html',
        'HTTPclient': 'https://javadoc.io/doc/org.apache.httpcomponents/httpclient/latest/index.html',
        'Imap': 'https://javadoc.io/doc/com.sun.mail/imap/latest/deprecated-list.html',
        'IText': 'https://javadoc.io/doc/com.itextpdf/itextpdf/latest/deprecated-list.html',
        'Jasper Reports': 'https://jasperreports.sourceforge.net/api/deprecated-list.html',
        'Jaxrpc': 'https://docs.jboss.org/jbossas/javadoc/4.0.3SP1/jaxrpc/deprecated-list.html',
        'Jcifs': 'https://javadoc.io/docs/org.codelibs/jcifs/latest/deprecated-list.html',
        'Jcommon': 'https://www.jfree.org/jcommon/api/index.html?deprecated-list.html',
        'Jdom': 'http://www.jdom.org/docs/apidocs/deprecated-list.html',
        'Jfreechart': 'https://www.jfree.org/jfreechart/javadoc/deprecated-list.html',
        'Jsch Documentation': 'https://epaul.github.io/jsch-documentation/javadoc/deprecated-list.html',
        'Juddi Client': 'https://juddi.apache.org/juddi-client/apidocs/deprecated-list.html',
        'Apache OpenJPA': 'https://openjpa.apache.org/builds/{package_version}/apidocs/deprecated-list.html',
        'Mybatis': 'https://mybatis.org/mybatis-3/apidocs/deprecated-list.html',
        'Mybatis Spring': 'https://mybatis.org/spring/apidocs/deprecated-list.html',
        'Opencsv': 'https://javadoc.io/doc/com.opencsv/opencsv/latest/deprecated-list.html',
        'POI': 'https://poi.apache.org/apidocs/{package_version}/deprecated-list.html',
        'Quartz scheduler': 'https://www.quartz-scheduler.org/api/{package_version}/deprecated-list.html',
        'slf4j': 'https://www.slf4j.org/api/deprecated-list.html',
        'wsdl4j': 'https://javadoc.io/doc/wsdl4j/wsdl4j/latest/index.html',
        'XMLBeans': 'https://xmlbeans.apache.org/docs/{package_version}/reference/deprecated-list.html'
        }

        print('Test 1')
        json_data = request.get_json()
        package_name = json_data['package_name']
        package_version = json_data['package_version']
        package_url = json_data['package_url']
        print('Test 2')

        if (package_url == ''):
            print('Test 3')
            if(package_name in urls_dict.keys()):
                print('Test 4')
                url = urls_dict[package_name]
                if '{package_version}' in url:
                    # If package version is not provided, extract available versions
                    if not package_version:
                        available_versions = extract_versions(package_name)
                        # Iterate over available versions and collect data
                        for version in available_versions:
                            version_url = url.replace('{package_version}', version)
                            result = deprecated_data_sheet(version_url, package_name, version)
                            print(result)
                    else:
                        print('Test 5')
                        url = url.replace('{package_version}', package_version)
                result = deprecated_data_sheet(url, package_name, package_version)
            else:
                print('Test 6')
                print("Package not found in the list")
        else:
            print('Test 7')
            result = deprecated_data_sheet(package_url, package_name, 1)
            print(result)
        return 'Package found successfully : Updated'

    except requests.exceptions.RequestException as e:
        return jsonify({'error': str(e)})

@authentication_controller.route('/openai', methods=['POST'])
def openai():
    import os
    import openai

    openai.api_type = "azure"
    openai.api_base = "https://azure-openai-demo1.openai.azure.com/"
    openai.api_version = "2023-07-01-preview"
    openai.api_key = "c2c49c25ad2946b98d9e7ca876f7f7cb"
    message_text = [
    {"role": "system", "content": "You are an expert middleware architect."},
    {"role": "user", "content": "2+2"}
    ]
    completion = openai.ChatCompletion.create(
        engine="gpt-4-32k",
        messages = message_text,
        temperature=0.0,
        max_tokens=800,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None
        )
    result = completion.choices[0].message['content']
    print(result)
    return result

@authentication_controller.route('/check_website', methods=['POST'])
def check_website():
    data = request.get_json()

    if data and 'url' in data:
        website_url = data['url']
        try:
            response = requests.head(website_url, timeout=10)
            if response.status_code == 200:
                return jsonify({'status': 'reachable'}), 200
            else:
                return jsonify({'status': 'unreachable', 'reason': 'HTTP Status Code: {}'.format(response.status_code)})
        except requests.ConnectionError:
            return jsonify({'status': 'unreachable', 'reason': 'Connection Error'}), 404
        except requests.Timeout:
            return jsonify({'status': 'unreachable', 'reason': 'Timeout'}), 404
        except Exception as e:
            return jsonify({'status': 'unreachable', 'reason': str(e)}), 404
    else:
        return jsonify({'error': 'Missing or invalid URL in the request body'}), 404

@authentication_controller.route('/extract_middleware', methods=['POST'])
def extract_middleware():
    try:
        print(request)
        print('Test')
        tool_analysis_type = request.form['tool_analysis_type']
        print('Test1')
        file = request.files['file_name']
        c_analysis_types = ['C', 'C++', 'Pro*C', 'C/Pro*C', 'C++/Pro*C']
        print('Test1')
        if(tool_analysis_type == 'Java'):
            tool_analysis_type = 'javaanalysis'
            print('Test2')
        elif(tool_analysis_type in c_analysis_types):
            if(tool_analysis_type == 'Pro*C'):
                tool_analysis_type = 'canalysis'
            elif(tool_analysis_type == 'C' or tool_analysis_type == 'C++'):
                print('Accepting C Parameters!')
                tool_analysis_type = 'canalysis'
            elif(tool_analysis_type == 'C/Pro*C' or tool_analysis_type == 'C++/Pro*C'):
                tool_analysis_type = 'canalysis'
        elif(tool_analysis_type == 'Shell'):  
            tool_analysis_type = 'shellanalysis'

        analysis_types = ['javaanalysis', 'canalysis', 'shellanalysis']
        if tool_analysis_type not in analysis_types:
            error_message = {'error' :"Invalid analysis type"}
            return jsonify(error_message), 404

        file_name = file.filename
        temp_tuple = os.path.splitext(file_name)
        if temp_tuple[1] != '.zip':
            error_message = {'error' :"Upload .zip files only"}
            return jsonify(error_message), 404

        cur_date = datetime.datetime.now()

        file_name = temp_tuple[0]

        year_month_day = str(cur_date).split(" ")
        global yMD
        yMD = year_month_day[0].replace("-", "")

        hour_min_sec = os.path.splitext(year_month_day[1])
        global hMS
        hMS = hour_min_sec[0].replace(":", "")

        file_name = file_name + "_" + yMD + "_" + hMS 
        directory = '/usr/u2l/u2l_backend/midextract/' + file_name
        os.makedirs(directory)
        file.save(os.path.join(directory, file_name +'.zip'))

        zip_file_path = directory + '/' + file_name + '.zip'

        if(tool_analysis_type == 'javaanalysis'):
            middleware_json_path = '/usr/u2l/u2l_backend/middleware_list/java_middleware_list.json'
            print('Test3')
        elif(tool_analysis_type == 'shellanalysis'):
            middleware_json_path = '/usr/u2l/u2l_backend/middleware_list/shell_middleware_list.json'
        elif(tool_analysis_type == 'canalysis'):
            middleware_json_path = '/usr/u2l/u2l_backend/middleware_list/c_middleware_list.json'

        output_folder = '/usr/u2l/u2l_backend/middleware-output/java-middleware'
        print('Before result')
        results = java_middleware_extraction(zip_file_path, output_folder, middleware_json_path)
        print('After result')
        if (results):
            print('In result')
            return jsonify({'message': 'Middleware extraction completed', 'results': results}), 200
        else:
            return jsonify({'message': 'No results found.'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500     

    return 'Success'

@authentication_controller.route('/invoke_container2', methods=['POST'])
def invoke_container2():
    import docker
    client = docker.from_env()
    image, logs = client.images.build(path='dockerfile/', tag='my-image')
    java_version = '8'
    command = f'/usr/lib/jvm/java-{java_version}-openjdk-amd64/bin/javac MyFile.java'
    container = client.containers.run(image, command=command, remove=True)

    print(container.logs().decode('utf-8'))

@authentication_controller.route('/invoke_container', methods=['POST'])
def invoke_container():
    java_source_code = request.json.get('source_code')
    java_version = request.json.get('java_version')

    if java_version == "8":
        docker_command = ["docker", "run", "--rm", "-i", "openjdk:8", "javac", "-"]
    elif java_version == "11":
        docker_command = ["docker", "run", "--rm", "-i", "openjdk:11", "javac", "-"]
    else:
        return jsonify({"error": "Unsupported Java version"}), 400

    try:
        result = subprocess.run(docker_command, input=java_source_code.encode(), capture_output=True, text=True, check=True)
        return jsonify({"stdout": result.stdout, "stderr": result.stderr})
    except subprocess.CalledProcessError as e:
        return jsonify({"error": "Compilation failed", "stderr": e.stderr}), 500

@authentication_controller.route('/compile_java2', methods=['POST'])
def compile_java2():
    from javax.tools import DiagnosticCollector, JavaCompiler, ToolProvider
    try:
        source_code = request.form['source_code']
        compiler = ToolProvider.getSystemJavaCompiler()
        diagnostics = DiagnosticCollector()
        file_manager = compiler.getStandardFileManager(diagnostics, None, None)
        compilation_units = [SimpleJavaFileObject('Main.java', JavaFileObject.Kind.SOURCE, source_code)]
        compiler.getTask(None, file_manager, diagnostics, None, None, compilation_units).call()
        
        errors = []
        for diagnostic in diagnostics.getDiagnostics():
            error = {
                'message': str(diagnostic.getMessage(None)),
                'line': diagnostic.getLineNumber(),
                'column': diagnostic.getColumnNumber()
            }
            errors.append(error)
        return jsonify({'message': 'Compilation successful'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@authentication_controller.route('/compile_java', methods=['POST'])
def compile_java():
    try:
        jdk_version = request.json.get('jdk_version')
        jdk_version = '11'
        # java_home = subprocess.check_output(['which', f'java-{jdk_version}']).decode().strip()
        # if not java_home:
        #     return jsonify({'error': f'JDK version {jdk_version} not found'}), 400

        # os.environ["JAVA_HOME"] = java_home
        # source_folder = request.json.get('source_folder')
        source_folder = "/usr/u2l/u2l_backend/pet-store-master"

        for file_name in os.listdir("/usr/u2l/u2l_backend/pet-store-master"):
            if file_name.endswith('.java'):
                print('1')
                file_path = os.path.join(source_folder, file_name)
                compile_process = subprocess.Popen(['javac', file_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                stdout, stderr = compile_process.communicate()
                if compile_process.returncode != 0:
                    return jsonify({'error': f'Compilation failed for {file_name}', 'stderr': stderr.decode()}), 400

        return jsonify({'message': 'Compilation successful'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@authentication_controller.route('/creports', methods=['POST'])
def creports():
    c_report_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'
    c_remediation_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_remediation_report.xlsx'
    c_inventory_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_inventory_report.xlsx'

    form_source_ver = 'HP-UX'
    # redis_client.rpush(redis_queue_name, '7/40::Executing Scan Source Script')
    source_versions = ["AIX", "Solaris", "HP-UX"]
    if form_source_ver not in source_versions:
        error_message = {'error' :"Invalid Source Vesion"}
        return jsonify(error_message), 404

    # file_name = "UNIX_HPE_20231128_054027"
    file_path = re.sub(r'_\d{8}_\d{6}', '', file_name)
    print(file_path)
    source_path = '/usr/u2l/u2l_backend/projects/'+ file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/' + file_path
    print(source_path)
    script_path = '/usr/u2l/u2l_backend/scan_source.sh'
    source_ver = form_source_ver
    target_ver = 'RHEL'
    command = ['sh', script_path, source_path, source_ver, target_ver]
    process = subprocess.Popen(command, stdout=subprocess.PIPE)
    output = process.communicate()[0].decode().strip()
    lines = output.split('\n')
    print(lines)
    # redis_client.rpush(redis_queue_name, '8/40::Completed Executing Scan Source Script')
    # redis_client.rpush(redis_queue_name, '9/40::Generating Final Report')

    ic_lines = []
    with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
        for line in file:
            if line.strip().startswith('/usr/'):
                ic_lines.append(line)

    with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
        file.write(''.join(ic_lines))

    ic_lines = []
    with open('/usr/u2l/u2l_backend/final_report.csv', 'r') as file:
        for line in file:
            modified_line = line.replace(':', ',', 1)
            ic_lines.append(modified_line)

    with open('/usr/u2l/u2l_backend/final_report.csv', 'w') as file:
        file.write(''.join(ic_lines))

    final_report_path = '/usr/u2l/u2l_backend/final_report.csv'
    if os.path.isfile(final_report_path) and os.path.getsize(final_report_path) > 0:
        df = pd.read_csv(final_report_path, sep='?', engine='python', header=None, index_col=None)
        df.columns = ["FILE PATH","LINE NUMBER","SOURCE CODE","ERROR MESSAGE","REMEDIATION"]
        df = df.drop(df.index[0])
        O28 = df['FILE PATH'].nunique()
        P28 = df['FILE PATH'].count()
        print(df)

    generate_workbook(df, c_report_path,'Source Scan Remedy Detail')
    # redis_client.rpush(redis_queue_name, '10/40::Reports almost ready')

    # redis_client.rpush(redis_queue_name, '11/40::Performing cppcheck !!')

    source_path = '/usr/u2l/u2l_backend/projects/'+ file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/' + file_path
    # source_path = '/usr/u2l/u2l_backend/projects/' + file_name +'/'+ file_name + '.zip'
    stdout, stderr = run_cppcheck(source_path)

    with open("cppcheck_output.txt", "w") as file:
        file.write(stderr)

    with open('cppcheck_output.txt', 'r') as file:
        content = file.readlines()

    file_paths = []
    line_numbers = []
    error_messages = []
    error_lines = []

    pattern = re.compile(r'^(.*?):(\d+):\d+: (.*?): (.*)$')

    current_error_lines = []

    for line in content:
        match = pattern.match(line)
        if match:
            if current_error_lines:
                error_lines.append(current_error_lines)
                current_error_lines = []
            
            file_paths.append(match.group(1))
            line_numbers.append(int(match.group(2)))
            error_messages.append(match.group(4))
        else:
            current_error_lines.append(line)

    if current_error_lines:
        error_lines.append(current_error_lines)

    data = {
        'File Path': file_paths,
        'Line Number': line_numbers,
        'Error Message': error_messages,
        'Error Lines': [''.join(lines) for lines in error_lines]
    }

    df = pd.DataFrame(data)
    df.columns = ['FILE PATH', 'LINE NUMBER', 'ERROR MESSAGE', 'ERROR LINES']
    O29 = df['FILE PATH'].nunique()
    P29 = df['FILE PATH'].count()

    df.reset_index(drop=True, inplace=True)
    df.index = df.index + 1

    book = load_workbook(c_report_path)
    # writer = pd.ExcelWriter(c_report_path, engine='openpyxl')
    # writer.book = book
    sheet_name = 'Compilation Warning Detail'
    sheet = book[sheet_name]
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 0):
            sheet.cell(row=r_idx, column=c_idx+1, value=value)
    book.save(c_report_path)
    # writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    # df.to_excel(writer, sheet_name=sheet_name, startcol=0, startrow=1, header=False)
    # writer.save()
    # redis_client.rpush(redis_queue_name, '12/40::Compilation reports are ready')

    sheet_transfer('Source Scan Remedy Detail', '1. Source Scan Remedy Detail', 6, 1, c_report_path, c_remediation_path)
    sheet_transfer('Compilation Warning Detail', '2. Compilation Warning Detail', 6, 1, c_report_path, c_remediation_path)
    # sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, c_report_path, c_inventory_path)
    # sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, c_report_path, c_remediation_path)

    print("Start 13 : Source Code Inventory")
    # redis_client.rpush(redis_queue_name, '35/40::Generating logs for Inventory')
    regex = r'^.*$'
    source_code_c_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
    print(source_code_c_path)
    if os.path.isfile(source_code_c_path) and os.path.getsize(source_code_c_path) > 0:
        
        with open(source_code_c_path, 'r') as f:
            data = f.read()
        lines = data.split('\n')
        log_entries = []
        for line in lines:
            match = re.match(regex, line)
            if match:
                log_entries.append(match.group())
        df = pd.DataFrame(log_entries, columns=['log_message'])

        start_str = '###########  Execution 0001_stepcounter Starts  ###########'
        end_str = '###########  Execution 0001_stepcounter Ends ###########'
        start_df = df[df['log_message'].str.contains(start_str)]
        end_df = df[df['log_message'].str.contains(end_str)]

        start_index = start_df.index[0]
        end_index = end_df.index[0]
        extracted_df = df.iloc[start_index+1:end_index]

        result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
        result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
        result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
        result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
        result_df = result_df.drop(['log_message', 'col2'], axis=1)
        result_df = result_df.reset_index(drop=True)
        result_df.index += 1

        artefact_type_dict = result_df['Type'].unique()                                     # ['C/C++' 'CShell' 'KShell']
        type_counts = result_df['Type'].value_counts()
        type_counts_dict = type_counts.to_dict()                                            # {'C/C++': 690, 'KShell': 107, 'CShell': 1}
        F6 = sum(type_counts_dict.values())
        result_df['Actual Nr of Lines'] = result_df['Actual Nr of Lines'].astype(int)
        autual_by_type = result_df.groupby('Type')['Actual Nr of Lines'].sum().reset_index()
        autual_by_type_dict = autual_by_type.set_index('Type')['Actual Nr of Lines'].to_dict()    # {'Java': 231, 'KShell': 3716}
        O6 = sum(autual_by_type_dict.values())
        result_df['Total Nr LoC'] = result_df['Total Nr LoC'].astype(int)
        total_by_type = result_df.groupby('Type')['Total Nr LoC'].sum().reset_index()
        total_by_type_dict = total_by_type.set_index('Type')['Total Nr LoC'].to_dict()      # {'C/C++': 958392, 'CShell': 9, 'KShell': 9689}

        generate_workbook(result_df, c_report_path,'Source Code Inventory')

        sheet_transfer('Source Code Inventory', 'Source Code Inventory', 1, 0, c_report_path, c_inventory_path)
        sheet_transfer('Source Code Inventory', 'Source Code Inventory', 6, 1, c_report_path, c_remediation_path)
        print('Transfer Completed !!')
        # redis_client.rpush(redis_queue_name, '36/40::Inventory reports are ready')

    wb = load_workbook(c_remediation_path)
    print('1')
    ws = wb['Environment Information']
    print('2')
    start_row = 22
    start_col = 2
    col_actual = 2
    col_total = 3

    for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
        col_idx = start_col
        ws.cell(row=row_idx, column=col_idx, value=artefact_type)

        value = type_counts_dict.get(artefact_type, 0)
        col_idx = start_col + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

        actual_value = autual_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_actual
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

        actual_value = total_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_total
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

    wb.save(c_remediation_path)

    wb = load_workbook(c_inventory_path)
    print('1')
    ws = wb['Scope of Artefacts']
    print('2')
    start_row = 2
    start_col = 1
    col_total = 2

    for row_idx, artefact_type in enumerate(artefact_type_dict, start=start_row):
        col_idx = start_col
        ws.cell(row=row_idx, column=col_idx, value=artefact_type)

        value = type_counts_dict.get(artefact_type, 0)
        col_idx = start_col + 1
        ws.cell(row=row_idx, column=col_idx, value=value)

        actual_value = total_by_type_dict.get(artefact_type, 0)
        col_idx = start_col + col_total
        ws.cell(row=row_idx, column=col_idx, value=actual_value)

    wb.save(c_inventory_path)

    summary_sheet_generation(c_remediation_path, '2. Compilation Warning Detail', '2. Compilation Warning Summary', 5, 3, c_remediation_path, 7, 2)
    summary_sheet_generation(c_remediation_path, '1. Source Scan Remedy Detail', '1. Source Scan Remedy Summary', 5, 5, c_remediation_path, 7, 2)

    print("Sheet : Source Code Inventory : CREATED")
    print("End 13 : Source Code Inventory")

    print("Start 17 : Other inventory")
    # redis_client.rpush(redis_queue_name, '37/40::Generating logs for other inventory if present')
    yyyy_mmdd = yMD[:4] + '-' + yMD[4:]

    # Convert check_files.out
    check_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/check_files.out'
    other_inventory_df1 = None
    if os.path.isfile(check_files_dir_path) and os.path.getsize(check_files_dir_path) > 0:
        other_inventory_df1 = pd.read_csv(check_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df1.index = other_inventory_df1.index + 1
        print(other_inventory_df1)

    # Convert exclude_files.out
    exclude_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/exclude_files.out'
    other_inventory_df2 = None
    if os.path.isfile(exclude_files_dir_path) and os.path.getsize(exclude_files_dir_path) > 0:
        other_inventory_df2 = pd.read_csv(exclude_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df2.index = other_inventory_df2.index + 1
        print(other_inventory_df2)

    # Convert c_files.out
    c_files_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/c_files.out'
    other_inventory_df3 = None
    if os.path.isfile(c_files_dir_path) and os.path.getsize(c_files_dir_path) > 0:
        other_inventory_df3 = pd.read_csv(c_files_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df3.index = other_inventory_df3.index + 1
        print(other_inventory_df3)

    # Convert csh_list.out
    csh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/csh_list.out'
    other_inventory_df4 = None
    if os.path.isfile(csh_list_dir_path) and os.path.getsize(csh_list_dir_path) > 0:
        other_inventory_df4 = pd.read_csv(csh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df4.index = other_inventory_df4.index + 1
        print(other_inventory_df4)    

    # Convert ksh_list.out
    ksh_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ksh_list.out'
    other_inventory_df5 = None
    if os.path.isfile(ksh_list_dir_path) and os.path.getsize(ksh_list_dir_path) > 0:
        other_inventory_df5 = pd.read_csv(ksh_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df5.index = other_inventory_df5.index + 1
        print(other_inventory_df5)

    # Convert ot2_list.out
    ot2_list_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/srccheck/' + yyyy_mmdd + '/ot2_list.out'
    other_inventory_df6 = None
    if os.path.isfile(ot2_list_dir_path) and os.path.getsize(ot2_list_dir_path) > 0:
        other_inventory_df6 = pd.read_csv(ot2_list_dir_path, sep='\n', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
        other_inventory_df6.index = other_inventory_df6.index + 1
        print(other_inventory_df6)   


    other_inventory_combined_df = pd.concat([other_inventory_df1, other_inventory_df2, other_inventory_df3, other_inventory_df4, other_inventory_df5, other_inventory_df6], ignore_index=True)
    other_inventory_combined_df = other_inventory_combined_df.drop_duplicates()
    print(other_inventory_combined_df)

    writer = pd.ExcelWriter(c_inventory_path, engine='openpyxl', mode='a')
    book = load_workbook(c_inventory_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
    writer.save()
    print('c_inventory other inventory added')

    writer = pd.ExcelWriter(c_remediation_path, engine='openpyxl', mode='a')
    book = load_workbook(c_remediation_path)
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    other_inventory_combined_df.to_excel(writer, sheet_name='Other Inventory', index=False, startrow=6, startcol=1, header=False)
    writer.save()
    print('c_remediation other inventory added')

    print("Sheet : Other Inventory : CREATED")
    print("End 17 : Other inventory\n")
    # redis_client.rpush(redis_queue_name, '38/40::Other Inventory check is complete')

    middleware_type = 'Tuxedo'
    if(middleware_type != 'Tuxedo'):
        filename = c_remediation_path
        sheet_name = '6. Middleware-Tuxedo'
        workbook = load_workbook(filename)

        if sheet_name in workbook.sheetnames:
            sheet_to_delete = workbook[sheet_name]
            workbook.remove(sheet_to_delete)
            workbook.save(filename)
        print('Tuxedo sheet removed !!!')

    print('Migration Summary')
    wb = load_workbook(c_remediation_path)
    ws = wb['Migration Summary']
    start_col_1 = 15
    start_col_2 = 16
    start_row_1 = 28
    start_row_2 = 29
    if O28 != 0:
        ws.cell(row=start_row_1, column=start_col_1, value=O28)
    if O29 != 0:
        ws.cell(row=start_row_2, column=start_col_1, value=O29)
    if P28 != 0:
        ws.cell(row=start_row_1, column=start_col_2, value=P28)
    if P29 != 0:
        ws.cell(row=start_row_2, column=start_col_2, value=P29)

    O7 = P28 + P29
    F7 = O28 + O29

    if O7 != 0:
        ws.cell(row=7, column=15, value=O7)
    if O6 != 0:
        ws.cell(row=6, column=15, value=O6)
    if F7 != 0:
        ws.cell(row=7, column=6, value=F7)
    if F6 != 0:
        ws.cell(row=6, column=6, value=F6)

    wb.save(c_remediation_path)
    # redis_client.rpush(redis_queue_name, '39/40::Reports are ready')
    # redis_client.rpush(redis_queue_name, '40/40::CODE ASSESSMENT SUCCESSFUL')
