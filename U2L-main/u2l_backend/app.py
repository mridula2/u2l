from flask import Flask, request, jsonify, Response
import datetime
import os
import subprocess
import shutil
import pandas as pd
from openpyxl import load_workbook
import re
import numpy as np
import base64
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
from sqlalchemy.orm import class_mapper
import zipfile
import openpyxl

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'
CORS(app)

tool_analysis_type = ''
PJHOME = ''
APNAME = ''
file_name = ''
yMD = ''
hMS = ''
hM = ''
db_username = ''
db_password_encrypt = ''
user_dict = {}

#db connection
userpass = "mysql+pymysql://mysql:Password#123@"
basedir = "u2l_db:3306"
dbname = "/u2l_db"

app.config["SQLALCHEMY_DATABASE_URI"] = userpass + basedir + dbname
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
app.app_context().push()

class user_details(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    user_name = db.Column(db.String(50), nullable=False)
    password = db.Column(db.String(50), nullable=False)
    user_role = db.Column(db.String(50))
    created_at = db.Column(db.DateTime)
    updated_at = db.Column(db.DateTime)

    def __init__(self, user_name, password, user_role, created_at,updated_at):
        self.user_name = user_name
        self.password = password
        self.user_role = user_role
        self.created_at = created_at
        self.updated_at = updated_at

class project_details(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    user_name = db.Column(db.String(50))
    user_id = db.Column(db.Integer, db.ForeignKey('user_details.id'))
    project_client = db.Column(db.String(50))
    project_manager = db.Column(db.String(50))
    file_name = db.Column(db.String(50))
    file_size = db.Column(db.Integer)
    analysis_status = db.Column(db.String(50))
    created_at = db.Column(db.DateTime)
    user_details = db.relationship('user_details', backref=db.backref('project_details', lazy=True))

    def __init__(self, project_name, user_name, user_id, project_client, project_manager, file_name, file_size, analysis_status, created_at):
        self.project_name = project_name
        self.user_name = user_name
        self.user_id = user_id
        self.project_client = project_client
        self.project_manager = project_manager
        self.file_name = file_name
        self.file_size = file_size
        self.analysis_status = analysis_status
        self.created_at = created_at

class os_details(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    source_os = db.Column(db.String(50))
    source_os_version = db.Column(db.String(50))
    target_os = db.Column(db.String(50))
    target_os_version = db.Column(db.String(50))
    
    def __init__(self, project_name, source_os, source_os_version, target_os, target_os_version):
        self.project_name = project_name
        self.source_os = source_os
        self.source_os_version = source_os_version
        self.target_os = target_os
        self.target_os_version = target_os_version

class analysis_type(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    analysis_type = db.Column(db.String(50))
    
    def __init__(self, project_name, analysis_type):
        self.project_name = project_name
        self.analysis_type = analysis_type

class analysis_java(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    analysis_id = db.Column(db.Integer, db.ForeignKey('analysis_type.id'))
    source_jdk = db.Column(db.String(50))
    target_jdk = db.Column(db.String(50))
    source_jsp = db.Column(db.String(50))
    target_jsp = db.Column(db.String(50))
    source_servlet = db.Column(db.String(50))
    target_servlet = db.Column(db.String(50))
    analysis_type = db.relationship('analysis_type', backref=db.backref('analysis_java', lazy=True))
    

    def __init__(self, project_name, analysis_id, source_jdk, target_jdk, source_jsp, target_jsp, source_servlet, target_servlet):
        self.project_name = project_name
        self.analysis_id = analysis_id
        self.source_jdk = source_jdk
        self.target_jdk = target_jdk
        self.source_jsp = source_jsp
        self.target_jsp = target_jsp
        self.source_servlet = source_servlet
        self.target_servlet = target_servlet

class analysis_c(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    analysis_id = db.Column(db.Integer, db.ForeignKey('analysis_type.id'))
    source_compiler = db.Column(db.String(50))
    target_compiler = db.Column(db.String(50))
    source_compiler_version = db.Column(db.String(50))
    target_compiler_version = db.Column(db.String(50))
    source_oracle_version = db.Column(db.String(50))
    target_oracle_version = db.Column(db.String(50))
    analysis_type = db.relationship('analysis_type', backref=db.backref('analysis_c', lazy=True))
    

    def __init__(self, project_name, analysis_id, source_compiler, target_compiler, source_compiler_version, target_compiler_version, source_oracle_version, target_oracle_version):
        self.project_name = project_name
        self.analysis_id = analysis_id
        self.source_compiler = source_compiler
        self.target_compiler = target_compiler
        self.source_compiler_version = source_compiler_version
        self.target_compiler_version = target_compiler_version
        self.source_oracle_version = source_oracle_version
        self.target_oracle_version = target_oracle_version

class analysis_shell(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    analysis_id = db.Column(db.Integer, db.ForeignKey('analysis_type.id'))
    source_shell = db.Column(db.String(50))
    target_shell = db.Column(db.String(50))
    source_shell_version = db.Column(db.String(50))
    target_shell_version = db.Column(db.String(50))
    analysis_type = db.relationship('analysis_type', backref=db.backref('analysis_shell', lazy=True))

    def __init__(self, project_name, analysis_id, source_shell, target_shell, source_shell_version, target_shell_version):
        self.project_name = project_name
        self.analysis_id = analysis_id
        self.source_shell = source_shell
        self.target_shell = target_shell
        self.source_shell_version = source_shell_version
        self.target_shell_version = target_shell_version

class analysis_status(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    project_name = db.Column(db.String(50), unique=True, nullable=False)
    analysis_id = db.Column(db.String(50))
    user_name = db.Column(db.String(50))
    analysis_status = db.Column(db.String(50))
    file_name = db.Column(db.String(50))
    file_size = db.Column(db.Integer)
    analysis_start_time = db.Column(db.DateTime)
    analysis_end_time = db.Column(db.DateTime)

    def __init__(self, project_name, analysis_id, user_name, analysis_status, file_name, file_size, analysis_start_time, analysis_end_time):
        self.project_name = project_name
        self.analysis_id = analysis_id
        self.user_name = user_name
        self.analysis_status = analysis_status
        self.file_name = file_name
        self.file_size = file_size
        self.analysis_start_time = analysis_start_time
        self.analysis_end_time = analysis_end_time                          

db.create_all()

# db_username = 'admin'
# db_password = 'Password@123'
# db_password_encrypt = base64.b64encode(db_password.encode()).decode()
# db_user_role = 'admin'
# db_created_at = datetime.datetime.now()
# db_updated_at = datetime.datetime.now()
# user_data = user_details(db_username, db_password_encrypt, db_user_role, db_created_at, db_updated_at)
# db.session.add(user_data)
# db.session.commit()


@app.route('/login', methods=['POST'])
def authentication():

    if request.method == 'POST':
        json_data = request.get_json()
        form_username = json_data['user_name']
        form_password = json_data['password']

        user_data_validate = user_details.query.filter_by(user_name=form_username).first()

        if user_data_validate:
            if user_data_validate.password == form_password:
                user = user_details.query.filter_by(user_name=form_username).first()
                if user and user.password == form_password:
                    global user_dict
                    user_dict = {
                        'id': user.id,
                        'user_name': user.user_name,
                        'password': user.password,
                        'user_role': user.user_role,
                        'created_at': user.created_at,
                        'updated_at': user.updated_at
                    }
                return jsonify({'message': 'authentication success'}), 200
            else:
                return jsonify({'message': 'authentication failed'}), 401
        else:
            return jsonify({'message': 'authentication failed'}), 401

@app.route('/analysis', methods=['POST'])
def upload():
    try:

        file = request.files['file_name']
        # file_size = os.path.getsize(file.filename)
        # print(file_size)
        form_project_name = request.form['project_name']
        form_project_client = request.form['project_client']
        form_project_manager = request.form['project_manager']
        global tool_analysis_type
        tool_analysis_type = request.form['analysis_type']

        form_source_os = request.form['source_os']
        form_source_os_version = request.form['source_os_version']
        form_target_os = request.form['target_os']
        form_target_os_version = request.form['target_os_version']


        if(tool_analysis_type == 'Java'):
            tool_analysis_type = 'javaanalysis'
            form_source_jdk = request.form['source_jdk']
            form_target_jdk = request.form['target_jdk']
            form_source_jsp = request.form['source_jsp']
            form_target_jsp = request.form['target_jsp']
            form_source_servlet = request.form['source_servlet']
            form_target_servlet = request.form['target_servlet']
        elif(tool_analysis_type == 'C/C++/Pro*C'):
            tool_analysis_type = 'canalysis'
            form_source_compiler = request.form['source_compiler']
            form_target_compiler = request.form['target_compiler']
            form_source_compiler_version = request.form['source_compiler_version']
            form_target_compiler_version = request.form['target_compiler_version']
            form_source_oracle_version = request.form['source_oracle_version']
            form_target_oracle_version = request.form['target_oracle_version']
        elif(tool_analysis_type == 'Shell'):  
            tool_analysis_type = 'shellanalysis'
            form_source_shell = request.form['source_shell']
            form_target_shell = request.form['target_shell']
            form_source_shell_version = request.form['source_shell_version']
            form_target_shell_version = request.form['target_shell_version']


        analysis_types = ['javaanalysis', 'canalysis', 'shellanalysis']
        if tool_analysis_type not in analysis_types:
            error_message = {'error' :"Invalid analysis type"}
            return jsonify(error_message), 404

        cur_date = datetime.datetime.now()
    
        global file_name
        file_name = file.filename
        temp_tuple = os.path.splitext(file_name)
        if temp_tuple[1] != '.zip':
            error_message = {'error' :"Upload .zip files only"}
            return jsonify(error_message), 404
        file_name = temp_tuple[0]

        year_month_day = str(cur_date).split(" ")
        global yMD
        yMD = year_month_day[0].replace("-", "")

        hour_min_sec = os.path.splitext(year_month_day[1])
        global hMS
        hMS = hour_min_sec[0].replace(":", "")

        file_name = file_name + "_" + yMD + "_" + hMS 
        directory = './projects/' + file_name
        os.makedirs(directory)
        file.save(os.path.join(directory, file_name +'.zip'))

        try:
            source_dir = '../'
            dest_dir = directory
            source_path = os.path.join(source_dir, 'U2L')
            dest_path = os.path.join(dest_dir, 'U2L')
            shutil.copytree(source_path, dest_path)  
            
            if(tool_analysis_type == 'javaanalysis'):
                source_path = os.path.join(source_dir, 'Java')
                dest_path = os.path.join(dest_dir, 'Java')
                shutil.copytree(source_path, dest_path)

                source_path = os.path.join(source_dir, 'java_report.xlsx')
                dest_path = os.path.join(dest_dir, 'java_report.xlsx')
                shutil.copy(source_path, dest_path)

                source_path = os.path.join(source_dir, 'java_inventory_report.xlsx')
                dest_path = os.path.join(dest_dir, 'java_inventory_report.xlsx')
                shutil.copy(source_path, dest_path)

            elif(tool_analysis_type == 'shellanalysis'):
                source_path = os.path.join(source_dir, 'shell_report.xlsx')
                dest_path = os.path.join(dest_dir, 'shell_report.xlsx')
                shutil.copy(source_path, dest_path)

                source_path = os.path.join(source_dir, 'shell_inventory_report.xlsx')
                dest_path = os.path.join(dest_dir, 'shell_inventory_report.xlsx')
                shutil.copy(source_path, dest_path)

            elif(tool_analysis_type == 'canalysis'):
                source_path = os.path.join(source_dir, 'c_report.xlsx')
                dest_path = os.path.join(dest_dir, 'c_report.xlsx')
                shutil.copy(source_path, dest_path)

                source_path = os.path.join(source_dir, 'c_inventory_report.xlsx')
                dest_path = os.path.join(dest_dir, 'c_inventory_report.xlsx')
                shutil.copy(source_path, dest_path)
        
        except Exception as e:
            response = jsonify({'error': str(e)})
            response.status_code = 500
            return response 

        #storing values in db project_details table
        db_project_details = project_details(form_project_name, 'admin', 1, form_project_client, form_project_manager, file_name, 23, 'analysis started', cur_date)
        db.session.add(db_project_details)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'Project name already exists'}), 409
            # return jsonify({'message': 'project_details : Project name already exists'}), 409

        #storing values in db analysis_type table
        db_analysis_type = analysis_type(form_project_name, tool_analysis_type)
        db.session.add(db_analysis_type)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'analysis_type : Project name already exists'}), 409

        #storing values in db os_details table
        db_os_details = os_details(form_project_name, form_source_os, form_source_os_version, form_target_os, form_target_os_version)
        db.session.add(db_os_details)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'os_details : Project name already exists'}), 409
        
        #storing values in db analysis_status table
        analysis_start_time = datetime.datetime.now()
        analysis_end_time = datetime.datetime.now()
        db_analysis_status = analysis_status(form_project_name, db_analysis_type.id, 'admin', 'analysis started', file_name, 15, analysis_start_time, analysis_end_time)
        db.session.add(db_analysis_status)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'message': 'analysis_status : Project name already exists'}), 409

        # Execute U2LTool_Install.sh Script
        script_path = './projects/' + file_name + '/U2L/U2LTool_Install.sh'
    
        if(db_analysis_type.analysis_type == 'javaanalysis'):
            db_analysis_java = analysis_java(form_project_name, db_analysis_type.id, form_source_jdk, form_target_jdk, form_source_jsp, form_target_jsp, form_source_servlet, form_target_servlet)
            db.session.add(db_analysis_java)
            db.session.commit()
            file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_' + yMD + '_' + hMS
        elif(db_analysis_type.analysis_type == 'canalysis'):
            db_analysis_c = analysis_c(form_project_name, db_analysis_type.id, form_source_compiler, form_target_compiler, form_source_compiler_version, form_target_compiler_version, form_source_oracle_version, form_target_oracle_version)
            db.session.add(db_analysis_c)
            db.session.commit()
            file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS
        elif(db_analysis_type.analysis_type == 'shellanalysis'):
            db_analysis_shell = analysis_shell(form_project_name, db_analysis_type.id, form_source_shell, form_target_shell, form_source_shell_version, form_target_shell_version)
            db.session.add(db_analysis_shell)
            db.session.commit()
            file_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS

        try:
            command = ['sh', script_path, file_path, tool_analysis_type]
            process = subprocess.Popen(command, stdout=subprocess.PIPE)
            output = process.communicate()[0].decode().strip()
            lines = output.split('\n')
            print(lines)
            print('\nTool Installation Completed !!') 
        except Exception as e:
            return jsonify({'error': str(e)})

        global PJHOME
        PJHOME = lines[5].split(':')[1]
        global APNAME
        APNAME = lines[6].split(':')[1]
        
        # Export 'PJHOME' & 'APNAME' variables
        os.environ['PJHOME'] = PJHOME
        os.environ['APNAME'] = APNAME

        # Copy source code to java/c/shell analysis directory
        source_path = '/usr/u2l/u2l_backend/projects/' + file_name +'/'+ file_name + '.zip'

        try:
            if(tool_analysis_type == 'javaanalysis'):
                destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
            elif(tool_analysis_type == 'canalysis'):
                destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_'+ yMD + '_' + hMS +'/work/canalysis/'
            elif(tool_analysis_type == 'shellanalysis'):
                destination_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_'+ yMD + '_' + hMS +'/work/shellanalysis/'
            shutil.unpack_archive(source_path, destination_path, 'zip')
            
        except Exception as e:
            response = jsonify({'error': str(e)})
            response.status_code = 500
            return response

        # Execute U2LTool_Analysis.sh Script for c/shell analysis
        try:
            if(tool_analysis_type == 'canalysis' or tool_analysis_type == 'shellanalysis'):
                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/U2LTool_Analysis.sh'
                command = ['sh', script_path, PJHOME, APNAME]
                process = subprocess.Popen(command, stdout=subprocess.PIPE)
                output = process.communicate()[0].decode().strip()
                lines = output.split('\n')
                print(lines)
                print(output)
                print('\nCompleted running U2LTool_Analysis.sh')
        
        except Exception as e:
            return jsonify({'error': str(e)})
        
        # Execute U2LTool_Analysis.sh Script for java analysis
        if(tool_analysis_type == 'javaanalysis'):
            try:

                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/U2LTool_Analysis.sh'

                command = ['sh', script_path, PJHOME, APNAME]
                process = subprocess.Popen(command, stdout=subprocess.PIPE)
                process.wait()
                output = process.communicate()[0].decode().strip()

                print('\nCompleted running U2LTool_Analysis.sh')
            
            except Exception as e:
                return jsonify({'error': str(e)})

            # To run the javaanalysis scripts
            # Export 'PJHOME' variable
            PJHOME = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java'
            os.environ['PJHOME'] = PJHOME

            # Run javaSourceDiscovery.sh scriptPassword@123

            try:
                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaSourceDiscovery.sh'
                code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
                command = ['sh', script_path, code_path]
                process = subprocess.Popen(command, stdout=subprocess.PIPE)
                output = process.communicate()[0].decode().strip()
                lines = output.split('\n')
                print(lines)
                print('\nCompleted running javaSourceDiscovery.sh')
            except Exception as e:
                return jsonify({'error': str(e)})    

            # Run javaRulesScan.sh script
            try:
                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaRulesScan.sh'
                code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
                command = ['sh', script_path, code_path]
                process = subprocess.Popen(command, stdout=subprocess.PIPE)
                output = process.communicate()[0].decode().strip()
                lines = output.split('\n')
                print(lines)
                print('\nCompleted running javaRulesScan.sh')
            except Exception as e:
                return jsonify({'error': str(e)})   

            # Run javaFrameworkScan.sh script
            try:
                script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaFrameworkScan.sh'
                code_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
                framework_type = 'Spring'
                command = ['sh', script_path, code_path, framework_type]
                process = subprocess.Popen(command, stdout=subprocess.PIPE)
                output = process.communicate()[0].decode().strip()
                lines = output.split('\n')
                print(lines)
                print('\nCompleted running javaFrameworkScan.sh\n')
            except Exception as e:
                return jsonify({'error': str(e)}) 

            # Run javaRulesRemedy.sh script
            # try:
            #     script_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/Java/javaRulesRemedy.sh'
            #     code_path = '/usr/u2l/u2l_backend/projects/' + file_name + 'U2L/java-vfunction_'+ yMD + '_' + hMS +'/work/javaanalysis/'
            #     command = subprocess.run([script_path, code_path], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            #     output = command.communicate()[0].decode().strip()
            #     lines = output.split('\n')
            #     print(lines)
            #     print('Completed running javaRulesRemedy.sh')
            # except Exception as e:
            #     return jsonify({'error': str(e)})

        # Converting logs into dataframes and generating reports
        try:

            if(tool_analysis_type == 'javaanalysis'):

                yyyy_mmdd = yMD[:4] + '-' + yMD[4:]
                
                # Convert javacIssue.log files into .csv
                javacIssue_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/javacLog/' + yyyy_mmdd + '/javacIssue.log'
                if os.path.isfile(javacIssue_dir_path) and os.path.getsize(javacIssue_dir_path) > 0:
                    df = pd.read_csv(javacIssue_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['FILE NAME', 'LINE NUMBER', 'COMMAND', 'ERROR MESSAGE']
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Compilation Error Report', index=True)
                    writer.save()

                # Convert javacWarnings.list files into .csv
                javacWarnings_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/javacLog/' + yyyy_mmdd + '/javacWarnings.list'
                if os.path.isfile(javacWarnings_dir_path) and os.path.getsize(javacWarnings_dir_path) > 0:
                    df = pd.read_csv(javacWarnings_dir_path, sep=':', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Compilation Warning Report', index=True)
                    writer.save()

                # Convert import_lines.out
                importLines_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines.out'
                if os.path.isfile(importLines_dir_path) and os.path.getsize(importLines_dir_path) > 0:
                    df = pd.read_csv(importLines_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['IMPORTS', 'FILE NAME', 'LINE NUMBER']
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Import Class Report', index=True)
                    writer.save()

                # Convert import_lines_jsp.out
                importLinesJsp_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/import/' + yyyy_mmdd + '/import_lines_jsp.out'
                if os.path.isfile(importLinesJsp_dir_path) and os.path.getsize(importLinesJsp_dir_path) > 0:
                    df = pd.read_csv(importLinesJsp_dir_path, sep='\t', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Import JSP Report', index=True)
                    writer.save()

                # Convert OS_diff.out
                OSDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/OS_diff.out'
                if os.path.isfile(OSDiff_dir_path) and os.path.getsize(OSDiff_dir_path) > 0:
                    df = pd.read_csv(OSDiff_dir_path, sep='\s+', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                    df = df.drop(columns=['Column1'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='OS Analysis Details', index=True)
                    writer.save()

                # Convert jdk_diff.out
                jdkDiff_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/java/versiondiff/' + yyyy_mmdd + '/jdk_diff.out'
                if os.path.isfile(jdkDiff_dir_path) and os.path.getsize(jdkDiff_dir_path) > 0:
                    df = pd.read_csv(jdkDiff_dir_path, sep='\s+', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df = df.drop(columns=['Column1'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='JDK Anaylsis Details', index=True)
                    writer.save()
                
                #Source Code Inventory
                regex = r'^.*$'
                source_code_java_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/java-vfunction_'+ yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
                print(source_code_java_path)
                if os.path.isfile(source_code_java_path) and os.path.getsize(source_code_java_path) > 0:
                    
                    with open(source_code_java_path, 'r') as f:
                        data = f.read()
                    lines = data.split('\n')
                    log_entries = []
                    for line in lines:
                        match = re.match(regex, line)
                        if match:
                            log_entries.append(match.group())
                    df = pd.DataFrame(log_entries, columns=['log_message'])

                    start_str = '###########  Execution 0001_stepcounter Starts  ###########'
                    end_str = '###########  Execution 0001_stepcounter Ends ###########'
                    start_df = df[df['log_message'].str.contains(start_str)]
                    end_df = df[df['log_message'].str.contains(end_str)]

                    start_index = start_df.index[0]
                    end_index = end_df.index[0]
                    extracted_df = df.iloc[start_index+1:end_index]

                    result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
                    result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
                    result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
                    result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
                    result_df = result_df.drop(['log_message', 'col2'], axis=1)
                    result_df = result_df.reset_index(drop=True)
                    result_df.index += 1

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/java_inventory_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

            elif(tool_analysis_type == 'canalysis'):

                yyyy_mmdd = yMD[:4] + '-' + yMD[4:]

                # Convert stk-YYYY-MMDD-detail.log
                stk_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/stk/' + yyyy_mmdd + '/stk-' + yyyy_mmdd + '-detail.log'
                if os.path.isfile(stk_dir_path) and os.path.getsize(stk_dir_path) > 0:
                    df = pd.read_csv(stk_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='STK Analysis Detail', index=True)
                    writer.save()

                # Convert 64bit_C_reportlog
                bitC_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/checkanalysis/64bit_C_reportlog'
                if os.path.isfile(bitC_dir_path) and os.path.getsize(bitC_dir_path) > 0:
                    df = pd.read_csv(bitC_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE', 'Column4', 'Column5']
                    df = df.drop(columns=['Column4', 'Column5'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='64-bit Analysis Detail', index=True)
                    writer.save()

                # Convert endian_C_reportlog
                endian_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/checkanalysis/endian_C_reportlog'
                if os.path.isfile(endian_dir_path) and os.path.getsize(endian_dir_path) > 0:
                    df = pd.read_csv(endian_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Endianness Analysis Detail', index=True)
                    writer.save()

                # Convert char-array.out
                charArray_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/charinv/' + yyyy_mmdd + '/char-array.out'
                if os.path.isfile(charArray_dir_path) and os.path.getsize(charArray_dir_path) > 0:
                    df = pd.read_csv(charArray_dir_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='String-Mem Analysis Detail', index=True)
                    writer.save()

                # Convert Log_Envcheck_CFiles_YYYY-MMDD.log
                envCheck_dir_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Envcheck_CFiles_' + yyyy_mmdd + '.log'
                if os.path.isfile(envCheck_dir_path) and os.path.getsize(envCheck_dir_path) > 0:
                    df = pd.read_csv(envCheck_dir_path, sep=':', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['FILE PATH', 'LINE NUMBER', 'SOURCE CODE']
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='ExecutionEnv Analysis Detail', index=True)
                    writer.save()

                #Source Code Inventory
                regex = r'^.*$'
                source_code_c_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/c-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
                if os.path.isfile(source_code_c_path) and os.path.getsize(source_code_c_path) > 0:
                    with open(source_code_c_path, 'r') as f:
                        data = f.read()
                    lines = data.split('\n')
                    log_entries = []
                    for line in lines:
                        match = re.match(regex, line)
                        if match:
                            log_entries.append(match.group())
                    df = pd.DataFrame(log_entries, columns=['log_message'])

                    start_str = '###########  Execution 0001_stepcounter Starts  ###########'
                    end_str = '###########  Execution 0001_stepcounter Ends ###########'
                    start_df = df[df['log_message'].str.contains(start_str)]
                    end_df = df[df['log_message'].str.contains(end_str)]

                    start_index = start_df.index[0]
                    end_index = end_df.index[0]
                    extracted_df = df.iloc[start_index+1:end_index]

                    result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
                    result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
                    result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
                    result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
                    result_df = result_df.drop(['log_message', 'col2'], axis=1)
                    result_df = result_df.reset_index(drop=True)
                    result_df.index += 1

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/c_inventory_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

            elif(tool_analysis_type == 'shellanalysis'):
            
                yyyy_mmdd = yMD[:4] + '-' + yMD[4:]

                # Convert ksh_env_common.txt
                ksh_env_common_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/env/' + yyyy_mmdd + '/ksh_env_common.txt'
                if os.path.isfile(ksh_env_common_path) and os.path.getsize(ksh_env_common_path) > 0:
                    df = pd.read_csv(ksh_env_common_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                    df = df.drop(columns=['Column1'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='OS Specific ENV Var Details', index=True)
                    writer.save()

                # Convert ksh_filepath_common.txt
                ksh_filepath_common_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/env/' + yyyy_mmdd + '/ksh_filepath_common.txt'
                if os.path.isfile(ksh_filepath_common_path) and os.path.getsize(ksh_filepath_common_path) > 0:
                    df = pd.read_csv(ksh_filepath_common_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                    df = df.drop(columns=['Column1'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='OS Specific File Path Details', index=True)
                    writer.save()

                # Convert ksh_parsed_uniq.out
                ksh_parsed_uniq_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/parse/' + yyyy_mmdd + '/ksh_parsed_uniq.out'
                if os.path.isfile(ksh_parsed_uniq_path) and os.path.getsize(ksh_parsed_uniq_path) > 0:
                    df = pd.read_csv(ksh_parsed_uniq_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['Column1', 'Column2', 'COMMAND', 'Column4', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                    df = df.drop(columns=['Column1', 'Column2', 'Column4'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Command and option Details', index=True)
                    writer.save()

                # Convert ksh_shebang.txt
                ksh_shebang_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/shebang/' + yyyy_mmdd + '/ksh_shebang.txt'
                if os.path.isfile(ksh_shebang_path) and os.path.getsize(ksh_shebang_path) > 0:
                    with open(ksh_shebang_path, 'r') as f:
                        lines = f.readlines()
                    data = []
                    prev_column1 = None
                    for line in lines:
                        split_line = line.strip().split('#!')
                        if split_line[0].strip():
                            prev_column1 = split_line[0].strip()
                        data.append([prev_column1, '#!' + split_line[1].strip() if len(split_line) > 1 else ''])
                    df = pd.DataFrame(data, columns=['column1', 'column2']).dropna()
                    df['column2'] = df.groupby(df.index // 2)['column2'].transform(lambda x: x.iloc[1])
                    df.columns = ['FILE NAME', 'SOURCE CODE']
                    df.index = df.index + 1
                    df = df.drop(df[df.index % 2 == 0].index).reset_index(drop=True)

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'
                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl')
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Shebang Line Analysis', index=True)
                    writer.save()

                # Convert ksh_heredocument.txt
                ksh_heredocument_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/shell/syntax/' + yyyy_mmdd + '/ksh_heredocument.txt'
                if os.path.isfile(ksh_heredocument_path) and os.path.getsize(ksh_heredocument_path) > 0:
                    df = pd.read_csv(ksh_heredocument_path, sep='\t', engine='python', header=None, index_col=None, error_bad_lines=False, warn_bad_lines=True)
                    df.columns = ['Column1', 'FILE NAME', 'LINE NUMBER', 'SOURCE CODE']
                    df = df.drop(columns=['Column1'])
                    df.index = df.index + 1
                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    df.to_excel(writer, sheet_name='Heredoc Analysis', index=True)
                    writer.save()

                # Source Code Inventory
                regex = r'^.*$'
                source_code_shell_path = '/usr/u2l/u2l_backend/projects/' + file_name + '/U2L/shell-vfunction_' + yMD + '_' + hMS + '/log/' + tool_analysis_type + '/Log_Source_Analysis_' + yMD + '.log'
                if os.path.isfile(source_code_shell_path) and os.path.getsize(source_code_shell_path) > 0:
                    with open(source_code_shell_path, 'r') as f:
                        data = f.read()
                    lines = data.split('\n')
                    log_entries = []
                    for line in lines:
                        match = re.match(regex, line)
                        if match:
                            log_entries.append(match.group())
                    df = pd.DataFrame(log_entries, columns=['log_message'])

                    start_str = '###########  Execution 0001_stepcounter Starts  ###########'
                    end_str = '###########  Execution 0001_stepcounter Ends ###########'
                    start_df = df[df['log_message'].str.contains(start_str)]
                    end_df = df[df['log_message'].str.contains(end_str)]

                    start_index = start_df.index[0]
                    end_index = end_df.index[0]
                    extracted_df = df.iloc[start_index+1:end_index]

                    result_df = extracted_df[extracted_df['log_message'].str.startswith('./')]
                    result_df[['Directory', 'Type', 'col2', 'Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']] = result_df['log_message'].str.split(',', expand=True)
                    result_df = result_df.replace(r'^\s*$', np.nan, regex=True)
                    result_df = result_df.dropna(subset=['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC'])
                    result_df = result_df.drop(['log_message', 'col2'], axis=1)
                    result_df = result_df.reset_index(drop=True)
                    result_df.index += 1

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

                    save_location = '/usr/u2l/u2l_backend/projects/' + file_name + '/shell_inventory_report.xlsx'

                    book = load_workbook(save_location)
                    writer = pd.ExcelWriter(save_location, engine='openpyxl') 
                    writer.book = book
                    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                    result_df.to_excel(writer, sheet_name='Source Code Inventory', index=True)
                    writer.save()

        except Exception as e:
            return jsonify({'error': str(e)})
    
    except Exception as e:
            analysis_end_time = datetime.datetime.now()
            if db_project_details is not None:
                db_project_details.analysis_status = 'analysis failed'
                db.session.commit()
            if db_analysis_status is not None:
                db_analysis_status.analysis_status = 'analysis failed'
                db_analysis_status.analysis_start_time = analysis_start_time
                db_analysis_status.analysis_end_time = analysis_end_time
                db.session.commit()
            return jsonify({'message': 'analysis failed'}), 401
    
    analysis_end_time = datetime.datetime.now()
    if db_project_details is not None:
        db_project_details.analysis_status = 'analysis successful'
        db.session.commit()
    if db_analysis_status is not None:
        db_analysis_status.analysis_status = 'analysis successful'
        db_analysis_status.analysis_start_time = analysis_start_time
        db_analysis_status.analysis_end_time = analysis_end_time
        db.session.commit()
    return jsonify({'message': 'analysis successful'}), 200

@app.route('/projects/<query_user_name>', methods=['GET']) #{user_name}
def projects(query_user_name):
    
    query = project_details.query.filter_by(user_name=query_user_name)
    json_project_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_project_details.append(row_dict)

    return jsonify({'project_details': json_project_details}), 200

@app.route('/report/<query_project_name>', methods=['GET']) #{project_name} excel reports zip
def report(query_project_name):

    query = project_details.query.filter_by(project_name=query_project_name, user_name='admin')
    print(query.first().file_name)
    fileName = query.first().file_name

    query = analysis_type.query.filter_by(project_name=query_project_name)
    analysisType = query.first().analysis_type
    print(analysisType)

    if analysisType == 'javaanalysis':      
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_inventory_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_report.xlsx'
    elif analysisType == 'canalysis':
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_inventory_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_report.xlsx'
    elif analysisType == 'shellalysis':
        file1_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_inventory_report.xlsx'
        file2_path = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_report.xlsx'

    if not os.path.exists(file1_path):
        return 'File 1 not found', 404
    if not os.path.exists(file2_path):
        return 'File 2 not found', 404
    
    zip_buffer = zipfile.ZipFile('files.zip', 'w', zipfile.ZIP_DEFLATED)
    zip_buffer.write(file1_path, os.path.basename(file1_path))
    zip_buffer.write(file2_path, os.path.basename(file2_path))
    zip_buffer.close()

    os.chmod('files.zip', 0o777)

    return Response(
        open('files.zip', 'rb').read(),
        # mimetype='application/zip',
        headers={
            'Content-Disposition': 'attachment;filename=files.zip',
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/zip',
            'Cache-Control': 'no-cache, no-store, must-revalidate'
        }
    )

@app.route('/pdf/<query_project_name>', methods=['GET']) #{project_name} pdf generation
def pdf(query_project_name):

    query = project_details.query.filter_by(project_name=query_project_name)
    json_project_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_project_details.append(row_dict)

    fileName = query.first().file_name

    query = os_details.query.filter_by(project_name=query_project_name)
    json_os_details = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_os_details.append(row_dict)

    query = analysis_type.query.filter_by(project_name=query_project_name)
    json_analysis_type = []
    for row in query.all():
        row_dict = {}
        for column in row.__table__.columns:
            row_dict[column.name] = str(getattr(row, column.name))
        json_analysis_type.append(row_dict)

    analysisType = query.first().analysis_type
    print(analysisType)
    sheetName = 'Source Code Inventory'
    selected_cols = ['Actual Nr of Lines', 'Nr Blank Lines', 'Nr Commented Lines', 'Total Nr LoC']

    if analysisType == 'javaanalysis':      
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_inventory_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/java_report.xlsx'

        sheet_names = ['OS Analysis Details', 'Compilation Error Report', 'Import Class Report', 'JDK Anaylsis Details', 'Import JSP Report', 'Compilation Warning Report']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)
        sheets = [wb[sheet_name] for sheet_name in sheet_names]
        count_of_ids = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1

    elif analysisType == 'canalysis':
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_inventory_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/c_report.xlsx'

        sheet_names = ['STK Analysis Detail', '64-bit Analysis Detail', 'Endianness Analysis Detail', 'String-Mem Analysis Detail', 'ExecutionEnv Analysis Detail']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)
        sheets = [wb[sheet_name] for sheet_name in sheet_names]
        count_of_ids = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1

    elif analysisType == 'shellanalysis':
        filepath = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_inventory_report.xlsx'
        filepath_percentage = '/usr/u2l/u2l_backend/projects/'+ fileName +'/shell_report.xlsx'

        sheet_names = ['OS Specific ENV Var Details', 'OS Specific File Path Details', 'Command and option Details', 'Shebang Line Analysis', 'Heredoc Analysis']
        wb = load_workbook(filename=filepath_percentage, read_only=True, data_only=True)
        sheets = [wb[sheet_name] for sheet_name in sheet_names]
        count_of_ids = 0
        for sheet in sheets:
            if sheet.max_row < 2:
                continue
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] is not None:
                    count_of_ids += 1

    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    df = df.apply(pd.to_numeric, errors='coerce')
    column_sums = df.sum()


    sheetName = 'Source Code Inventory'
    selected_cols = ['Type']
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheetName]
    data = ws.values
    cols = next(data)[0:]
    df = pd.DataFrame(data, columns=cols)[selected_cols]
    type_counts = df['Type'].value_counts().reset_index()
    counts = type_counts.rename(columns={'index': 'Type', 'Type': 'number'})
    counts_list = counts.to_dict('records')

    total_no_of_lines = column_sums['Total Nr LoC']
    percentage = (count_of_ids/total_no_of_lines) * 100
    rounded_percentage = round(percentage, 2)

    return jsonify({'project_details': json_project_details, 'os_details': json_os_details, 'analysis_type': json_analysis_type, 'chart1': column_sums.to_dict(), 'chart2' : counts_list, 'percent' : rounded_percentage}), 200


if __name__ == '__main__':
    app.run(debug=True)
